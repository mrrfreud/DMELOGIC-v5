"""
Nova UI Server
==============
Serves the Nova web chat interface and handles WebSocket communication.
Runs on port 8401. Open http://localhost:8401 in Chrome.

The terminal windows run hidden in the background.
This is the only window you need to see.

Usage (handled by start_nova.bat):
    python nova_ui_server.py
"""

from __future__ import annotations
import os, sys, json, asyncio, logging, base64, re, sqlite3, subprocess, time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from datetime import datetime, timedelta
import requests

# Add project root to path
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

try:
    from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request
    from fastapi.responses import HTMLResponse, FileResponse
    from fastapi.middleware.cors import CORSMiddleware
    import uvicorn
except ImportError:
    sys.exit("Run: pip install fastapi uvicorn")

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from dmelogic import nova_phone
log = logging.getLogger("nova_ui")

ENV_PATH = Path(_HERE) / ".env"
NOVA_ICON_PNG = Path(_HERE) / "assets" / "nova_icon.png"
NOVA_ICON_ICO = Path(_HERE) / "assets" / "Nova Icon.ico"
WEBPHONE_JS = Path(_HERE) / "static" / "vendor" / "ringcentral-web-phone.min.js"


def _env_or_default(name: str, default: str) -> str:
  value = str(os.getenv(name, "") or "").strip()
  placeholders = {
    "your-key",
    "your-key-here",
    "your-strong-key",
    "your-strong-key-here",
    "changeme",
    "change-me",
  }
  if not value or value.lower() in placeholders:
    return default
  return value


def _env_bool(name: str, default: bool = False) -> bool:
  value = str(os.getenv(name, "") or "").strip().lower()
  if not value:
    return bool(default)
  return value in {"1", "true", "yes", "on"}


def _env_int(name: str, default: int) -> int:
  value = str(os.getenv(name, "") or "").strip()
  if not value:
    return int(default)
  try:
    return int(value)
  except Exception:
    return int(default)

ELEVENLABS_API_KEY = os.getenv("ELEVENLABS_API_KEY", "")
ELEVENLABS_VOICE_ID = os.getenv("ELEVENLABS_VOICE_ID", "EXAVITQu4vr4xnSDxMaL")
ELEVENLABS_VOICE_ID_ES = os.getenv("ELEVENLABS_VOICE_ID_ES", "").strip() or ELEVENLABS_VOICE_ID
ELEVENLABS_MODEL_ID = os.getenv("ELEVENLABS_MODEL_ID", "eleven_turbo_v2").strip() or "eleven_turbo_v2"
ELEVENLABS_MODEL_ID_ES = os.getenv("ELEVENLABS_MODEL_ID_ES", "eleven_multilingual_v2").strip() or "eleven_multilingual_v2"
NOVA_VOICE_ENABLED = _env_bool("NOVA_VOICE_ENABLED", default=bool(ELEVENLABS_API_KEY))
NOVA_AUTO_SUMMARIZE_ON_DISCONNECT = _env_bool("NOVA_AUTO_SUMMARIZE_ON_DISCONNECT", default=False)
NOVA_AUTOSEND_PENDING_ON_STARTUP = _env_bool("NOVA_AUTOSEND_PENDING_ON_STARTUP", default=False)
NOVA_AUTO_SONNET_FOR_ATTACHMENTS = _env_bool("NOVA_AUTO_SONNET_FOR_ATTACHMENTS", default=True)
NOVA_AUTO_SONNET_CHAT_MODEL = os.getenv("NOVA_AUTO_SONNET_CHAT_MODEL", "claude-sonnet-4-5-20250929").strip() or "claude-sonnet-4-5-20250929"
NOVA_AUTO_SONNET_VISION_MODEL = os.getenv("NOVA_AUTO_SONNET_VISION_MODEL", "claude-sonnet-4-5-20250929").strip() or "claude-sonnet-4-5-20250929"
NOVA_ECONOMY_CHAT_MODEL = os.getenv("NOVA_ECONOMY_CHAT_MODEL", "claude-haiku-4-5-20251001").strip() or "claude-haiku-4-5-20251001"
NOVA_ECONOMY_VISION_MODEL = os.getenv("NOVA_ECONOMY_VISION_MODEL", "claude-haiku-4-5-20251001").strip() or "claude-haiku-4-5-20251001"
NOVA_MODEL_IDLE_RESET_SECONDS = max(60, _env_int("NOVA_MODEL_IDLE_RESET_SECONDS", 420))
_eleven_client = None
EXECUTOR = ThreadPoolExecutor(max_workers=max(4, (os.cpu_count() or 2)))


async def _send_audio_later(websocket: WebSocket, text: str, context: str = "response"):
  """Generate ElevenLabs audio off the main request path and send when ready."""
  try:
    if not NOVA_VOICE_ENABLED:
      return
    loop = asyncio.get_event_loop()
    audio_b64 = await loop.run_in_executor(EXECUTOR, _synthesize_elevenlabs_b64, text)
    if not audio_b64:
      return
    await websocket.send_json({
      "type": "nova_audio",
      "context": context,
      "audio_b64": audio_b64,
    })
  except Exception as e:
    log.warning(f"Deferred audio send failed: {e}")


def _synthesize_elevenlabs_b64(text: str, force: bool = False, language: str = "en") -> str | None:
  """Return base64 MP3 audio for text using ElevenLabs, or None if unavailable.

  force=True bypasses the UI voice toggle — used for phone-call audio, which
  must be synthesized regardless of the chat voice setting.

  language="es" selects the Spanish voice + multilingual model so Spanish is
  spoken with a natural accent instead of an English-accented one.
  """
  global _eleven_client

  if (not NOVA_VOICE_ENABLED and not force) or not ELEVENLABS_API_KEY or not text:
    return None

  clean = str(text)
  clean = clean.replace("**", "")
  clean = clean.replace("`", "")
  clean = clean.replace("#", "")
  clean = clean.replace("\n", " ").strip()
  if not clean:
    return None

  if str(language).lower().startswith("es"):
    voice_id = ELEVENLABS_VOICE_ID_ES
    model_id = ELEVENLABS_MODEL_ID_ES
  else:
    voice_id = ELEVENLABS_VOICE_ID
    model_id = ELEVENLABS_MODEL_ID

  last_err = None
  for attempt in range(3):
    try:
      if _eleven_client is None:
        from elevenlabs.client import ElevenLabs
        _eleven_client = ElevenLabs(api_key=ELEVENLABS_API_KEY)

      audio_iter = _eleven_client.text_to_speech.convert(
        text=clean,
        voice_id=voice_id,
        model_id=model_id,
        output_format="mp3_44100_128",
      )
      audio_bytes = b"".join(audio_iter)
      if audio_bytes:
        return base64.b64encode(audio_bytes).decode("ascii")
      last_err = "empty audio"
    except Exception as e:
      last_err = e
      log.warning(f"ElevenLabs synthesis failed (attempt {attempt + 1}/3): {e}")
      # Rebuild the client in case it went stale, then retry.
      _eleven_client = None
      time.sleep(0.4)
  log.error(f"ElevenLabs synthesis gave up after retries: {last_err}")
  return None


app = FastAPI(title="Nova UI")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Active WebSocket registry (for wake_trigger broadcast) ────────────
_active_websockets: list = []
NOVA_MEMORY_DB_PATH = Path(_HERE) / "nova_memory.db"

# ── RingCentral incoming-call live monitor state ───────────────────────
CALL_MONITOR_ENABLED = _env_bool("NOVA_CALL_MONITOR_ENABLED", default=True)
CALL_MONITOR_INTERVAL_SECONDS = max(3, int(os.getenv("NOVA_CALL_MONITOR_INTERVAL_SECONDS", "6") or "6"))
CALL_MONITOR_LOOKBACK_MINUTES = max(1, int(os.getenv("NOVA_CALL_MONITOR_LOOKBACK_MINUTES", "20") or "20"))
CALL_MONITOR_VERBOSE = _env_bool("NOVA_CALL_MONITOR_VERBOSE", default=True)
_call_monitor_task: asyncio.Task | None = None
_seen_incoming_calls: set[str] = set()
_seen_live_call_sessions: set[str] = set()
_missed_call_bootstrapped = False
_telephony_sessions_unavailable = False
_calllog_backoff_until: datetime | None = None
_calllog_unavailable = False
_next_missed_poll_at: datetime | None = None

# ── Scheduled reminder monitor state ────────────────────────────────────
REMINDER_MONITOR_ENABLED = _env_bool("NOVA_REMINDER_MONITOR_ENABLED", default=True)
REMINDER_MONITOR_VERBOSE = _env_bool("NOVA_REMINDER_MONITOR_VERBOSE", default=True)
REMINDER_MONITOR_INTERVAL_SECONDS = max(15, int(os.getenv("NOVA_REMINDER_MONITOR_INTERVAL_SECONDS", "30") or "30"))
REMINDER_DEFAULT_REPEAT_MINUTES = max(1, int(os.getenv("NOVA_REMINDER_REPEAT_MINUTES", "30") or "30"))
_reminder_monitor_task: asyncio.Task | None = None


def _calllog_in_backoff() -> bool:
  if _calllog_unavailable:
    return True
  if _calllog_backoff_until is None:
    return False
  return datetime.utcnow() < _calllog_backoff_until


def _set_calllog_backoff(seconds: int = 75) -> None:
  global _calllog_backoff_until
  _calllog_backoff_until = datetime.utcnow() + timedelta(seconds=max(15, int(seconds)))
  if CALL_MONITOR_VERBOSE:
    remaining = int((_calllog_backoff_until - datetime.utcnow()).total_seconds())
    log.info(f"Call monitor call-log backoff set for {remaining}s")


def _handle_calllog_error(error_text: str) -> None:
  global _calllog_unavailable
  err = str(error_text or "")
  if "401" in err or "unauthorized" in err.lower():
    _calllog_unavailable = True
    log.info("Call monitor call-log unavailable (401); disabling call-log polling paths")
    return
  if "429" in err or "too many requests" in err.lower():
    _set_calllog_backoff()


def _parse_iso_utc_naive(value: str) -> datetime | None:
  raw = str(value or "").strip()
  if not raw:
    return None
  try:
    return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
  except Exception:
    return None


def _ensure_reminder_schedule_schema(conn: sqlite3.Connection) -> None:
  cols = {
    str(r[1])
    for r in conn.execute("PRAGMA table_info(reminders)").fetchall()
    if len(r) > 1
  }
  if "due_at" not in cols:
    conn.execute("ALTER TABLE reminders ADD COLUMN due_at TEXT")
  if "remind_every_minutes" not in cols:
    conn.execute("ALTER TABLE reminders ADD COLUMN remind_every_minutes INTEGER NOT NULL DEFAULT 30")
  if "last_notified_at" not in cols:
    conn.execute("ALTER TABLE reminders ADD COLUMN last_notified_at TEXT")
  if "notification_count" not in cols:
    conn.execute("ALTER TABLE reminders ADD COLUMN notification_count INTEGER NOT NULL DEFAULT 0")
  conn.execute("CREATE INDEX IF NOT EXISTS idx_rem_due ON reminders(due_at)")


def _fetch_due_reminder_notifications() -> list[dict]:
  if not NOVA_MEMORY_DB_PATH.exists():
    return []

  now = datetime.now()
  now_iso = now.isoformat()
  out: list[dict] = []

  try:
    conn = sqlite3.connect(str(NOVA_MEMORY_DB_PATH))
    conn.row_factory = sqlite3.Row
    _ensure_reminder_schedule_schema(conn)

    rows = conn.execute(
      """
      SELECT id, content, tag, due_at, remind_every_minutes, last_notified_at, notification_count
      FROM reminders
      WHERE status='active' AND due_at IS NOT NULL AND TRIM(COALESCE(due_at, '')) <> ''
      ORDER BY due_at ASC
      """
    ).fetchall()

    for r in rows:
      due_dt = _parse_iso_utc_naive(r["due_at"])
      if due_dt is None or now < due_dt:
        continue

      try:
        cadence = int(r["remind_every_minutes"] or REMINDER_DEFAULT_REPEAT_MINUTES)
      except Exception:
        cadence = REMINDER_DEFAULT_REPEAT_MINUTES
      cadence = max(1, cadence)

      last_dt = _parse_iso_utc_naive(r["last_notified_at"])
      if last_dt is not None:
        elapsed = (now - last_dt).total_seconds()
        if elapsed < cadence * 60:
          continue

      out.append({
        "id": int(r["id"]),
        "content": str(r["content"] or ""),
        "tag": str(r["tag"] or "general"),
        "due_at": str(r["due_at"] or ""),
        "remind_every_minutes": cadence,
        "notification_count": int(r["notification_count"] or 0),
      })

      conn.execute(
        "UPDATE reminders SET last_notified_at = ?, notification_count = COALESCE(notification_count, 0) + 1 WHERE id = ?",
        (now_iso, int(r["id"])),
      )

    conn.commit()
    conn.close()
  except Exception as e:
    log.warning(f"Reminder monitor DB scan failed: {e}")
    return []

  return out


def _format_due_reminder_alert(row: dict) -> str:
  content = str(row.get("content") or "Reminder")
  due_at = str(row.get("due_at") or "")
  every = int(row.get("remind_every_minutes") or REMINDER_DEFAULT_REPEAT_MINUTES)
  due_label = due_at if due_at else "scheduled time"
  return f"Reminder: {content} (due {due_label}). I will keep reminding every {every} minutes until you mark it done."

# ── Nova agent — fresh instance per browser session ───────────────────
def get_nova():
  """Create a fresh Nova instance — clean history guaranteed."""
  try:
    from dmelogic.nova_agent import NovaAgent
    nova = NovaAgent(voice=NOVA_VOICE_ENABLED)
    log.info("Fresh Nova agent created")
    return nova
  except Exception as e:
    log.error(f"Failed to load Nova: {e}")
    raise


async def _broadcast_json(payload: dict) -> int:
  """Send a JSON payload to all active websocket clients."""
  sent = 0
  dead = []
  for ws in list(_active_websockets):
    try:
      await ws.send_json(payload)
      sent += 1
    except Exception:
      dead.append(ws)
  for ws in dead:
    try:
      _active_websockets.remove(ws)
    except ValueError:
      pass
  return sent


def _incoming_call_key(item: dict) -> str:
  number = str(item.get("number") or "")
  time_raw = str(item.get("time") or "")
  duration = str(item.get("duration") or "")
  direction = str(item.get("direction") or "")
  result = str(item.get("result") or "")
  return "|".join([number, time_raw, duration, direction, result])


def _live_call_key(item: dict) -> str:
  call_id = str(item.get("call_id") or "")
  session_id = str(item.get("session_id") or item.get("id") or "")
  number = str(item.get("number") or "")
  start_time = str(item.get("time") or item.get("start_time") or "")
  if call_id:
    return "|".join([session_id, call_id, number])
  if session_id and number:
    return "|".join([session_id, number])
  return "|".join([session_id, number, start_time])


def _fetch_new_missed_incoming_calls() -> list[dict]:
  """Fetch missed inbound calls from RingCentral and return only unseen ones."""
  if _calllog_in_backoff():
    return []

  try:
    import dmelogic.nova_ringcentral as rc_tools
  except Exception as e:
    log.warning(f"Call monitor import failed: {e}")
    return []

  date_from = (datetime.utcnow() - timedelta(minutes=CALL_MONITOR_LOOKBACK_MINUTES)).strftime("%Y-%m-%dT%H:%M:%SZ")
  result = rc_tools.get_call_log(direction="Inbound", limit=100, date_from=date_from)
  if not isinstance(result, dict) or result.get("error"):
    if isinstance(result, dict) and result.get("error") and CALL_MONITOR_VERBOSE:
      log.info(f"Call monitor missed-call path error: {result.get('error')}")
    if isinstance(result, dict):
      _handle_calllog_error(str(result.get("error") or ""))
    return []

  rows = [r for r in (result.get("records") or []) if bool(r.get("missed"))]
  # oldest -> newest so notifications are emitted in order
  rows = sorted(rows, key=lambda r: str(r.get("time") or ""))

  new_rows = []
  for row in rows:
    key = _incoming_call_key(row)
    if key in _seen_incoming_calls:
      continue
    _seen_incoming_calls.add(key)
    new_rows.append(dict(row))

  # Bound dedupe cache to avoid unbounded growth.
  if len(_seen_incoming_calls) > 2000:
    # Keep only keys from most recent rows.
    latest_keys = {_incoming_call_key(r) for r in rows[-500:]}
    _seen_incoming_calls.clear()
    _seen_incoming_calls.update(latest_keys)

  # enrich known party match
  for row in new_rows:
    number = row.get("number")
    if not number:
      row["known_party"] = {"known": False}
      continue
    try:
      match = rc_tools.match_caller_to_patient(number)
      row["known_party"] = match if isinstance(match, dict) else {"known": False}
    except Exception:
      row["known_party"] = {"known": False}

  return new_rows


def _prime_missed_call_cache() -> None:
  """Seed missed-call dedupe cache without emitting notifications."""
  if _calllog_in_backoff():
    return

  try:
    import dmelogic.nova_ringcentral as rc_tools
  except Exception:
    return

  date_from = (datetime.utcnow() - timedelta(minutes=CALL_MONITOR_LOOKBACK_MINUTES)).strftime("%Y-%m-%dT%H:%M:%SZ")
  result = rc_tools.get_call_log(direction="Inbound", limit=100, date_from=date_from)
  if not isinstance(result, dict) or result.get("error"):
    if isinstance(result, dict) and result.get("error") and CALL_MONITOR_VERBOSE:
      log.info(f"Call monitor prime path error: {result.get('error')}")
    if isinstance(result, dict):
      _handle_calllog_error(str(result.get("error") or ""))
    return

  rows = [r for r in (result.get("records") or []) if bool(r.get("missed"))]
  for row in rows:
    _seen_incoming_calls.add(_incoming_call_key(row))


def _fetch_live_incoming_calls() -> list[dict]:
  """Fetch currently active inbound calls from telephony sessions."""
  global _telephony_sessions_unavailable

  try:
    import dmelogic.nova_ringcentral as rc_tools
  except Exception as e:
    log.warning(f"Live call monitor import failed: {e}")
    return []

  result = None if _telephony_sessions_unavailable else rc_tools.get_active_calls(limit=200)
  rows = []

  # Primary path: telephony sessions (best real-time source when scope permits).
  if isinstance(result, dict) and not result.get("error"):
    active_status_markers = {"setup", "proceeding", "ringing", "connected", "onhold", "parked"}
    for sess in (result.get("records") or []):
      direction = str(sess.get("direction") or "")
      telephony_status = str(sess.get("telephony_status") or "")
      status_l = telephony_status.lower()
      if direction.lower() != "inbound":
        continue
      if status_l and status_l not in active_status_markers:
        continue

      parties = sess.get("parties") or []
      caller_number = ""
      for p in parties:
        candidate = str((p or {}).get("from") or "").strip()
        if candidate:
          caller_number = candidate
          break

      row = {
        "session_id": sess.get("id"),
        "caller": caller_number or "Unknown caller",
        "number": caller_number,
        "time": sess.get("start_time"),
        "direction": "Inbound",
        "result": "Live",
        "live": True,
        "telephony_status": telephony_status,
      }

      if caller_number:
        try:
          match = rc_tools.match_caller_to_patient(caller_number)
          row["known_party"] = match if isinstance(match, dict) else {"known": False}
        except Exception:
          row["known_party"] = {"known": False}
      else:
        row["known_party"] = {"known": False}

      rows.append(row)

  # Fallback path: some accounts cannot query telephony sessions in real-time scope.
  # Use very recent inbound call-log records that are not yet marked missed/failed.
  elif isinstance(result, dict) and result.get("error"):
    err = str(result.get("error") or "")
    if "404" in err:
      _telephony_sessions_unavailable = True
      log.info("Call monitor live-session endpoint unavailable (404); disabling live-session path")
    elif CALL_MONITOR_VERBOSE:
      log.info(f"Call monitor live-session path error: {err}")

  # Secondary path: extension presence with detailed telephony state.
  if not rows:
    try:
      svc = rc_tools._get_rc_service()
      if svc and getattr(svc, "is_connected", False):
        presence_url = f"{svc.config.server_url}/restapi/v1.0/account/~/extension/~/presence"
        resp = requests.get(
          presence_url,
          headers=svc._get_auth_header(),
          params={"detailedTelephonyState": "true"},
          timeout=8,
        )
        if resp.ok:
          pdata = resp.json() if resp.content else {}
          if not isinstance(pdata, dict):
            pdata = {}

          telephony_status = str(pdata.get("telephonyStatus") or "")
          if telephony_status.lower() in {"ringing", "callconnected", "onhold", "busy"}:
            active_calls = pdata.get("activeCalls") or []
            if not isinstance(active_calls, list):
              active_calls = []
            caller_number = ""
            caller_display = ""
            call_id = ""
            call_started = ""
            for call in active_calls:
              if not isinstance(call, dict):
                continue

              call_direction = str((call or {}).get("direction") or "").strip().lower()
              if call_direction not in {"inbound", "outbound"}:
                call_direction = "inbound"

              if not call_id:
                call_id = str(
                  (call or {}).get("telephonySessionId")
                  or (call or {}).get("sessionId")
                  or (call or {}).get("id")
                  or ""
                ).strip()
              if not call_started:
                call_started = str(
                  (call or {}).get("startTime")
                  or (call or {}).get("sessionStartTime")
                  or ""
                ).strip()

              # Use the remote party number based on direction:
              # inbound -> caller is in "from"
              # outbound -> destination is in "to"
              source_key = "to" if call_direction == "outbound" else "from"
              source_fallback_name = "toName" if call_direction == "outbound" else "fromName"
              source_fallback_number = "toNumber" if call_direction == "outbound" else "fromNumber"
              from_info = (call or {}).get(source_key)
              num = ""
              if isinstance(from_info, dict):
                num = str(from_info.get("phoneNumber") or from_info.get("extensionNumber") or "").strip()
                if not caller_display:
                  caller_display = str(from_info.get("name") or "").strip()
              elif isinstance(from_info, str):
                num = from_info.strip()
                if not caller_display:
                  caller_display = num

              if not num:
                num = str((call or {}).get(source_fallback_number) or (call or {}).get("phoneNumber") or "").strip()

              if not caller_display:
                caller_display = str((call or {}).get(source_fallback_name) or (call or {}).get("name") or "").strip()

              if num.lower().startswith("tel:"):
                num = num[4:].strip()

              if num:
                caller_number = num
                break

            row = {
              "session_id": str(pdata.get("uri") or "presence"),
              "call_id": call_id,
              "caller": caller_number or caller_display or "Unknown caller",
              "number": caller_number,
              "time": call_started or (datetime.utcnow().isoformat() + "Z"),
              "direction": ("Outbound" if call_direction == "outbound" else "Inbound"),
              "result": "Live",
              "live": True,
              "telephony_status": telephony_status,
            }
            if caller_number:
              try:
                match = rc_tools.match_caller_to_patient(caller_number)
                row["known_party"] = match if isinstance(match, dict) else {"known": False}
              except Exception:
                row["known_party"] = {"known": False}
            else:
              row["known_party"] = {"known": False}

            rows.append(row)
    except Exception as e:
      if CALL_MONITOR_VERBOSE:
        log.info(f"Call monitor presence path error: {e}")

  if not rows and (not _calllog_in_backoff()):
    date_from = (datetime.utcnow() - timedelta(minutes=2)).strftime("%Y-%m-%dT%H:%M:%SZ")
    log_result = rc_tools.get_call_log(direction="Inbound", limit=50, date_from=date_from)
    if isinstance(log_result, dict) and not log_result.get("error"):
      now = datetime.utcnow()
      for rec in (log_result.get("records") or []):
        result_label = str(rec.get("result") or "").strip().lower()
        if result_label in {"missed", "no answer", "rejected", "cancelled", "failed", "voicemail"}:
          continue

        ts_raw = str(rec.get("time") or "")
        ts = None
        try:
          ts = datetime.fromisoformat(ts_raw.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
          ts = None

        # Keep only very recent records to approximate live ringing.
        if ts is not None and (now - ts).total_seconds() > 45:
          continue

        caller_number = str(rec.get("number") or "").strip()
        row = {
          "session_id": "",
          "caller": rec.get("caller") or caller_number or "Unknown caller",
          "number": caller_number,
          "time": rec.get("time"),
          "direction": "Inbound",
          "result": "Live",
          "live": True,
          "telephony_status": "Ringing",
        }
        if caller_number:
          try:
            match = rc_tools.match_caller_to_patient(caller_number)
            row["known_party"] = match if isinstance(match, dict) else {"known": False}
          except Exception:
            row["known_party"] = {"known": False}
        else:
          row["known_party"] = {"known": False}

        rows.append(row)
    elif isinstance(log_result, dict) and log_result.get("error") and CALL_MONITOR_VERBOSE:
      log.info(f"Call monitor live-fallback path error: {log_result.get('error')}")
      _handle_calllog_error(str(log_result.get("error") or ""))

  rows = sorted(rows, key=lambda r: str(r.get("time") or ""))
  return rows


def _format_call_alert(row: dict) -> str:
  caller = str(row.get("caller") or row.get("number") or "Unknown caller")
  number = str(row.get("number") or "")
  direction = str(row.get("direction") or "Inbound").strip() or "Inbound"
  result = str(row.get("result") or "")
  known = (row.get("known_party") or {}) if isinstance(row.get("known_party"), dict) else {}
  if known.get("known"):
    kind = str(known.get("kind") or "contact")
    name = str(known.get("name") or "")
    who = f" ({kind}: {name})" if name else f" ({kind})"
  else:
    who = ""
  result_part = f" | {result}" if result else ""
  if row.get("live"):
    status = str(row.get("telephony_status") or "Ringing")
    result_part = f" | LIVE ({status})"
  elif row.get("missed"):
    result_part = " | Missed"
  number_part = f" [{number}]" if number else ""
  prefix = "Outgoing call" if direction.lower() == "outbound" else "Incoming call"
  return f"{prefix}: {caller}{number_part}{who}{result_part}"


async def _run_incoming_call_monitor():
  """Background task: poll RingCentral and push immediate inbound-call alerts."""
  global _missed_call_bootstrapped, _next_missed_poll_at
  log.info(
    f"Incoming call monitor started (enabled={CALL_MONITOR_ENABLED}, interval={CALL_MONITOR_INTERVAL_SECONDS}s, lookback={CALL_MONITOR_LOOKBACK_MINUTES}m)"
  )
  while True:
    try:
      # Skip polling when nobody is connected to the UI.
      if _active_websockets:
        loop = asyncio.get_event_loop()
        alerts_sent = 0

        live_calls = await loop.run_in_executor(EXECUTOR, _fetch_live_incoming_calls)
        current_live_keys = set()
        for row in live_calls:
          key = _live_call_key(row)
          current_live_keys.add(key)
          if key in _seen_live_call_sessions:
            continue
          _seen_live_call_sessions.add(key)
          text = _format_call_alert(row)
          await _broadcast_json({
            "type": "incoming_call_alert",
            "text": text,
            "call": row,
          })
          alerts_sent += 1

        # Prune stale live session ids so set does not grow forever.
        if _seen_live_call_sessions:
          _seen_live_call_sessions.intersection_update(current_live_keys)

        if not _missed_call_bootstrapped:
          await loop.run_in_executor(EXECUTOR, _prime_missed_call_cache)
          _missed_call_bootstrapped = True
          continue

        now = datetime.utcnow()
        missed_calls = []
        if _next_missed_poll_at is None or now >= _next_missed_poll_at:
          missed_calls = await loop.run_in_executor(EXECUTOR, _fetch_new_missed_incoming_calls)
          _next_missed_poll_at = now + timedelta(seconds=60)

        for row in missed_calls:
          text = _format_call_alert(row)
          await _broadcast_json({
            "type": "incoming_call_alert",
            "text": text,
            "call": row,
          })
          alerts_sent += 1

        if CALL_MONITOR_VERBOSE:
          log.info(
            f"Call monitor tick: websockets={len(_active_websockets)} live={len(live_calls)} missed={len(missed_calls)} alerts_sent={alerts_sent}"
          )
    except Exception as e:
      log.warning(f"Incoming call monitor tick failed: {e}")

    await asyncio.sleep(CALL_MONITOR_INTERVAL_SECONDS)


async def _run_scheduled_reminder_monitor():
  """Background task: push due reminders and repeat until user marks done."""
  log.info(
    f"Reminder monitor started (enabled={REMINDER_MONITOR_ENABLED}, interval={REMINDER_MONITOR_INTERVAL_SECONDS}s, repeat_default={REMINDER_DEFAULT_REPEAT_MINUTES}m)"
  )
  while True:
    try:
      if _active_websockets:
        loop = asyncio.get_event_loop()
        due_rows = await loop.run_in_executor(EXECUTOR, _fetch_due_reminder_notifications)
        sent = 0
        for row in due_rows:
          text = _format_due_reminder_alert(row)
          await _broadcast_json({
            "type": "reminder_due",
            "text": text,
            "reminder": row,
          })
          sent += 1
        if REMINDER_MONITOR_VERBOSE:
          log.info(
            f"Reminder monitor tick: websockets={len(_active_websockets)} due={len(due_rows)} alerts_sent={sent}"
          )
    except Exception as e:
      log.warning(f"Reminder monitor tick failed: {e}")

    await asyncio.sleep(REMINDER_MONITOR_INTERVAL_SECONDS)


@app.on_event("startup")
async def _startup_background_monitors():
  global _call_monitor_task, _reminder_monitor_task
  if not CALL_MONITOR_ENABLED:
    log.info("Incoming call monitor disabled via NOVA_CALL_MONITOR_ENABLED")
  elif not (_call_monitor_task and not _call_monitor_task.done()):
    _call_monitor_task = asyncio.create_task(_run_incoming_call_monitor())

  if not REMINDER_MONITOR_ENABLED:
    log.info("Reminder monitor disabled via NOVA_REMINDER_MONITOR_ENABLED")
  elif not (_reminder_monitor_task and not _reminder_monitor_task.done()):
    _reminder_monitor_task = asyncio.create_task(_run_scheduled_reminder_monitor())


@app.on_event("shutdown")
async def _shutdown_background_monitors():
  global _call_monitor_task, _reminder_monitor_task
  if _call_monitor_task and not _call_monitor_task.done():
    _call_monitor_task.cancel()
    try:
      await _call_monitor_task
    except Exception:
      pass
  _call_monitor_task = None

  if _reminder_monitor_task and not _reminder_monitor_task.done():
    _reminder_monitor_task.cancel()
    try:
      await _reminder_monitor_task
    except Exception:
      pass
  _reminder_monitor_task = None


def _upsert_env_value(key: str, value: str) -> None:
    """Create/update key=value in .env without disturbing unrelated lines."""
    try:
        lines = []
        if ENV_PATH.exists():
            lines = ENV_PATH.read_text(encoding="utf-8").splitlines()

        prefix = f"{key}="
        updated = False
        out = []
        for ln in lines:
            if ln.startswith(prefix):
                out.append(f"{key}={value}")
                updated = True
            else:
                out.append(ln)
        if not updated:
            out.append(f"{key}={value}")

        ENV_PATH.write_text("\n".join(out).strip() + "\n", encoding="utf-8")
    except Exception as e:
        log.warning(f"Could not persist {key} to .env: {e}")


_ANTHROPIC_MODEL_ALIASES = {
  "claude-3-5-haiku-latest": "claude-haiku-4-5-20251001",
  "claude-3-5-haiku-20241022": "claude-haiku-4-5-20251001",
  "claude-3-5-sonnet-latest": "claude-sonnet-4-5-20250929",
  "claude-3-5-sonnet-20241022": "claude-sonnet-4-5-20250929",
  "claude-3-7-sonnet-latest": "claude-sonnet-4-5-20250929",
  "claude-3-7-sonnet-20250219": "claude-sonnet-4-5-20250929",
  "claude-sonnet-4-latest": "claude-sonnet-4-5-20250929",
  "claude-sonnet-4-20250514": "claude-sonnet-4-5-20250929",
  "claude-opus-4-latest": "claude-opus-4-5-20251101",
  "claude-opus-4-20250514": "claude-opus-4-5-20251101",
}

_DEFAULT_MODEL_FALLBACKS = [
  "claude-haiku-4-5-20251001",
  "claude-sonnet-4-5-20250929",
  "claude-opus-4-5-20251101",
]


def _normalize_anthropic_model_name(name: str, *, default: str) -> str:
  raw = str(name or "").strip()
  if not raw:
    return default
  return _ANTHROPIC_MODEL_ALIASES.get(raw.lower(), raw)


def _normalize_anthropic_fallbacks_csv(csv_text: str) -> str:
  parts = []
  for token in str(csv_text or "").split(","):
    normalized = _normalize_anthropic_model_name(token, default="").strip()
    if normalized and normalized not in parts:
      parts.append(normalized)
  for model_name in _DEFAULT_MODEL_FALLBACKS:
    if model_name not in parts:
      parts.append(model_name)
  return ",".join(parts)


def _set_anthropic_models(na_module, model: str | None = None, vision_model: str | None = None,
                          fallbacks: str | None = None, persist: bool = True) -> dict:
    """Apply model overrides to current process and optional .env persistence."""
    current_model = model if model is not None else getattr(na_module, "CLAUDE_MODEL", os.getenv("CLAUDE_MODEL", ""))
    current_vision = vision_model if vision_model is not None else getattr(na_module, "CLAUDE_VISION_MODEL", os.getenv("CLAUDE_VISION_MODEL", ""))
    current_fallbacks = fallbacks if fallbacks is not None else getattr(na_module, "CLAUDE_MODEL_FALLBACKS", os.getenv("CLAUDE_MODEL_FALLBACKS", ""))

    model_val = _normalize_anthropic_model_name(current_model, default=NOVA_ECONOMY_CHAT_MODEL)
    vision_val = _normalize_anthropic_model_name(current_vision, default=NOVA_ECONOMY_VISION_MODEL)
    fallbacks_val = _normalize_anthropic_fallbacks_csv(current_fallbacks)

    os.environ["CLAUDE_MODEL"] = model_val
    na_module.CLAUDE_MODEL = model_val
    os.environ["CLAUDE_VISION_MODEL"] = vision_val
    na_module.CLAUDE_VISION_MODEL = vision_val
    os.environ["CLAUDE_MODEL_FALLBACKS"] = fallbacks_val
    na_module.CLAUDE_MODEL_FALLBACKS = fallbacks_val

    if persist:
      _upsert_env_value("CLAUDE_MODEL", model_val)
      _upsert_env_value("CLAUDE_VISION_MODEL", vision_val)
      _upsert_env_value("CLAUDE_MODEL_FALLBACKS", fallbacks_val)

    return {
        "model": getattr(na_module, "CLAUDE_MODEL", ""),
        "vision_model": getattr(na_module, "CLAUDE_VISION_MODEL", ""),
        "fallbacks": getattr(na_module, "CLAUDE_MODEL_FALLBACKS", ""),
    }


def _is_sonnet_or_opus(model_name: str) -> bool:
    lowered = str(model_name or "").strip().lower()
    return ("sonnet" in lowered) or ("opus" in lowered)


def _reset_models_to_haiku(na_module, persist: bool = False) -> dict:
    # Economy baseline for both chat and vision after heavy/idle periods.
    return _set_anthropic_models(
        na_module,
        model=NOVA_ECONOMY_CHAT_MODEL,
        vision_model=NOVA_ECONOMY_VISION_MODEL,
        persist=persist,
    )

# ── HTML UI ────────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NOVA Assistant - Central Pharmacy Group</title>
<meta name="application-name" content="NOVA Assistant">
<meta name="apple-mobile-web-app-title" content="NOVA Assistant">
<meta name="theme-color" content="#00d4aa">
<link rel="icon" type="image/x-icon" href="/favicon.ico">
<link rel="icon" type="image/png" sizes="512x512" href="/nova-icon.png">
<link rel="apple-touch-icon" href="/nova-icon.png">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {
    --bg:        #0a0f1e;
    --surface:   #111827;
    --surface2:  #1a2235;
    --border:    #1f2d45;
    --accent:    #00d4aa;
    --accent2:   #0099ff;
    --text:      #e8edf5;
    --text-dim:  #7a8ba0;
    --nova-bg:   #0d1f35;
    --user-bg:   #0f2a1a;
    --error:     #ff4757;
    --warning:   #ffa502;
    --radius:    14px;
    --font:      'DM Sans', sans-serif;
    --mono:      'DM Mono', monospace;
  }

  * { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    font-family: var(--font);
    background: var(--bg);
    color: var(--text);
    height: 100vh;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }

  /* ── Header ── */
  header {
    display: flex;
    align-items: center;
    gap: 14px;
    padding: 16px 24px;
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    flex-shrink: 0;
  }

  .nova-avatar {
    width: 42px; height: 42px;
    border-radius: 50%;
    background: linear-gradient(135deg, var(--accent), var(--accent2));
    display: flex; align-items: center; justify-content: center;
    font-size: 20px;
    flex-shrink: 0;
    box-shadow: 0 0 20px rgba(0,212,170,0.3);
  }

  .nova-avatar.speaking {
    animation: pulse 1.2s ease-in-out infinite;
  }

  @keyframes pulse {
    0%, 100% { box-shadow: 0 0 20px rgba(0,212,170,0.3); }
    50% { box-shadow: 0 0 35px rgba(0,212,170,0.7); }
  }

  .header-info h1 {
    font-size: 15px;
    font-weight: 600;
    letter-spacing: 0.3px;
  }

  .header-info p {
    font-size: 12px;
    color: var(--text-dim);
    margin-top: 1px;
  }

  .status-dot {
    width: 7px; height: 7px;
    border-radius: 50%;
    background: var(--accent);
    display: inline-block;
    margin-right: 5px;
    animation: blink 2s ease-in-out infinite;
  }

  @keyframes blink {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.4; }
  }

  .header-right {
    margin-left: auto;
    display: flex;
    align-items: center;
    gap: 10px;
    flex-wrap: wrap;
    justify-content: flex-end;
  }

  .pill {
    font-size: 11px;
    font-weight: 500;
    padding: 4px 10px;
    border-radius: 20px;
    background: rgba(0,212,170,0.1);
    color: var(--accent);
    border: 1px solid rgba(0,212,170,0.2);
    font-family: var(--mono);
  }

  .pill.active {
    background: rgba(0,212,170,0.22);
    border-color: rgba(0,212,170,0.6);
    color: #9dfbe8;
  }

  .model-controls {
    display: flex;
    align-items: center;
    gap: 6px;
    padding: 4px 8px;
    border-radius: 10px;
    border: 1px solid var(--border);
    background: rgba(17,24,39,0.85);
  }

  .model-controls select,
  .model-controls input {
    background: var(--surface2);
    color: var(--text);
    border: 1px solid var(--border);
    border-radius: 7px;
    padding: 4px 6px;
    font-size: 11px;
    font-family: var(--mono);
    min-width: 165px;
  }

  .model-controls button {
    background: rgba(0,212,170,0.16);
    color: var(--accent);
    border: 1px solid rgba(0,212,170,0.35);
    border-radius: 7px;
    padding: 4px 8px;
    cursor: pointer;
    font-size: 11px;
    font-family: var(--mono);
  }

  .model-controls button:hover {
    background: rgba(0,212,170,0.24);
  }

  /* ── Main layout ── */
  main {
    display: flex;
    flex: 1;
    overflow: hidden;
  }

  /* ── Sidebar ── */
  aside {
    width: clamp(260px, 28vw, 340px);
    background: var(--surface);
    border-right: 1px solid var(--border);
    display: flex;
    flex-direction: column;
    flex-shrink: 0;
    overflow-y: auto;
  }

  .sidebar-section {
    padding: 16px;
    border-bottom: 1px solid var(--border);
  }

  .sidebar-section h3 {
    font-size: 10px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1.2px;
    color: var(--text-dim);
    margin-bottom: 10px;
  }

  .quick-btn {
    display: block;
    width: 100%;
    text-align: left;
    padding: 8px 10px;
    border-radius: 8px;
    background: transparent;
    border: 1px solid transparent;
    color: var(--text-dim);
    font-size: 13px;
    font-family: var(--font);
    cursor: pointer;
    transition: all 0.15s;
    margin-bottom: 4px;
  }

  .quick-btn:hover {
    background: var(--surface2);
    border-color: var(--border);
    color: var(--text);
  }

  .quick-btn .btn-icon {
    margin-right: 8px;
    font-size: 14px;
  }

  #reminders-list {
    font-size: 12px;
    color: var(--text-dim);
  }

  .reminder-actions {
    display: flex;
    align-items: center;
    gap: 6px;
    margin-bottom: 8px;
  }

  .reminder-actions button {
    font-size: 10px;
    font-family: var(--mono);
    padding: 4px 6px;
    border-radius: 6px;
    border: 1px solid var(--border);
    background: var(--surface2);
    color: var(--text-dim);
    cursor: pointer;
  }

  .reminder-actions button:hover {
    border-color: rgba(0,212,170,0.45);
    color: var(--accent);
  }

  .reminder-actions button.danger:hover {
    border-color: rgba(255,71,87,0.55);
    color: #ff8a95;
  }

  .reminder-item {
    padding: 6px 0;
    border-bottom: 1px solid var(--border);
    display: grid;
    grid-template-columns: auto auto minmax(0, 1fr);
    column-gap: 6px;
    row-gap: 4px;
    align-items: start;
  }

  .reminder-item:last-child { border-bottom: none; }

  .reminder-check {
    margin-top: 2px;
    accent-color: var(--accent);
    cursor: pointer;
    grid-column: 1;
    grid-row: 1;
  }

  .reminder-text {
    grid-column: 3;
    grid-row: 1;
    color: var(--text);
    line-height: 1.35;
    min-width: 0;
    overflow-wrap: anywhere;
  }

  .reminder-meta {
    display: block;
    color: var(--text-dim);
    font-size: 10px;
    margin-top: 2px;
    line-height: 1.3;
  }

  .reminder-row-actions {
    grid-column: 2 / 4;
    grid-row: 2;
    display: flex;
    flex-direction: row;
    flex-wrap: wrap;
    gap: 4px;
    margin-left: 0;
    flex-shrink: 0;
  }

  .reminder-row-actions button {
    font-size: 10px;
    font-family: var(--mono);
    padding: 3px 6px;
    border-radius: 6px;
    border: 1px solid var(--border);
    background: var(--surface2);
    color: var(--text-dim);
    cursor: pointer;
  }

  .reminder-row-actions button:hover {
    border-color: rgba(0,212,170,0.45);
    color: var(--accent);
  }

  .reminder-row-actions button.danger:hover {
    border-color: rgba(255,71,87,0.55);
    color: #ff8a95;
  }

  .reminder-edit {
    grid-column: 2 / 4;
    display: grid;
    grid-template-columns: 1fr;
    gap: 6px;
    width: 100%;
  }

  .reminder-edit input,
  .reminder-edit textarea,
  .reminder-edit select {
    width: 100%;
    background: #141821;
    border: 1px solid var(--border);
    border-radius: 6px;
    color: var(--text);
    font-size: 11px;
    padding: 6px;
    font-family: var(--font);
    box-sizing: border-box;
  }

  .reminder-edit textarea {
    min-height: 54px;
    resize: vertical;
  }

  .reminder-edit-row {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 6px;
  }

  .reminder-tag {
    grid-column: 2;
    grid-row: 1;
    font-size: 9px;
    font-weight: 600;
    padding: 2px 5px;
    border-radius: 4px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    flex-shrink: 0;
    margin-top: 1px;
  }

  .tag-ordering { background: rgba(0,153,255,0.15); color: #4da6ff; }
  .tag-calls    { background: rgba(255,165,0,0.15);  color: #ffa500; }
  .tag-billing  { background: rgba(0,212,170,0.15);  color: var(--accent); }
  .tag-general  { background: rgba(122,139,160,0.15); color: var(--text-dim); }
  .tag-clinical { background: rgba(255,71,87,0.15);  color: #ff6b7a; }
  .tag-follow_up{ background: rgba(155,89,182,0.15); color: #c39bd3; }

  /* ── Chat area ── */
  .chat-area {
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }

  #messages {
    flex: 1;
    overflow-y: auto;
    padding: 20px 24px;
    display: flex;
    flex-direction: column;
    gap: 16px;
    scroll-behavior: smooth;
  }

  #messages::-webkit-scrollbar { width: 4px; }
  #messages::-webkit-scrollbar-track { background: transparent; }
  #messages::-webkit-scrollbar-thumb { background: var(--border); border-radius: 2px; }

  .message {
    display: flex;
    gap: 10px;
    max-width: 80%;
    animation: fadeUp 0.2s ease-out;
  }

  @keyframes fadeUp {
    from { opacity: 0; transform: translateY(8px); }
    to   { opacity: 1; transform: translateY(0); }
  }

  .message.nova { align-self: flex-start; }
  .message.user { align-self: flex-end; flex-direction: row-reverse; }

  .msg-avatar {
    width: 30px; height: 30px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 14px;
    flex-shrink: 0;
    margin-top: 2px;
  }

  .message.nova .msg-avatar {
    background: linear-gradient(135deg, var(--accent), var(--accent2));
  }

  .message.user .msg-avatar {
    background: var(--surface2);
    border: 1px solid var(--border);
  }

  .msg-bubble {
    padding: 10px 14px;
    border-radius: var(--radius);
    font-size: 14px;
    line-height: 1.55;
    max-width: 100%;
    word-wrap: break-word;
  }

  .message.nova .msg-bubble {
    background: var(--nova-bg);
    border: 1px solid var(--border);
    border-radius: 4px var(--radius) var(--radius) var(--radius);
    color: var(--text);
  }

  .message.user .msg-bubble {
    background: var(--user-bg);
    border: 1px solid rgba(0,212,170,0.15);
    border-radius: var(--radius) 4px var(--radius) var(--radius);
    color: var(--text);
  }

  .msg-bubble strong { color: var(--accent); font-weight: 600; }
  .msg-bubble em     { color: var(--text-dim); font-style: normal; }

  .msg-time {
    font-size: 10px;
    color: var(--text-dim);
    margin-top: 4px;
    font-family: var(--mono);
  }

  .message.nova .msg-time { margin-left: 40px; }
  .message.user .msg-time { margin-right: 40px; text-align: right; }

  /* Tool call indicator */
  .tool-call {
    align-self: flex-start;
    margin-left: 40px;
    font-size: 11px;
    color: var(--text-dim);
    font-family: var(--mono);
    padding: 4px 10px;
    background: var(--surface2);
    border-radius: 6px;
    border-left: 2px solid var(--accent);
    animation: fadeUp 0.2s ease-out;
  }

  /* Typing indicator */
  .typing-indicator {
    align-self: flex-start;
    display: flex;
    gap: 10px;
    align-items: center;
  }

  .typing-dots {
    display: flex;
    gap: 4px;
    padding: 12px 16px;
    background: var(--nova-bg);
    border: 1px solid var(--border);
    border-radius: 4px var(--radius) var(--radius) var(--radius);
  }

  .typing-dots span {
    width: 6px; height: 6px;
    border-radius: 50%;
    background: var(--accent);
    animation: bounce 1.2s ease-in-out infinite;
  }

  .typing-dots span:nth-child(2) { animation-delay: 0.2s; }
  .typing-dots span:nth-child(3) { animation-delay: 0.4s; }

  @keyframes bounce {
    0%, 60%, 100% { transform: translateY(0); opacity: 0.4; }
    30% { transform: translateY(-6px); opacity: 1; }
  }

  /* ── Input area ── */
  .input-area {
    padding: 16px 24px;
    background: var(--surface);
    border-top: 1px solid var(--border);
    display: flex;
    gap: 10px;
    align-items: flex-end;
    flex-shrink: 0;
  }

  .input-wrap {
    flex: 1;
    min-width: 0;
    position: relative;
  }

  #msg-input {
    width: 100%;
    display: block;
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 10px 14px;
    color: var(--text);
    font-family: var(--font);
    font-size: 14px;
    resize: none;
    min-height: 42px;
    max-height: 120px;
    outline: none;
    transition: border-color 0.15s;
    line-height: 1.4;
  }

  #msg-input:focus { border-color: var(--accent); }
  #msg-input::placeholder { color: var(--text-dim); }

  .send-btn {
    width: 42px; height: 42px;
    border-radius: 10px;
    background: var(--accent);
    border: none;
    cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px;
    transition: all 0.15s;
    flex-shrink: 0;
    color: #000;
  }

  .send-btn:hover { background: #00efc0; transform: scale(1.05); }
  .send-btn:disabled { background: var(--border); cursor: not-allowed; transform: none; color: var(--text-dim); }

  .attach-btn {
    width: 42px; height: 42px;
    border-radius: 10px;
    background: var(--surface2);
    border: 1px solid var(--border);
    cursor: pointer;
    font-size: 18px;
    transition: all 0.15s;
    flex-shrink: 0;
    display: flex; align-items: center; justify-content: center;
  }
  .attach-btn:hover { background: var(--border); }

  .mic-btn {
    width: 42px; height: 42px;
    border-radius: 10px;
    background: var(--surface2);
    border: 1px solid var(--border);
    cursor: pointer;
    font-size: 18px;
    transition: all 0.15s;
    flex-shrink: 0;
    display: flex; align-items: center; justify-content: center;
    color: var(--text);
  }
  .mic-btn:hover { background: var(--border); }
  .mic-btn.recording {
    background: rgba(255, 71, 87, 0.18);
    border-color: rgba(255, 71, 87, 0.8);
    color: #ff8a95;
    box-shadow: 0 0 0 3px rgba(255, 71, 87, 0.15);
  }

  /* ── Alerts banner ── */
  #alert-bar {
    display: none;
    background: rgba(255,165,0,0.08);
    border-bottom: 1px solid rgba(255,165,0,0.2);
    padding: 8px 24px;
    font-size: 12px;
    color: var(--warning);
    flex-shrink: 0;
  }

  #alert-bar.show { display: flex; align-items: center; gap: 8px; }

  /* ── Connection status ── */
  #conn-status {
    position: fixed;
    bottom: 80px;
    right: 20px;
    font-size: 11px;
    padding: 6px 12px;
    border-radius: 20px;
    font-family: var(--mono);
    display: none;
    z-index: 100;
  }

  #conn-status.error {
    display: block;
    background: rgba(255,71,87,0.15);
    border: 1px solid var(--error);
    color: var(--error);
  }

  #conn-status.show {
    display: block;
    background: rgba(0,212,170,0.12);
    border: 1px solid rgba(0,212,170,0.45);
    color: var(--accent);
  }

  /* ── Mic button ── */
  .mic-btn {
    width: 42px; height: 42px;
    border-radius: 10px;
    background: var(--surface2);
    border: 1px solid var(--border);
    cursor: pointer;
    font-size: 18px;
    transition: all 0.15s;
    flex-shrink: 0;
    display: flex; align-items: center; justify-content: center;
  }
  .mic-btn:hover { background: var(--border); }
  .mic-btn.recording {
    background: rgba(255,71,87,0.15);
    border-color: var(--error);
    animation: pulse 1s ease-in-out infinite;
  }

  /* ── Active pill (wake mode on) ── */
  .pill.active {
    background: rgba(0,212,170,0.2);
    border-color: var(--accent);
    color: var(--accent);
  }

  /* ── Connection status — info variant ── */
  #conn-status.show {
    display: block;
    background: rgba(0,153,255,0.1);
    border: 1px solid rgba(0,153,255,0.3);
    color: var(--accent2);
  }

  /* ── Scrollbar for sidebar ── */
  aside::-webkit-scrollbar { width: 3px; }
  aside::-webkit-scrollbar-thumb { background: var(--border); }

  /* ── 📞 Call card ── */
  #call-card {
    display: none;
    position: fixed;
    right: 18px;
    bottom: 84px;
    width: 320px;
    background: var(--bg-card, #141a26);
    border: 1px solid var(--accent);
    border-radius: 12px;
    padding: 14px;
    z-index: 300;
    box-shadow: 0 8px 30px rgba(0,0,0,0.5);
  }
  #call-card .cc-title { font-size: 12px; color: var(--accent); letter-spacing: 1px; margin-bottom: 6px; }
  #call-card .cc-who { font-size: 15px; font-weight: 600; margin-bottom: 4px; }
  #call-card .cc-patient { font-size: 12px; color: var(--text-dim); margin-bottom: 8px; }
  #call-card .cc-timer { font-family: monospace; font-size: 13px; color: var(--accent2); margin-bottom: 10px; }
  #call-card .cc-transcript {
    max-height: 140px; overflow-y: auto; font-size: 12px;
    border-top: 1px solid var(--border); padding-top: 6px; margin-bottom: 10px;
  }
  #call-card .cc-transcript div { margin-bottom: 4px; }
  #call-card .cc-buttons { display: flex; gap: 8px; }
  #call-card .cc-buttons button {
    flex: 1; padding: 8px; border-radius: 8px; border: 1px solid var(--border);
    background: var(--bg, #0b0f17); color: inherit; cursor: pointer; font-size: 12px;
  }
  #call-card .cc-buttons button.danger { border-color: var(--error); color: var(--error); }
</style>
</head>
<body>

<!-- Header -->
<header>
  <div class="nova-avatar" id="nova-avatar">🤖</div>
  <div class="header-info">
    <h1><span class="status-dot"></span>Nova</h1>
    <p>Central Pharmacy Group · Bronx, NY</p>
  </div>
  <div class="header-right">
    <span class="pill" id="model-pill">Haiku</span>
    <button class="pill" id="voice-pill" type="button" onclick="toggleVoice()">🔊 Voice</button>
    <button class="pill" id="wake-pill" type="button" onclick="toggleWakeMode()">👂 Wake Off</button>
    <button class="pill" id="phone-pill" type="button" onclick="toggleAnswerCalls()" style="opacity:0.6">📞 Answer Calls</button>
    <button class="pill" id="active-listen-pill" type="button" onclick="activateActiveListen()">🎙 Active Listen</button>
    <span class="pill" id="memory-pill">🧠 Memory</span>
    <div class="model-controls">
      <select id="model-select" title="Chat model">
        <option value="">Chat model</option>
        <option value="claude-3-5-haiku-20241022">claude-3-5-haiku-20241022</option>
        <option value="claude-3-5-sonnet-20241022">claude-3-5-sonnet-20241022</option>
        <option value="claude-3-7-sonnet-20250219">claude-3-7-sonnet-20250219</option>
        <option value="claude-sonnet-4-20250514">claude-sonnet-4-20250514</option>
        <option value="claude-opus-4-20250514">claude-opus-4-20250514</option>
      </select>
      <select id="vision-model-select" title="Vision model">
        <option value="">Vision model</option>
        <option value="claude-3-5-haiku-20241022">claude-3-5-haiku-20241022</option>
        <option value="claude-3-5-sonnet-20241022">claude-3-5-sonnet-20241022</option>
        <option value="claude-3-7-sonnet-20250219">claude-3-7-sonnet-20250219</option>
        <option value="claude-sonnet-4-20250514">claude-sonnet-4-20250514</option>
        <option value="claude-opus-4-20250514">claude-opus-4-20250514</option>
      </select>
      <button type="button" onclick="applyModelSettings()">Apply</button>
      <button type="button" onclick="refreshModelSettings()">Show</button>
    </div>
  </div>
</header>

<!-- Alert bar -->
<div id="alert-bar">
  <span>⚠️</span>
  <span id="alert-text"></span>
</div>

<main>
  <!-- Sidebar -->
  <aside>
    <div class="sidebar-section">
      <h3>Quick Actions</h3>
      <button class="quick-btn" onclick="sendMsg('morning summary')">
        <span class="btn-icon">🌅</span>Morning Summary
      </button>
      <button class="quick-btn" onclick="sendMsg('refills due this week')">
        <span class="btn-icon">💊</span>Refills Due
      </button>
      <button class="quick-btn" onclick="sendMsg('what items need to be reordered')">
        <span class="btn-icon">📦</span>Reorder List
      </button>
      <button class="quick-btn" onclick="sendMsg('show pending approvals')">
        <span class="btn-icon">✅</span>Pending Approvals
      </button>
      <button class="quick-btn" onclick="sendMsg('billing summary')">
        <span class="btn-icon">💵</span>Billing Summary
      </button>
      <button class="quick-btn" onclick="sendMsg('show my reminders')">
        <span class="btn-icon">🔔</span>My Reminders
      </button>
    </div>

    <div class="sidebar-section" style="flex:1">
      <h3>Active Reminders</h3>
      <div class="reminder-actions">
        <button type="button" onclick="toggleAllReminders(true)">Select all</button>
        <button type="button" onclick="toggleAllReminders(false)">Clear</button>
        <button type="button" class="danger" onclick="deleteSelectedReminders()">Delete selected</button>
      </div>
      <div id="reminders-list">
        <div style="color:var(--text-dim);font-size:12px">No active reminders</div>
      </div>
    </div>
  </aside>

  <!-- Chat -->
  <div class="chat-area">
    <div id="messages"></div>
    <div class="input-area">
      <input type="file" id="img-file-input" multiple accept="image/*,.pdf,application/pdf,.xlsx,.xls,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel" style="display:none" onchange="handleFileSelect(event)">
      <button class="attach-btn" onclick="document.getElementById('img-file-input').click()" title="Attach file">📎</button>
      <button class="mic-btn" id="mic-btn" type="button" onclick="toggleMicInput()" title="Start voice input">🎤</button>
      <button class="mic-btn" id="interrupt-btn" type="button" onclick="interruptAndRedirect()" title="Stop Nova and redirect">⛔</button>
      <div class="input-wrap">
        <textarea id="msg-input" placeholder="Ask Nova anything... (Ctrl+V to paste screenshot)" rows="1"></textarea>
        <div id="img-preview-wrap" style="display:none">
          <img id="img-preview" style="max-height:80px;border-radius:6px;margin:4px 0;border:1px solid var(--border)">
          <button onclick="clearImage()" style="position:absolute;top:2px;right:2px;background:var(--error);border:none;color:white;border-radius:50%;width:18px;height:18px;cursor:pointer;font-size:11px;line-height:18px;text-align:center">×</button>
        </div>
      </div>
      <button class="send-btn" id="send-btn" onclick="sendFromInput()">➤</button>
    </div>
  </div>
</main>

<div id="conn-status"></div>

<!-- 📞 Live call card -->
<div id="call-card">
  <div class="cc-title">📞 NOVA IS ON A CALL</div>
  <div class="cc-who" id="call-card-who">Unknown caller</div>
  <div class="cc-patient" id="call-card-patient"></div>
  <div class="cc-timer" id="call-card-timer">00:00</div>
  <div class="cc-transcript" id="call-card-transcript"></div>
  <div class="cc-buttons">
    <button type="button" id="call-card-transfer" onclick="phoneTransfer()">👤 Take over</button>
    <button type="button" class="danger" onclick="phoneHangup()">⏹ Hang up</button>
  </div>
</div>
<audio id="phone-remote-audio" style="display:none"></audio>

<script src="/vendor/ringcentral-web-phone.min.js"></script>

<script>
const messagesEl = document.getElementById('messages');
const inputEl    = document.getElementById('msg-input');
const sendBtn    = document.getElementById('send-btn');
const avatarEl   = document.getElementById('nova-avatar');
const voicePill  = document.getElementById('voice-pill');
const wakePill   = document.getElementById('wake-pill');
const activeListenPill = document.getElementById('active-listen-pill');
const micBtn     = document.getElementById('mic-btn');
const interruptBtn = document.getElementById('interrupt-btn');
const modelSelect = document.getElementById('model-select');
const visionModelSelect = document.getElementById('vision-model-select');
let ws = null;
let isProcessing = false;
let typingEl = null;
let greeted = false;
window.NOVA_AUTOSEND_PENDING_ON_STARTUP = false;
let voiceEnabled = true;
let speechQueue = [];
let speakingNow = false;
let audioUnlocked = false;
let serverAudioQueue = [];
let serverAudioPlaying = false;
let currentServerAudio = null;
let recognition = null;
let recognitionActive = false;
let speechFinalText = '';
let speechInterimText = '';
let speechDraftPrefix = '';
const SpeechRecognitionCtor = window.SpeechRecognition || window.webkitSpeechRecognition;
let wakeRecognition = null;
let activeListenBusy = false;
let wakeModeEnabled = false;
let wakeAwaitingCommand = false;
const WAKE_PHRASES = ['hey nova', 'okay nova', 'ok nova', 'nova'];
const WAKE_END_PHRASES = [
  'thats all',
  "that's all",
  'thank you nova',
  'thanks nova',
  'goodbye nova',
  'bye nova',
  'we are done',
  "we're done",
  'end conversation',
  'stop listening',
  'never mind',
];
const CLOSE_WINDOW_PHRASES = [
  'close your window',
  'close the window',
  'close this window',
  'close nova window',
  'close nova',
  'exit nova',
  'quit nova',
];
const WAKE_CONVERSATION_TIMEOUT_MS = 60 * 1000;
let wakeConversationActive = false;
let wakeSilenceTimer = null;
let wakePausedForResponse = false;
let wakeLastHandledText = '';
let wakeLastHandledAt = 0;
let wakePendingUtterance = '';
let wakeCommitTimer = null;
let wakeRestartTimer = null;
let wakeStartInFlight = false;
let wakeEchoGuardUntil = 0;
let wakeRecognitionActive = false;
let wakeNoSpeechCount = 0;
let wakeHealthTimer = null;
let wakeLastStartAt = 0;
const WAKE_HEALTH_CHECK_MS = 5000;
const WAKE_RESTART_AFTER_MS = 4 * 60 * 1000;
const WAKE_OWNER_KEY = 'novaWakeOwner';
const WAKE_OWNER_TTL_MS = 12000;
const wakeInstanceId = `nova-${Date.now()}-${Math.random().toString(16).slice(2)}`;
let reminderSelection = new Set();
let remindersById = new Map();
let reminderEditState = new Set();
const NOVA_UI_STATE_KEY = 'novaUIState';
const BILLING_REMINDER_HOURS = [10, 12, 14, 16, 18];
const BILLING_REMINDER_STORAGE_KEY = 'novaBillingReminderSlots';
let billingReminderTimer = null;

function saveNovaUIState() {
  try {
    localStorage.setItem(NOVA_UI_STATE_KEY, JSON.stringify({
      voiceEnabled,
      wakeModeEnabled,
    }));
  } catch (_) {}
}

function loadNovaUIState() {
  try {
    const raw = localStorage.getItem(NOVA_UI_STATE_KEY);
    if (!raw) return;
    const state = JSON.parse(raw);
    if (typeof state.voiceEnabled === 'boolean') voiceEnabled = state.voiceEnabled;
    // wakeModeEnabled is activated via toggleWakeMode() after initSpeechInput()
    return state;
  } catch (_) {}
  return null;
}

function localGreeting() {
  const hour = new Date().getHours();
  const part = hour < 12 ? 'morning' : (hour < 17 ? 'afternoon' : 'evening');
  return `Good ${part}. Nova online.`;
}

/* ══════════════ 📞 Answer Calls — Nova softphone (Phase 1) ══════════════ */
const PHONE_OWNER_KEY = 'novaPhoneOwner';
const PHONE_OWNER_TTL_MS = 12000;
let answerCallsEnabled = false;
let novaWebPhone = null;
let phoneAudioCtx = null;
let phoneTtsDest = null;
let phoneSilentOsc = null;   // keeps the synthetic outbound track continuously live
let phoneCurrentSession = null;
let phoneConsultSession = null;
let phoneCurrentSource = null;
let phoneCurrentAudioEl = null;
let phonePlaybackStartedAt = 0;
let phoneLastSelfHealAt = 0;
let phoneSenderKeepaliveInt = null;
let phoneLastSpokenB64 = '';
let phoneLastRtpHealAt = 0;
let phoneCallStartedAt = 0;
let phoneCallTimerInt = null;
let phoneMaxCallTimer = null;
let phoneOwnerBeat = null;
let phoneConfig = { transfer_number: '', max_call_seconds: 600, greeting: '' };
let phoneDebug = true;  // surface WebRTC media diagnostics into the chat (temporary)

// Step-by-step call tracer: writes to the server's nova_call_debug.log so the
// whole answer flow can be followed without opening browser DevTools.
function dbg(step, detail) {
  try {
    fetch('/phone/debug-log', {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ step: String(step), detail: (detail === undefined ? '' : detail) }),
      keepalive: true,
    }).catch(() => {});
  } catch (_) {}
}
const PHONE_CANT_HEAR_RE = /\b(i\s*(can't|cannot|do\s*not|don't)\s*hear\s*(you|anything)?|can't\s*hear\s*you|cannot\s*hear\s*you|no\s*te\s*escucho|no\s*puedo\s*escucharte|no\s*oigo)\b/i;

function setPhonePillState(mode) {
  const el = document.getElementById('phone-pill');
  if (!el) return;
  if (mode === 'on')        { el.textContent = '📞 Answering';    el.classList.add('active');    el.style.opacity = '1'; }
  else if (mode === 'call') { el.textContent = '📞 On a call';    el.classList.add('active');    el.style.opacity = '1'; }
  else if (mode === 'err')  { el.textContent = '📞 Phone error';  el.classList.remove('active'); el.style.opacity = '0.6'; }
  else                      { el.textContent = '📞 Answer Calls'; el.classList.remove('active'); el.style.opacity = '0.6'; }
}

function claimPhoneOwner() {
  try { localStorage.setItem(PHONE_OWNER_KEY, JSON.stringify({ id: wakeInstanceId, ts: Date.now() })); } catch (_) {}
}
function releasePhoneOwner() {
  try {
    const raw = localStorage.getItem(PHONE_OWNER_KEY);
    const o = raw ? JSON.parse(raw) : null;
    if (o && o.id === wakeInstanceId) localStorage.removeItem(PHONE_OWNER_KEY);
  } catch (_) {}
}
window.addEventListener('storage', (ev) => {
  if (ev.key === PHONE_OWNER_KEY && answerCallsEnabled) {
    try {
      const o = ev.newValue ? JSON.parse(ev.newValue) : null;
      if (o && o.id && o.id !== wakeInstanceId) disableAnswerCalls('another Nova window took over');
    } catch (_) {}
  }
});

function ensurePhoneAudio() {
  if (!phoneAudioCtx) {
    phoneAudioCtx = new (window.AudioContext || window.webkitAudioContext)();
    phoneTtsDest = phoneAudioCtx.createMediaStreamDestination();
  }
  if (phoneAudioCtx.state === 'suspended') phoneAudioCtx.resume().catch(() => {});
}

async function toggleAnswerCalls() {
  if (answerCallsEnabled) disableAnswerCalls('toggled off');
  else await enableAnswerCalls();
}

async function enableAnswerCalls() {
  const el = document.getElementById('phone-pill');
  try {
    ensurePhoneAudio();  // created inside the click gesture
    if (el) el.textContent = '📞 Connecting…';
    phoneConfig = await fetch('/phone/config').then(r => r.json()).catch(() => phoneConfig);
    const res = await fetch('/phone/sip-provision', { method: 'POST' });
    const data = await res.json().catch(() => ({}));
    if (!res.ok || !data.ok) {
      setPhonePillState('err');
      appendMessage('nova', '📞 Could not enable call answering: ' + (data.detail || data.error || ('HTTP ' + res.status)));
      return;
    }
    claimPhoneOwner();
    if (phoneOwnerBeat) clearInterval(phoneOwnerBeat);
    phoneOwnerBeat = setInterval(claimPhoneOwner, 5000);

    novaWebPhone = new WebPhone(data.provision, { appName: 'NovaPhone', appVersion: '1.0', logLevel: 1 });
    dbg('webphone_created');
    novaWebPhone.userAgent.on('registered', () => { dbg('sip_registered'); setPhonePillState(phoneCurrentSession ? 'call' : 'on'); });
    novaWebPhone.userAgent.on('registrationFailed', (e) => {
      dbg('sip_registration_failed', String((e && e.message) || e || ''));
      console.error('Softphone registration failed', e);
      setPhonePillState('err');
    });
    novaWebPhone.userAgent.on('invite', (session) => { dbg('sip_invite_received'); phoneOnInvite(session); });

    answerCallsEnabled = true;
    saveNovaUIState();
    setPhonePillState('on');
    fetch('/phone/state', { method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ enabled: true, owner: wakeInstanceId }) }).catch(() => {});
    appendMessage('nova', '📞 Answer Calls is ON — I will pick up incoming calls to the pharmacy line.');
  } catch (e) {
    console.error('enableAnswerCalls failed', e);
    setPhonePillState('err');
    appendMessage('nova', '📞 Could not enable call answering: ' + ((e && e.message) || e));
  }
}

function disableAnswerCalls(reason) {
  answerCallsEnabled = false;
  if (phoneOwnerBeat) { clearInterval(phoneOwnerBeat); phoneOwnerBeat = null; }
  releasePhoneOwner();
  try { if (phoneCurrentSession) phoneHangup(); } catch (_) {}
  try {
    if (novaWebPhone && novaWebPhone.userAgent) {
      try { novaWebPhone.userAgent.unregister(); } catch (_) {}
      try { novaWebPhone.userAgent.stop(); } catch (_) {}
    }
  } catch (_) {}
  novaWebPhone = null;
  saveNovaUIState();
  setPhonePillState('off');
  fetch('/phone/state', { method: 'POST', headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ enabled: false }) }).catch(() => {});
  if (reason) console.log('Answer Calls disabled:', reason);
}

function phoneCallerNumber(session) {
  try {
    const uri = (session.remoteIdentity && session.remoteIdentity.uri)
      || (session.request && session.request.from && session.request.from.uri);
    return (uri && (uri.user || '')) || '';
  } catch (_) { return ''; }
}
function phoneCallerName(session) {
  try { return (session.remoteIdentity && session.remoteIdentity.displayName) || ''; } catch (_) { return ''; }
}

async function phoneOnInvite(session) {
  if (!answerCallsEnabled) { dbg('invite_ignored_answer_off'); return; }
  if (phoneCurrentSession) { dbg('invite_ignored_already_on_call'); console.log('Already on a call — second call keeps ringing normally.'); return; }
  phoneCurrentSession = session;
  const num = phoneCallerNumber(session);
  const name = phoneCallerName(session);
  dbg('invite_accepting', { num: num, name: name });
  try {
    ensurePhoneAudio();
    dbg('audio_ctx_state', phoneAudioCtx ? phoneAudioCtx.state : 'none');
    try { session.on('terminated', () => phoneCleanupCall('caller hung up')); } catch (_) {}
    try { session.on('failed', () => phoneCleanupCall('call failed')); } catch (_) {}
    try { session.on('bye', () => phoneCleanupCall('caller hung up')); } catch (_) {}
    try { session.on('rejected', () => phoneCleanupCall('call rejected')); } catch (_) {}
    try { session.on('cancel', () => phoneCleanupCall('caller canceled')); } catch (_) {}
    await phoneAcceptCall(session);
    dbg('session_accepted');
    setPhonePillState('call');
    phoneWireAudio(session);
    showCallCard(name, num);
    startCallTimer();
    appendMessage('nova', '📞 Answered incoming call from ' + (name || 'Unknown') + (num ? ' [' + num + ']' : '') + '.');
    if (num) {
      fetch('/phone/match-caller?number=' + encodeURIComponent(num))
        .then(r => r.json()).then(m => updateCallCardPatient(m)).catch(() => {});
    }
    // Phase 2: open the per-call voice loop (greeting arrives over this WS)
    try {
      openCallAudioWs(num);
      dbg('call_audio_ws_opening');
      startCallVad(session);
      if (phoneSenderKeepaliveInt) { clearInterval(phoneSenderKeepaliveInt); phoneSenderKeepaliveInt = null; }
      phoneSenderKeepaliveInt = setInterval(() => {
        if (!phoneCurrentSession) return;
        phoneAttachTtsTrack(phoneCurrentSession, 2, 80).catch(() => {});
      }, 2000);
    } catch (e) {
      dbg('call_audio_setup_failed', String((e && e.message) || e || ''));
      console.error('call-audio setup failed, falling back to greeting only', e);
      const g = await fetch('/phone/greeting', {
        method: 'POST', headers: { 'Content-Type': 'application/json' }, body: '{}',
      }).then(r => r.json()).catch(() => null);
      if (g && g.audio_b64) playIntoCall(g.audio_b64);
    }
    if (phoneMaxCallTimer) clearTimeout(phoneMaxCallTimer);
    phoneMaxCallTimer = setTimeout(() => { try { phoneHangup(); } catch (_) {} },
      Math.max(60, Number(phoneConfig.max_call_seconds || 600)) * 1000);
  } catch (e) {
    dbg('invite_failed', String((e && e.message) || e || ''));
    console.error('phoneOnInvite failed', e);
    phoneCleanupCall('error answering');
  }
}

// Nova answers on a server PC that has no microphone. SIP.js still calls
// getUserMedia({audio:true}) to build the WebRTC offer; on a mic-less machine
// that rejects with NotFoundError, so the peerConnection is never created and
// the caller hears nothing (call falls through to voicemail). Nova never needs
// a real mic — she speaks via TTS and hears the caller through server-side
// transcription. So we hand SIP.js Nova's TTS output bus as a synthetic
// "microphone": negotiation succeeds, the peerConnection is created normally,
// and whatever Nova plays into the TTS bus is what the caller hears.
function phoneSyntheticMicStream() {
  ensurePhoneAudio();
  let track = phoneTtsDest && phoneTtsDest.stream.getAudioTracks()[0];
  if (!track || track.readyState === 'ended') {
    // A previous call closed and stopped the track; rebuild the TTS bus.
    phoneTtsDest = phoneAudioCtx.createMediaStreamDestination();
    phoneSilentOsc = null;
  }
  if (!phoneSilentOsc) {
    try {
      const osc = phoneAudioCtx.createOscillator();
      const g = phoneAudioCtx.createGain();
      g.gain.value = 0;                       // inaudible; keeps the track live
      osc.connect(g); g.connect(phoneTtsDest);
      osc.start();
      phoneSilentOsc = osc;
    } catch (_) {}
  }
  return phoneTtsDest.stream;
}

async function phoneAcceptCall(session) {
  const md = navigator.mediaDevices;
  const orig = (md && md.getUserMedia) ? md.getUserMedia.bind(md) : null;
  const synth = phoneSyntheticMicStream();
  try {
    md.getUserMedia = async (constraints) => {
      if (constraints && constraints.audio) { dbg('gum_synthetic_mic'); return synth; }
      if (orig) return orig(constraints);
      throw new DOMException('No media devices', 'NotFoundError');
    };
    await session.accept({ sessionDescriptionHandlerOptions: { constraints: { audio: true, video: false } } });
  } finally {
    if (orig) md.getUserMedia = orig; else { try { delete md.getUserMedia; } catch (_) {} }
  }
}

function phoneGetPeerConnection(session) {
  try {
    const sdh = session && session.sessionDescriptionHandler;
    if (!sdh) return null;
    return sdh.peerConnection || sdh._peerConnection
        || (sdh.sessionDescriptionHandler && sdh.sessionDescriptionHandler.peerConnection)
        || null;
  } catch (_) { return null; }
}
function phoneWireAudio(session, _attempt) {
  const attempt = _attempt || 0;
  try {
    const sdh = session.sessionDescriptionHandler;
    const pc = phoneGetPeerConnection(session);
    if (!pc) {
      // The peerConnection is created asynchronously during SDP negotiation;
      // wait for it (up to ~4s) instead of giving up, which was the cause of
      // 'answered but no audio' calls.
      if (attempt < 40 && phoneCurrentSession === session) {
        if (attempt === 0) dbg('wire_audio_awaiting_peerconnection');
        setTimeout(() => phoneWireAudio(session, attempt + 1), 100);
        return;
      }
      dbg('wire_audio_no_peerconnection'); console.error('No peerConnection on session'); return;
    }
    if (attempt > 0) dbg('wire_audio_peerconnection_ready', { attempts: attempt });
    dbg('wire_audio_start', { conn: pc.connectionState || '?', ice: pc.iceConnectionState || '?', sig: pc.signalingState || '?' });
    // Reliable remote-hangup detection across SIP.js versions: watch the
    // WebRTC connection state. If the media path drops (caller hung up or the
    // network died) end the call locally so the call card always clears.
    const onPcDown = () => {
      const st = (pc.connectionState || pc.iceConnectionState || '');
      dbg('pc_state_change', { conn: pc.connectionState || '?', ice: pc.iceConnectionState || '?' });
      if (st === 'failed' || st === 'closed') {
        console.log('peerConnection ' + st + ' — cleaning up call.');
        phoneCleanupCall('caller hung up');
      } else if (st === 'disconnected') {
        // Could be a transient ICE blip; confirm after a short grace period.
        setTimeout(() => {
          if (!phoneCurrentSession) return;
          const st2 = (pc.connectionState || pc.iceConnectionState || '');
          if (st2 === 'disconnected' || st2 === 'failed' || st2 === 'closed') {
            console.log('peerConnection still ' + st2 + ' — cleaning up call.');
            phoneCleanupCall('caller hung up');
          }
        }, 4000);
      }
    };
    try { pc.addEventListener('connectionstatechange', onPcDown); } catch (_) {}
    try { pc.addEventListener('iceconnectionstatechange', onPcDown); } catch (_) {}
    try { pc.addEventListener('icegatheringstatechange', () => dbg('ice_gathering', pc.iceGatheringState || '?')); } catch (_) {}
    // Outbound: attach Nova's TTS track to the call's audio sender.
    phoneAttachTtsTrack(session).then((ok) => {
      if (!ok) console.error('Could not attach TTS outbound track; caller may not hear Nova.');
      dbg('tts_track_attached', ok);
    });
    // Inbound: gather the caller's audio for local monitoring (and P2 VAD).
    const remoteStream = new MediaStream();
    pc.getReceivers().forEach(r => { if (r.track && r.track.kind === 'audio') remoteStream.addTrack(r.track); });
    const monitorEl = document.getElementById('phone-remote-audio');
    if (monitorEl) { monitorEl.srcObject = remoteStream; monitorEl.volume = 0.4; monitorEl.play().catch(() => {}); }
    session._novaRemoteStream = remoteStream;
    // Media snapshots so we can diagnose "caller hears nothing" without DevTools.
    const mediaSnap = (tag) => {
      try {
        const senders = (pc.getSenders && pc.getSenders()) || [];
        const recvs = (pc.getReceivers && pc.getReceivers()) || [];
        const sAudio = senders.filter(s => s.track && s.track.kind === 'audio');
        const rAudio = recvs.filter(r => r.track && r.track.kind === 'audio');
        dbg('media_snap_' + tag, {
          conn: pc.connectionState || '?', ice: pc.iceConnectionState || '?',
          iceGather: pc.iceGatheringState || '?', sig: pc.signalingState || '?',
          senders_audio: sAudio.length,
          sender_tracks: sAudio.map(s => s.track ? (s.track.readyState + (s.track.enabled ? '' : '/disabled')) : 'none'),
          receivers_audio: rAudio.length,
          receiver_tracks: rAudio.map(r => r.track ? (r.track.readyState + (r.track.muted ? '/muted' : '')) : 'none'),
          audioCtx: phoneAudioCtx ? phoneAudioCtx.state : 'none',
        });
      } catch (e) { dbg('media_snap_failed', String((e && e.message) || e || '')); }
    };
    setTimeout(() => mediaSnap('2s'), 2500);
    setTimeout(() => mediaSnap('8s'), 8000);
  } catch (e) { dbg('wire_audio_failed', String((e && e.message) || e || '')); console.error('phoneWireAudio failed', e); }
}

// Attach Nova's persistent TTS track onto the call's outbound audio sender.
// Safe and idempotent: never destroys the audio context or the TTS stream,
// only swaps the sender's track (retrying until the sender is ready).
async function phoneAttachTtsTrack(session, retries = 15, delayMs = 150) {
  try {
    if (!session) return false;
    ensurePhoneAudio();
    if (phoneAudioCtx && phoneAudioCtx.state === 'suspended') {
      try { await phoneAudioCtx.resume(); } catch (_) {}
    }
    const pc = phoneGetPeerConnection(session);
    if (!pc) return false;

    const ttsTrack = phoneTtsDest && phoneTtsDest.stream && phoneTtsDest.stream.getAudioTracks()[0];
    if (!ttsTrack) { console.error('No TTS track available for outbound call audio.'); return false; }
    ttsTrack.enabled = true;

    const maxRetries = Math.max(1, Number(retries || 1));
    const waitMs = Math.max(20, Number(delayMs || 120));
    for (let i = 0; i < maxRetries; i++) {
      const senders = (pc.getSenders && pc.getSenders()) || [];
      const sender = senders.find(s => s.track && s.track.kind === 'audio')
                  || senders.find(s => !s.track)
                  || senders[0];
      if (sender) {
        if (sender.track === ttsTrack) return true;
        try {
          await sender.replaceTrack(ttsTrack);
          return true;
        } catch (e) {
          console.error('replaceTrack failed', e);
        }
      }
      if (i < maxRetries - 1) await new Promise(r => setTimeout(r, waitMs));
    }
  } catch (e) {
    console.error('phoneAttachTtsTrack failed', e);
  }
  return false;
}

function playIntoCall(b64, onDone) {
  const done = () => { if (onDone) onDone(); };
  try {
    if (!b64) { done(); return; }
    ensurePhoneAudio();
    if (phoneAudioCtx && phoneAudioCtx.state === 'suspended') { phoneAudioCtx.resume().catch(() => {}); }
    // Make sure our TTS track is on the call sender (idempotent, non-destructive).
    phoneAttachTtsTrack(phoneCurrentSession, 3, 80).catch(() => {});
    stopCallPlayback();
    const bin = atob(b64);
    const buf = new Uint8Array(bin.length);
    for (let i = 0; i < bin.length; i++) buf[i] = bin.charCodeAt(i);

    phoneAudioCtx.decodeAudioData(buf.buffer.slice(0)).then(audioBuf => {
      const src = phoneAudioCtx.createBufferSource();
      src.buffer = audioBuf;
      src.connect(phoneTtsDest);                       // into the call
      const g = phoneAudioCtx.createGain();            // quiet local monitor
      g.gain.value = 0.2;
      src.connect(g); g.connect(phoneAudioCtx.destination);
      phoneCurrentSource = src;
      phonePlaybackStartedAt = Date.now();
      src._cancelled = false;
      src.onended = () => {
        if (phoneCurrentSource === src) phoneCurrentSource = null;
        if (!src._cancelled) done();                   // skip action if barge-in stopped us
      };
      src.start();
      dbg('tts_playing', { seconds: Math.round((audioBuf.duration || 0) * 10) / 10, ctx: phoneAudioCtx.state });
    }).catch(e => { dbg('tts_decode_failed', String((e && e.message) || e || '')); console.error('decodeAudioData failed', e); done(); });
  } catch (e) { dbg('tts_play_failed', String((e && e.message) || e || '')); console.error('playIntoCall failed', e); done(); }
}
function stopCallPlayback() {
  try {
    if (phoneCurrentSource) {
      phoneCurrentSource._cancelled = true;
      phoneCurrentSource.stop();
      phoneCurrentSource = null;
    }
    phonePlaybackStartedAt = 0;
  } catch (_) {}
}

function showCallCard(name, num) {
  const card = document.getElementById('call-card');
  if (!card) return;
  document.getElementById('call-card-who').textContent = (name || 'Unknown caller') + (num ? '  ·  ' + num : '');
  document.getElementById('call-card-patient').textContent = 'Matching caller to patient…';
  document.getElementById('call-card-transcript').innerHTML = '';
  document.getElementById('call-card-transfer').style.display = phoneConfig.transfer_number ? '' : 'none';
  card.style.display = 'block';
}
function updateCallCardPatient(m) {
  const el = document.getElementById('call-card-patient');
  if (!el) return;
  const p = (m && (m.patient || (m.matched && m))) || null;
  if (p) {
    const nm = p.name || ((p.first_name || '') + ' ' + (p.last_name || '')).trim();
    el.textContent = '👤 Patient: ' + (nm || 'matched');
  } else {
    el.textContent = 'No patient match for this number';
  }
}
function appendCallTranscript(role, text) {
  const box = document.getElementById('call-card-transcript');
  if (!box || !text) return;
  const div = document.createElement('div');
  div.textContent = (role === 'caller' ? '🗣 ' : '🤖 ') + text;
  box.appendChild(div);
  box.scrollTop = box.scrollHeight;
}
function startCallTimer() {
  phoneCallStartedAt = Date.now();
  const t = document.getElementById('call-card-timer');
  if (phoneCallTimerInt) clearInterval(phoneCallTimerInt);
  phoneCallTimerInt = setInterval(() => {
    if (!t) return;
    const s = Math.floor((Date.now() - phoneCallStartedAt) / 1000);
    t.textContent = String(Math.floor(s / 60)).padStart(2, '0') + ':' + String(s % 60).padStart(2, '0');
  }, 1000);
}

function phoneHangup() {
  const s = phoneCurrentSession;
  if (s) {
    try { (s.terminate || s.bye || s.dispose || s.reject || s.cancel || function () {}).call(s); }
    catch (e) { console.error('hangup terminate failed', e); }
  }
  // Always tear down the UI/timers even if there is no live SIP session,
  // so the "Hang up" button reliably dismisses a lingering call card.
  phoneCleanupCall('hung up');
}
async function phoneTransfer() {
  const s = phoneCurrentSession;
  if (!s || !phoneConfig.transfer_number) { dbg('transfer_skip', { has_session: !!s, has_number: !!phoneConfig.transfer_number }); return; }
  if (phoneConsultSession) { dbg('transfer_skip_inprogress'); return; } // transfer already in progress
  const rawTarget = String(phoneConfig.transfer_number || '');
  // RingCentral WebPhone outbound invites reject the leading "+" and any
  // punctuation — dial only the raw digits (e.g. 13476472347).
  const target = rawTarget.replace(/[^0-9]/g, '');
  dbg('transfer_start', { raw: rawTarget, dial: target });
  appendMessage('nova', '📞 Warm transfer: ringing a team member…');
  // Nova stops talking/listening while the caller is placed on hold.
  stopCallPlayback();
  stopCallVad();
  try { await s.hold(); dbg('transfer_caller_held'); } catch (e) { dbg('transfer_hold_failed', String((e && e.message) || e)); }
  let consult = null;
  // This PC has no microphone. The outbound consult INVITE also calls
  // getUserMedia({audio:true}) to build its WebRTC offer, which throws
  // NotFoundError and terminates the call in a few milliseconds. Hand the
  // SDK Nova's TTS bus as a synthetic mic (same trick as answering a call),
  // and keep it installed until the consult leg is settled.
  const md = navigator.mediaDevices;
  const origGum = (md && md.getUserMedia) ? md.getUserMedia.bind(md) : null;
  const synthMic = phoneSyntheticMicStream();
  let gumRestored = false;
  const restoreGum = () => {
    if (gumRestored) return; gumRestored = true;
    if (!md) return;
    if (origGum) md.getUserMedia = origGum; else { try { delete md.getUserMedia; } catch (_) {} }
  };
  try {
    md.getUserMedia = async (constraints) => {
      if (constraints && constraints.audio) { dbg('gum_synthetic_mic_transfer'); return synthMic; }
      if (origGum) return origGum(constraints);
      throw new DOMException('No media devices', 'NotFoundError');
    };
  } catch (_) {}
  // Safety net: never leave getUserMedia overridden indefinitely.
  const gumSafety = setTimeout(restoreGum, 8000);
  try {
    // Place a NEW consultation call to staff FROM Nova's extension so the
    // team member's phone shows Nova's caller ID ("Nova Transfer").
    consult = novaWebPhone.userAgent.invite(target, {});
    phoneConsultSession = consult;
    dbg('transfer_invite_created');
  } catch (e) {
    dbg('transfer_invite_throw', String((e && e.message) || e));
    console.error('consult invite failed', e);
    clearTimeout(gumSafety); restoreGum();
    appendMessage('nova', '📞 Transfer failed — could not reach a team member.');
    phoneConsultSession = null;
    try { await s.unhold(); } catch (_) {}
    // Tell the server so Nova speaks a fallback and takes a message.
    try {
      if (phoneCallWs && phoneCallWs.readyState === 1) {
        phoneCallWs.send(JSON.stringify({ type: 'transfer_failed', turn: phoneTurnCounter, reason: 'invite_error' }));
      }
    } catch (_) {}
    return;
  }
  let settled = false;
  const finish = async () => {
    if (settled) return; settled = true;
    clearTimeout(ringTimer);
    clearTimeout(gumSafety); restoreGum();
    dbg('transfer_staff_answered');
    try {
      await s.warmTransfer(consult);
      dbg('transfer_bridged');
      appendMessage('nova', '📞 Transfer complete — caller connected to the team member.');
      setTimeout(() => phoneCleanupCall('transferred'), 1500);
    } catch (e) {
      dbg('transfer_bridge_failed', String((e && e.message) || e));
      console.error('warm transfer complete failed', e);
      appendMessage('nova', '📞 Transfer failed — please pick up the desk phone.');
      phoneConsultSession = null;
      try { await s.unhold(); } catch (_) {}
    }
  };
  const abort = async (msg, why) => {
    if (settled) return; settled = true;
    clearTimeout(ringTimer);
    clearTimeout(gumSafety); restoreGum();
    dbg('transfer_abort', { why: why || 'unknown' });
    try { (consult.terminate || consult.dispose || consult.bye || function(){}).call(consult); } catch (_) {}
    phoneConsultSession = null;
    appendMessage('nova', msg || '📞 No team member answered — staying with the caller.');
    // Bring the caller back and resume Nova's listening loop so she is not
    // left deaf after a failed transfer.
    try { await s.unhold(); } catch (_) {}
    try { phoneAttachTtsTrack(s, 4, 100); } catch (_) {}
    try { if (!phoneVadState) startCallVad(s); } catch (_) {}
    // Tell the server the transfer failed so Nova speaks a fallback to the
    // caller, logs a callback task, and blocks any further transfer retry.
    try {
      if (phoneCallWs && phoneCallWs.readyState === 1) {
        phoneCallWs.send(JSON.stringify({ type: 'transfer_failed', turn: phoneTurnCounter, reason: why || 'no_answer' }));
      }
    } catch (_) {}
  };
  // Staff answered → bridge the caller to them.
  try { consult.on('accepted', finish); } catch (_) {}
  // Ringing indication (helps confirm the outbound leg actually reached RC).
  try { consult.on('progress', (r) => dbg('transfer_ringing', String((r && r.statusCode) || ''))); } catch (_) {}
  // Staff phone hung up / rejected before answering → return to caller.
  try { consult.on('terminated', (r) => abort(null, 'terminated' + (r && r.statusCode ? '_' + r.statusCode : ''))); } catch (_) {}
  try { consult.on('rejected', (r) => abort(null, 'rejected' + (r && r.statusCode ? '_' + r.statusCode : ''))); } catch (_) {}
  try { consult.on('failed', (r) => abort(null, 'failed' + (r && r.statusCode ? '_' + r.statusCode : ''))); } catch (_) {}
  // No answer within 30s → give up and return to caller.
  const ringTimer = setTimeout(() => abort('📞 No answer — I will take a message instead.', 'ring_timeout_30s'), 30000);
}
function phoneCleanupCall(reason) {
  stopCallPlayback();
  stopCallVad();
  if (phoneConsultSession) {
    try { (phoneConsultSession.terminate || phoneConsultSession.dispose || phoneConsultSession.bye || function(){}).call(phoneConsultSession); } catch (_) {}
    phoneConsultSession = null;
  }
  if (phoneSenderKeepaliveInt) { clearInterval(phoneSenderKeepaliveInt); phoneSenderKeepaliveInt = null; }
  if (phoneCallWs) {
    try { phoneCallWs.send(JSON.stringify({ type: 'call_end', reason: reason || 'ended' })); } catch (_) {}
    try { phoneCallWs.close(); } catch (_) {}
    phoneCallWs = null;
  }
  if (phoneCallTimerInt) { clearInterval(phoneCallTimerInt); phoneCallTimerInt = null; }
  if (phoneMaxCallTimer) { clearTimeout(phoneMaxCallTimer); phoneMaxCallTimer = null; }
  const card = document.getElementById('call-card');
  if (card) card.style.display = 'none';
  const monitorEl = document.getElementById('phone-remote-audio');
  if (monitorEl) monitorEl.srcObject = null;
  if (phoneCurrentSession) {
    appendMessage('nova', '📞 Call ended (' + (reason || 'done') + ').');
    phoneCurrentSession = null;
  }
  if (answerCallsEnabled) setPhonePillState('on');
}

/* ── Phase 2: per-call voice loop (WS + VAD) ── */
let phoneCallWs = null;
let phoneTurnCounter = 0;
let phoneVadState = null;

function openCallAudioWs(num) {
  const proto = location.protocol === 'https:' ? 'wss' : 'ws';
  phoneTurnCounter = 0;
  phoneCallWs = new WebSocket(proto + '://' + location.host + '/call-audio');
  phoneCallWs.onopen = () => {
    dbg('call_audio_ws_open');
    phoneCallWs.send(JSON.stringify({
      type: 'call_start',
      call_id: 'call-' + Date.now(),
      caller_number: num || '',
    }));
  };
  phoneCallWs.onmessage = (ev) => {
    let data;
    try { data = JSON.parse(ev.data); } catch (_) { return; }
    if (data.type === 'greeting') {
      dbg('greeting_received', { has_audio: !!data.audio_b64, text_len: (data.text || '').length });
      if (data.text) appendCallTranscript('nova', data.text);
      if (data.audio_b64) playIntoCall(data.audio_b64);
    } else if (data.type === 'reply') {
      if (data.turn && data.turn < phoneTurnCounter) return;  // stale turn after barge-in
      if (data.user_text) {
        appendCallTranscript('caller', data.user_text);
        if (PHONE_CANT_HEAR_RE.test(String(data.user_text))) {
          phoneAttachTtsTrack(phoneCurrentSession, 6, 120).then((ok) => {
            if (!ok) console.error('Re-attach failed after caller reported no audio.');
          });
        }
      }
      if (data.text) appendCallTranscript('nova', data.text);
      if (!data.action) {
        playIntoCall(data.audio_b64);
      } else {
        let actionFired = false;
        const fireAction = () => {
          if (actionFired) return;
          actionFired = true;
          phonePostReplyAction(data.action);
        };
        const words = String(data.text || '').trim().split(/\s+/).filter(Boolean).length;
        const actionFallbackMs = Math.min(12000, Math.max(2200, Math.round((words / 2.4) * 1000 + 900)));
        const timer = setTimeout(fireAction, actionFallbackMs);
        playIntoCall(data.audio_b64, () => { clearTimeout(timer); fireAction(); });
      }
    }
  };
  phoneCallWs.onerror = (e) => { dbg('call_audio_ws_error'); console.error('call-audio WS error', e); };
  phoneCallWs.onclose = () => { dbg('call_audio_ws_closed'); phoneCallWs = null; };
}

function phonePostReplyAction(action) {
  if (!action) return;
  if (action === 'hangup') phoneHangup();
  else if (action === 'transfer') phoneTransfer();
}

function startCallVad(session) {
  const stream = session && session._novaRemoteStream;
  if (!stream || !phoneAudioCtx) { console.error('VAD: no remote stream/audio ctx'); return; }
  const srcNode = phoneAudioCtx.createMediaStreamSource(stream);
  const analyser = phoneAudioCtx.createAnalyser();
  analyser.fftSize = 2048;
  srcNode.connect(analyser);
  const state = phoneVadState = {
    analyser, srcNode, stream,
    buf: new Float32Array(analyser.fftSize),
    recorder: null, chunks: [], recStartAt: 0,
    speaking: false, speechStartAt: 0, lastVoiceAt: 0,
    noiseFloor: 0.008, calibrated: false, calibSamples: [], startedAt: Date.now(),
    interval: null,
  };
  state.interval = setInterval(() => vadTick(state), 50);
}

function vadTick(state) {
  let rms = 0;
  try {
    state.analyser.getFloatTimeDomainData(state.buf);
    let sum = 0;
    for (let i = 0; i < state.buf.length; i++) sum += state.buf[i] * state.buf[i];
    rms = Math.sqrt(sum / state.buf.length);
  } catch (_) { return; }
  const now = Date.now();

  if (!state.calibrated) {                 // learn the line's noise floor (first 600ms)
    state.calibSamples.push(rms);
    if (now - state.startedAt >= 600) {
      const avg = state.calibSamples.reduce((a, b) => a + b, 0) / Math.max(1, state.calibSamples.length);
      state.noiseFloor = Math.max(0.006, avg * 3.0);
      state.calibrated = true;
    }
    return;
  }

  const voiced = rms > state.noiseFloor;
  if (voiced) state.lastVoiceAt = now;

  if (!state.speaking) {
    if (voiced) {
      if (!state.speechStartAt) state.speechStartAt = now;
      if (now - state.speechStartAt >= 200) {   // sustained speech → start of utterance
        state.speaking = true;
        onCallerSpeechStart(state);
      }
    } else {
      state.speechStartAt = 0;
    }
  } else if (!voiced && (now - state.lastVoiceAt) >= 1200) {  // 1.2s silence → end (let caller finish)
    state.speaking = false;
    state.speechStartAt = 0;
    onCallerSpeechEnd(state);
  }
}

function onCallerSpeechStart(state) {
  const now = Date.now();
  const isNovaPlaying = !!phoneCurrentSource;
  const playbackAgeMs = phonePlaybackStartedAt ? (now - phonePlaybackStartedAt) : 0;
  // Avoid cancelling Nova instantly on line noise; only allow barge-in
  // once Nova has been audible long enough to be heard.
  if (isNovaPlaying && playbackAgeMs >= 2500) {
    stopCallPlayback();
  }
  try {
    state.chunks = [];
    state.recorder = new MediaRecorder(state.stream, { mimeType: 'audio/webm;codecs=opus' });
    state.recorder.ondataavailable = (e) => { if (e.data && e.data.size) state.chunks.push(e.data); };
    state.recStartAt = Date.now();
    state.recorder.start();
  } catch (e) { console.error('MediaRecorder failed', e); state.recorder = null; }
}

function onCallerSpeechEnd(state) {
  const rec = state.recorder;
  if (!rec) return;
  state.recorder = null;
  const durMs = Date.now() - (state.recStartAt || 0);
  rec.onstop = () => {
    const chunks = state.chunks;
    state.chunks = [];
    if (durMs < 500 || !chunks.length) return;   // too short to be speech
    phoneTurnCounter++; // only advance turns for real utterances we actually send
    const myTurn = phoneTurnCounter;
    const blob = new Blob(chunks, { type: 'audio/webm' });
    const fr = new FileReader();
    fr.onloadend = () => {
      const s = String(fr.result || '');
      const b64 = s.slice(s.indexOf(',') + 1);
      if (b64 && phoneCallWs && phoneCallWs.readyState === WebSocket.OPEN) {
        phoneCallWs.send(JSON.stringify({
          type: 'utterance', turn: myTurn, audio_b64: b64, mime: 'audio/webm',
        }));
      }
    };
    fr.readAsDataURL(blob);
  };
  try { rec.stop(); } catch (_) {}
}

function stopCallVad() {
  const s = phoneVadState;
  if (!s) return;
  if (s.interval) clearInterval(s.interval);
  try { if (s.recorder) s.recorder.stop(); } catch (_) {}
  try { s.srcNode.disconnect(); } catch (_) {}
  phoneVadState = null;
}

// Restore the toggle after a reload; unregister cleanly when the window closes.
window.addEventListener('load', () => {
  try {
    const raw = localStorage.getItem(NOVA_UI_STATE_KEY);
    const st = raw ? JSON.parse(raw) : null;
    if (st && st.answerCalls) setTimeout(() => { enableAnswerCalls(); }, 1200);
  } catch (_) {}
});
window.addEventListener('beforeunload', () => {
  if (answerCallsEnabled) {
    try {
      navigator.sendBeacon('/phone/state',
        new Blob([JSON.stringify({ enabled: false })], { type: 'application/json' }));
    } catch (_) {}
    releasePhoneOwner();
  }
});
/* ══════════════ end 📞 Answer Calls ══════════════ */

function setVoicePillState() {
  voicePill.textContent = voiceEnabled ? '🔊 Voice' : '🔇 Voice Off';
  voicePill.style.opacity = voiceEnabled ? '1' : '0.6';
}

function setWakePillState() {
  if (!wakePill) return;
  wakePill.textContent = wakeModeEnabled ? '👂 Wake On' : '👂 Wake Off';
  wakePill.classList.toggle('active', wakeModeEnabled);
}

function readWakeOwner() {
  try {
    const raw = localStorage.getItem(WAKE_OWNER_KEY);
    if (!raw) return null;
    const parsed = JSON.parse(raw);
    if (!parsed || !parsed.id || !parsed.ts) return null;
    if ((Date.now() - Number(parsed.ts)) > WAKE_OWNER_TTL_MS) return null;
    return parsed;
  } catch (_) {
    return null;
  }
}

function claimWakeOwner(force = false) {
  if (document.hidden || !document.hasFocus()) return false;
  const owner = readWakeOwner();
  if (owner && owner.id !== wakeInstanceId && !force) {
    return false;
  }
  try {
    localStorage.setItem(WAKE_OWNER_KEY, JSON.stringify({ id: wakeInstanceId, ts: Date.now() }));
    return true;
  } catch (_) {
    return true;
  }
}

function isWakeOwner() {
  const owner = readWakeOwner();
  return !owner || owner.id === wakeInstanceId;
}

function releaseWakeOwner() {
  try {
    const owner = readWakeOwner();
    if (owner && owner.id === wakeInstanceId) {
      localStorage.removeItem(WAKE_OWNER_KEY);
    }
  } catch (_) {}
}

function setWakeEchoGuard(ms = 1200) {
  wakeEchoGuardUntil = Date.now() + Math.max(0, ms);
}

function isWakeEchoGuardActive() {
  return Date.now() < wakeEchoGuardUntil;
}

function clearWakeSilenceTimer() {
  if (wakeSilenceTimer) {
    clearTimeout(wakeSilenceTimer);
    wakeSilenceTimer = null;
  }
}

function clearWakeCommitTimer() {
  if (wakeCommitTimer) {
    clearTimeout(wakeCommitTimer);
    wakeCommitTimer = null;
  }
}

function clearWakeRestartTimer() {
  if (wakeRestartTimer) {
    clearTimeout(wakeRestartTimer);
    wakeRestartTimer = null;
  }
}

function stopWakeHealthMonitor() {
  if (wakeHealthTimer) {
    clearInterval(wakeHealthTimer);
    wakeHealthTimer = null;
  }
}

function ensureWakeListenerHealthy() {
  if (!wakeModeEnabled || !wakeRecognition) return;
  if (document.hidden) return;
  if (!isWakeOwner()) {
    if (wakeRecognitionActive) {
      try { wakeRecognition.stop(); } catch (_) {}
      wakeRecognitionActive = false;
    }
    return;
  }
  if (wakePausedForResponse) return;
  if (isWakeEchoGuardActive()) return;
  if (isProcessing || serverAudioPlaying || speakingNow) return;

  // Browser speech engines can go stale over time; soft-refresh the session.
  if (wakeRecognitionActive && (Date.now() - wakeLastStartAt) > WAKE_RESTART_AFTER_MS) {
    try { wakeRecognition.stop(); } catch (_) {}
    wakeRecognitionActive = false;
    requestWakeStart(500);
    return;
  }

  if (!wakeRecognitionActive) {
    requestWakeStart(0);
  }
}

function startWakeHealthMonitor() {
  stopWakeHealthMonitor();
  wakeHealthTimer = setInterval(ensureWakeListenerHealthy, WAKE_HEALTH_CHECK_MS);
}

function requestWakeStart(delayMs = 250) {
  if (!wakeModeEnabled || !wakeRecognition) return;
  clearWakeRestartTimer();
  wakeRestartTimer = setTimeout(() => {
    wakeRestartTimer = null;
    if (!wakeModeEnabled || !wakeRecognition) return;
    if (document.hidden) return;
    if (!claimWakeOwner()) return;
    if (wakePausedForResponse) return;
    if (isWakeEchoGuardActive()) {
      requestWakeStart(250);
      return;
    }
    if (isProcessing || serverAudioPlaying || speakingNow) return;
    if (wakeStartInFlight) return;
    try {
      wakeStartInFlight = true;
      wakeRecognition.start();
    } catch (_) {
      wakeStartInFlight = false;
      // Browser can reject start during transitional states; retry shortly.
      requestWakeStart(700);
    }
  }, Math.max(0, delayMs));
}

function queueWakeConversationSegment(segment) {
  const piece = String(segment || '').trim();
  if (!piece) return;
  wakePendingUtterance = (wakePendingUtterance + ' ' + piece).trim();
  clearWakeCommitTimer();
  wakeCommitTimer = setTimeout(() => {
    const utterance = wakePendingUtterance.trim();
    wakePendingUtterance = '';
    clearWakeCommitTimer();
    if (!utterance) return;
    handleWakeTranscript(utterance);
  }, 1300);
}

function updateWakeListeningStatus() {
  if (!wakeModeEnabled) return;
  if (wakeConversationActive) {
    showConnStatus('Conversation mode: listening (auto-timeout 60s)', false);
  } else {
    showConnStatus('Wake listening on: say "hey nova"', false);
  }
}

async function activateActiveListen() {
  if (activeListenBusy) return;
  activeListenBusy = true;
  if (activeListenPill) {
    activeListenPill.style.opacity = '0.65';
    activeListenPill.textContent = '🎙 Arming...';
  }

  try {
    const res = await fetch('/active-listen/ensure', { method: 'POST' });
    const payload = await res.json();
    if (!res.ok || !payload.ok) {
      const err = (payload && payload.error) ? payload.error : 'Failed to arm active listening service.';
      showConnStatus(err, true);
      return;
    }

    const wakeState = payload.wake_listener_running ? 'running' : 'starting';
    showConnStatus(`Active listening ${wakeState}. Say "hey nova".`, false);
    if (wakeModeEnabled && wakeRecognition && !wakeRecognitionActive && !wakePausedForResponse) {
      requestWakeStart(0);
    }
  } catch (_) {
    showConnStatus('Could not reach Active Listen service endpoint.', true);
  } finally {
    activeListenBusy = false;
    if (activeListenPill) {
      activeListenPill.style.opacity = '1';
      activeListenPill.textContent = '🎙 Active Listen';
    }
  }
}

function startWakeConversation() {
  if (!wakeModeEnabled) return;
  wakeConversationActive = true;
  clearWakeSilenceTimer();
  wakeSilenceTimer = setTimeout(() => {
    wakeConversationActive = false;
    wakeAwaitingCommand = false;
    clearWakeSilenceTimer();
    updateWakeListeningStatus();
  }, WAKE_CONVERSATION_TIMEOUT_MS);
  updateWakeListeningStatus();
}

function endWakeConversation() {
  wakeConversationActive = false;
  wakeAwaitingCommand = false;
  clearWakeSilenceTimer();
  wakePendingUtterance = '';
  clearWakeCommitTimer();
  updateWakeListeningStatus();
}

function isConversationEndPhrase(rawText) {
  const normalized = normalizeSpeech(rawText);
  if (!normalized) return false;
  return WAKE_END_PHRASES.some((phrase) => normalized.includes(normalizeSpeech(phrase)));
}

function isCloseWindowPhrase(rawText) {
  const normalized = normalizeSpeech(rawText);
  if (!normalized) return false;
  return CLOSE_WINDOW_PHRASES.some((phrase) => normalized.includes(normalizeSpeech(phrase)));
}

function closeNovaWindowLocal() {
  stopNovaSpeech();
  endWakeConversation();
  showConnStatus('Closing Nova window...', false);
  setTimeout(() => {
    try { window.close(); } catch (_) {}
  }, 150);
}

function toggleVoice() {
  voiceEnabled = !voiceEnabled;
  setVoicePillState();
  saveNovaUIState();
  if (!voiceEnabled && window.speechSynthesis) {
    window.speechSynthesis.cancel();
    speechQueue = [];
    speakingNow = false;
    setWakeEchoGuard(800);
  }
  if (!voiceEnabled) {
    serverAudioQueue = [];
    if (currentServerAudio) {
      currentServerAudio.pause();
      currentServerAudio = null;
    }
    serverAudioPlaying = false;
    setWakeEchoGuard(800);
  } else {
    processServerAudioQueue();
  }
}

// Chrome silently pauses SpeechSynthesis after ~15s — keep-alive
setInterval(() => {
  if (window.speechSynthesis && window.speechSynthesis.speaking) {
    window.speechSynthesis.pause();
    window.speechSynthesis.resume();
  }
}, 10000);

// Voices load async in Chrome — cache them once available
let _voices = [];
function loadVoices() {
  _voices = window.speechSynthesis ? window.speechSynthesis.getVoices() : [];
}
if (window.speechSynthesis) {
  window.speechSynthesis.onvoiceschanged = loadVoices;
  loadVoices();
}

function getBestVoice() {
  if (!_voices.length) loadVoices();
  const preferred = ['Samantha', 'Google US English', 'Microsoft Zira', 'Microsoft David', 'Alex', 'Karen'];
  for (const name of preferred) {
    const v = _voices.find(v => v.name.includes(name));
    if (v) return v;
  }
  return _voices.find(v => v.lang === 'en-US') || _voices.find(v => v.lang && v.lang.startsWith('en')) || null;
}

// Mark audio as unlocked on first user interaction — no probe needed
function markUnlocked() {
  if (!audioUnlocked) {
    audioUnlocked = true;
    processSpeechQueue();
    processServerAudioQueue();
  }
}

function enqueueServerAudio(audioB64) {
  if (!voiceEnabled || !audioB64) return;
  setWakeEchoGuard(1500);
  serverAudioQueue.push(`data:audio/mpeg;base64,${audioB64}`);
  processServerAudioQueue();
}

function stopNovaSpeech() {
  serverAudioQueue = [];
  if (currentServerAudio) {
    try {
      currentServerAudio.pause();
      currentServerAudio.currentTime = 0;
    } catch (_) {}
    currentServerAudio = null;
  }
  serverAudioPlaying = false;

  speechQueue = [];
  if (window.speechSynthesis) {
    try { window.speechSynthesis.cancel(); } catch (_) {}
  }
  speakingNow = false;
  avatarEl.classList.remove('speaking');
  setWakeEchoGuard(1400);
}

function interruptAndRedirect() {
  markUnlocked();
  stopNovaSpeech();
  setProcessing(false);

  if (wakeModeEnabled) {
    wakePausedForResponse = false;
    wakeAwaitingCommand = true;
    wakeConversationActive = true;
    updateWakeListeningStatus();
    requestWakeStart(300);
    showConnStatus('Nova interrupted. Speak your new command.', false);
    return;
  }

  showConnStatus('Nova interrupted. Press mic and speak your new command.', false);
}

function pauseWakeListeningForResponse() {
  if (!wakeModeEnabled || !wakeRecognition) return;
  wakePausedForResponse = true;
  clearWakeRestartTimer();
  try {
    wakeRecognition.stop();
  } catch (_) {
    // Safe to ignore if already stopped.
  }
}

function maybeResumeWakeListening() {
  if (!wakeModeEnabled || !wakeRecognition) return;
  if (!wakePausedForResponse) return;
  if (isProcessing || serverAudioPlaying || speakingNow) return;
  wakePausedForResponse = false;
  requestWakeStart(450);
}

function processServerAudioQueue() {
  if (!voiceEnabled || !audioUnlocked) return;
  if (serverAudioPlaying || !serverAudioQueue.length) return;

  const src = serverAudioQueue.shift();
  const audio = new Audio(src);
  currentServerAudio = audio;
  serverAudioPlaying = true;

  audio.onended = () => {
    setWakeEchoGuard(800);
    serverAudioPlaying = false;
    currentServerAudio = null;
    processServerAudioQueue();
    maybeResumeWakeListening();
  };
  audio.onerror = () => {
    setWakeEchoGuard(800);
    serverAudioPlaying = false;
    currentServerAudio = null;
    processServerAudioQueue();
    maybeResumeWakeListening();
  };

  const playPromise = audio.play();
  if (playPromise && typeof playPromise.catch === 'function') {
    playPromise.catch(() => {
      serverAudioPlaying = false;
      currentServerAudio = null;
    });
  }
}

function speakText(text) {
  if (!voiceEnabled || !window.speechSynthesis || !text) return;
  setWakeEchoGuard(1200);
  const clean = String(text)
    .replace(/\*\*/g, '')
    .replace(/[*]/g, '')
    .replace(/#+\s*/g, '')
    .replace(/`+/g, '')
    .replace(/\[(.*?)\]\((.*?)\)/g, '$1')
    .replace(/\s+/g, ' ')
    .trim();
  if (!clean) return;
  speechQueue.push(clean);
  processSpeechQueue();
}

function processSpeechQueue() {
  if (!voiceEnabled || !window.speechSynthesis || !audioUnlocked) return;
  if (speakingNow) return;
  if (!speechQueue.length) return;

  speakingNow = true;
  const next = () => {
    if (!speechQueue.length || !voiceEnabled) {
      setWakeEchoGuard(800);
      speakingNow = false;
      avatarEl.classList.remove('speaking');
      maybeResumeWakeListening();
      return;
    }
    const utterance = new SpeechSynthesisUtterance(speechQueue.shift());
    utterance.rate = 1.03;
    utterance.pitch = 1.0;
    const v = getBestVoice();
    if (v) utterance.voice = v;
    utterance.onend = next;
    utterance.onerror = next;
    window.speechSynthesis.speak(utterance);
  };
  next();
}

function updateMicButton() {
  if (!micBtn) return;
  if (recognitionActive) {
    micBtn.classList.add('recording');
    micBtn.textContent = '⏹';
    micBtn.title = 'Stop voice input and send';
  } else {
    micBtn.classList.remove('recording');
    micBtn.textContent = '🎤';
    micBtn.title = 'Start voice input';
  }
}

function updateInterruptButton() {
  if (!interruptBtn) return;
  const active = Boolean(serverAudioPlaying || speakingNow || isProcessing);
  interruptBtn.style.opacity = active ? '1' : '0.65';
  interruptBtn.title = active ? 'Interrupt Nova and redirect now' : 'Interrupt (idle)';
}

function normalizeSpeech(text) {
  return String(text || '')
    .toLowerCase()
    .replace(/[^a-z0-9\s]/g, ' ')
    .replace(/\s+/g, ' ')
    .trim();
}

function extractWakeCommand(rawText) {
  const normalized = normalizeSpeech(rawText);
  // Tolerate common recognizer variants: nova/noah/novaah.
  const wakeRegex = /\b(?:hey|ok|okay)?\s*(nova|noah|novah|noba)\b\s*(.*)$/;
  const match = normalized.match(wakeRegex);
  if (match) {
    const command = (match[2] || '').trim();
    return { heard: true, command };
  }

  for (const phrase of WAKE_PHRASES) {
    const idx = normalized.indexOf(phrase);
    if (idx !== -1) {
      const command = normalized.slice(idx + phrase.length).trim();
      return { heard: true, command };
    }
  }
  return { heard: false, command: '' };
}

function shouldHandleWakeText(rawText) {
  const normalized = normalizeSpeech(rawText);
  if (!normalized) return false;
  const now = Date.now();
  // Drop near-duplicate transcripts that often occur with interim/final updates.
  if (normalized === wakeLastHandledText && (now - wakeLastHandledAt) < 2500) {
    return false;
  }
  wakeLastHandledText = normalized;
  wakeLastHandledAt = now;
  return true;
}

function submitWakeCommand(commandText) {
  const command = String(commandText || '').trim();
  if (!command) return;
  if (isCloseWindowPhrase(command)) {
    appendMessage('user', command);
    appendMessage('nova', 'Closing Nova window. Say "hey nova" any time to bring me back.');
    closeNovaWindowLocal();
    return;
  }
  if (isProcessing) {
    showConnStatus('Nova is busy. Try wake phrase again in a moment.', true);
    return;
  }
  pauseWakeListeningForResponse();
  inputEl.value = command;
  inputEl.dispatchEvent(new Event('input'));
  sendFromInput();
}

function handleWakeTranscript(rawText) {
  const text = String(rawText || '').trim();
  if (!text) return;

  if (isCloseWindowPhrase(text)) {
    appendMessage('user', text);
    appendMessage('nova', 'Closing Nova window. Say "hey nova" any time to bring me back.');
    closeNovaWindowLocal();
    return;
  }

  // Ignore recognitions while Nova is speaking/processing to prevent
  // accidental self-trigger loops from speaker audio bleed.
  if (isProcessing || serverAudioPlaying || speakingNow || isWakeEchoGuardActive()) {
    return;
  }

  showConnStatus(`Heard: "${text}"`, false);

  if (wakeConversationActive) {
    if (isConversationEndPhrase(text)) {
      endWakeConversation();
      return;
    }
    startWakeConversation();
    submitWakeCommand(text);
    return;
  }

  if (wakeAwaitingCommand) {
    if (isConversationEndPhrase(text)) {
      endWakeConversation();
      return;
    }
    const wakeOnly = extractWakeCommand(text);
    if (wakeOnly.heard && !wakeOnly.command) {
      showConnStatus('Wake word heard. Say your command...', false);
      return;
    }
    startWakeConversation();
    // Coalesce multi-part dictation before submit so longer commands are not cut off.
    queueWakeConversationSegment(text);
    return;
  }

  const result = extractWakeCommand(text);
  if (!result.heard) {
    showConnStatus('Listening... say "hey nova" to begin', false);
    return;
  }

  if (result.command) {
    if (isConversationEndPhrase(result.command)) {
      endWakeConversation();
      return;
    }
    startWakeConversation();
    wakeAwaitingCommand = true;
    // First command chunk after wake phrase; additional chunks will be merged.
    queueWakeConversationSegment(result.command);
  } else {
    wakeAwaitingCommand = true;
    showConnStatus('Wake word heard. Say your command...', false);
  }
}

async function toggleWakeMode() {
  markUnlocked();
  if (!SpeechRecognitionCtor) {
    showConnStatus('Wake listening is not supported in this browser.', true);
    return;
  }

  if (wakeModeEnabled) {
    wakeModeEnabled = false;
    releaseWakeOwner();
    endWakeConversation();
    setWakePillState();
    clearWakeRestartTimer();
    stopWakeHealthMonitor();
    wakeStartInFlight = false;
    wakeRecognitionActive = false;
    if (wakeRecognition) {
      try { wakeRecognition.stop(); } catch (_) {}
    }
    saveNovaUIState();
    showConnStatus('', false);
    return;
  }

  try {
    const stream = await navigator.mediaDevices.getUserMedia({ audio: true });
    stream.getTracks().forEach((t) => t.stop());
  } catch (e) {
    showConnStatus('Microphone access blocked. Allow mic in browser site settings.', true);
    return;
  }

  if (recognitionActive && recognition) {
    recognition.stop();
  }

  if (!wakeRecognition) {
    wakeRecognition = new SpeechRecognitionCtor();
    wakeRecognition.lang = 'en-US';
    wakeRecognition.continuous = true;
    wakeRecognition.interimResults = true;
    wakeRecognition.maxAlternatives = 1;

    wakeRecognition.onstart = () => {
      wakeStartInFlight = false;
      wakeRecognitionActive = true;
      wakeNoSpeechCount = 0;
      wakeLastStartAt = Date.now();
      if (wakeModeEnabled) {
        updateWakeListeningStatus();
      }
    };

    wakeRecognition.onresult = (event) => {
      for (let i = event.resultIndex; i < event.results.length; i++) {
        const alt = event.results[i][0];
        const transcript = alt && alt.transcript ? alt.transcript : '';
        if (!transcript) continue;

        const isFinal = Boolean(event.results[i].isFinal);
        const wakeFound = extractWakeCommand(transcript).heard;
        if (wakeConversationActive || wakeAwaitingCommand) {
          // In active conversation, only commit final chunks and coalesce
          // short segments so we submit one complete command.
          if (!isFinal) continue;
          queueWakeConversationSegment(transcript);
          continue;
        }

        const shouldEvaluate = isFinal || wakeFound;
        if (!shouldEvaluate) continue;
        if (!shouldHandleWakeText(transcript)) continue;

        console.log('[Nova wake] transcript:', transcript, '| final:', isFinal);
        handleWakeTranscript(transcript);
      }
    };

    wakeRecognition.onerror = (event) => {
      if (!wakeModeEnabled) return;
      wakeStartInFlight = false;
      wakeRecognitionActive = false;
      const err = (event && event.error) ? event.error : 'unknown';
      if (err === 'not-allowed' || err === 'service-not-allowed') {
        showConnStatus('Wake listening denied by browser mic permission.', true);
        wakeModeEnabled = false;
        stopWakeHealthMonitor();
        setWakePillState();
        return;
      }
      if (err === 'no-speech') {
        wakeNoSpeechCount += 1;
        if (wakeNoSpeechCount >= 3) {
          showConnStatus('Wake mic hears no speech. Check browser microphone input device.', true);
          wakeNoSpeechCount = 0;
        }
      }
      if (err !== 'no-speech' && err !== 'aborted') {
        showConnStatus(`Wake listening error: ${err}`, true);
      }
      if (wakeModeEnabled && !wakePausedForResponse) {
        requestWakeStart(900);
      }
    };

    wakeRecognition.onend = () => {
      if (!wakeModeEnabled) return;
      wakeStartInFlight = false;
      wakeRecognitionActive = false;
      if (wakePausedForResponse) return;
      if (document.hidden) return;
      requestWakeStart(400);
    };
  }

  wakeModeEnabled = true;
  claimWakeOwner(true);
  wakeAwaitingCommand = false;
  wakeConversationActive = false;
  clearWakeSilenceTimer();
  setWakePillState();
  updateWakeListeningStatus();
  wakePausedForResponse = false;
  wakeStartInFlight = false;
  wakeRecognitionActive = false;
  saveNovaUIState();
  startWakeHealthMonitor();
  if (document.hidden) {
    showConnStatus('Wake listening armed. Focus this Nova window to listen.', false);
  } else {
    requestWakeStart(0);
  }
}

function initSpeechInput() {
  if (!SpeechRecognitionCtor) {
    if (micBtn) {
      micBtn.disabled = true;
      micBtn.title = 'Voice input is not supported in this browser';
      micBtn.style.opacity = '0.45';
    }
    if (wakePill) {
      wakePill.disabled = true;
      wakePill.title = 'Wake listening is not supported in this browser';
      wakePill.style.opacity = '0.45';
    }
    return;
  }

  recognition = new SpeechRecognitionCtor();
  recognition.lang = 'en-US';
  recognition.continuous = false;
  recognition.interimResults = true;
  recognition.maxAlternatives = 1;

  recognition.onstart = () => {
    recognitionActive = true;
    updateMicButton();
    showConnStatus('Listening... speak now', true);
  };

  recognition.onresult = (event) => {
    let interim = '';
    for (let i = event.resultIndex; i < event.results.length; i++) {
      const alt = event.results[i] && event.results[i][0] ? event.results[i][0] : null;
      const part = alt && alt.transcript ? alt.transcript : '';
      if (event.results[i].isFinal) {
        speechFinalText += part + ' ';
      } else {
        interim += part;
      }
    }
    speechInterimText = interim;
    inputEl.value = (speechDraftPrefix + ' ' + speechFinalText + speechInterimText).trim();
    inputEl.dispatchEvent(new Event('input'));
  };

  recognition.onend = () => {
    const captured = (speechFinalText + ' ' + speechInterimText).trim();
    const hadText = Boolean(captured);
    recognitionActive = false;
    updateMicButton();
    if (hadText) {
      inputEl.value = (speechDraftPrefix + ' ' + captured).trim();
      inputEl.dispatchEvent(new Event('input'));
      if (!isProcessing) {
        sendFromInput();
      }
      showConnStatus('', false);
    } else {
      showConnStatus('No speech detected. Check mic permission and input device.', true);
    }
    speechFinalText = '';
    speechInterimText = '';
    speechDraftPrefix = '';
  };

  recognition.onerror = (event) => {
    recognitionActive = false;
    updateMicButton();
    const err = (event && event.error) ? event.error : 'unknown';
    if (err === 'not-allowed' || err === 'service-not-allowed') {
      showConnStatus('Mic permission denied by browser.', true);
    } else if (err === 'no-speech') {
      showConnStatus('No speech heard. Try speaking louder/closer.', true);
    } else if (err === 'audio-capture') {
      showConnStatus('No microphone available for this browser.', true);
    } else {
      showConnStatus(`Mic input error: ${err}`, true);
    }
    speechFinalText = '';
    speechInterimText = '';
    speechDraftPrefix = '';
  };
}

async function toggleMicInput() {
  markUnlocked();
  if (serverAudioPlaying || speakingNow) {
    stopNovaSpeech();
    if (wakeModeEnabled) {
      wakePausedForResponse = false;
      wakeAwaitingCommand = true;
      wakeConversationActive = true;
      updateWakeListeningStatus();
      requestWakeStart(300);
      showConnStatus('Nova interrupted. Speak your command now.', false);
      return;
    }
  }
  if (wakeModeEnabled) {
    showConnStatus('Disable Wake mode to use push-to-talk mic.', true);
    return;
  }
  if (!recognition) {
    showConnStatus('Voice input not supported in this browser', true);
    return;
  }
  if (isProcessing) return;

  if (recognitionActive) {
    recognition.stop();
    return;
  }

  try {
    // Prompt/validate microphone access before starting recognition.
    const stream = await navigator.mediaDevices.getUserMedia({ audio: true });
    stream.getTracks().forEach((t) => t.stop());
  } catch (e) {
    showConnStatus('Microphone access blocked. Allow mic in browser site settings.', true);
    return;
  }

  speechDraftPrefix = inputEl.value ? inputEl.value.trim() : '';
  speechFinalText = '';
  speechInterimText = '';
  try {
    recognition.start();
  } catch (e) {
    showConnStatus('Could not start microphone', true);
  }
}

// ── WebSocket connection ───────────────────────────────────────────────
let retryCount = 0;
const BASE_WINDOW_TITLE = 'NOVA Assistant - Central Pharmacy Group';

function setWindowTitle(state = '') {
  document.title = state ? `NOVA ${state} | Central Pharmacy Group` : BASE_WINDOW_TITLE;
}

function connect() {
  ws = new WebSocket(`ws://${location.host}/ws`);

  ws.onopen = () => {
    document.getElementById('conn-status').className = '';
    retryCount = 0;
    setWindowTitle('Online');
    setVoicePillState();
    startBillingReminderScheduler();
    console.log('Connected to Nova');
  };

  ws.onmessage = (e) => {
    try {
      const data = JSON.parse(e.data);
      handleMessage(data);
    } catch(err) {
      console.error('Parse error:', err);
    }
  };

  ws.onclose = () => {
    retryCount++;
    setWindowTitle(`Reconnecting ${retryCount}`);
    const delay = Math.min(1000 * retryCount, 5000);
    showConnStatus(`Connecting... (${retryCount})`, true);
    setTimeout(connect, delay);
  };

  ws.onerror = (e) => {
    console.log('WS error, will retry');
  };
}

function getTodayKey() {
  const d = new Date();
  const yyyy = d.getFullYear();
  const mm = String(d.getMonth() + 1).padStart(2, '0');
  const dd = String(d.getDate()).padStart(2, '0');
  return `${yyyy}-${mm}-${dd}`;
}

function getBillingReminderSlots() {
  try {
    const raw = localStorage.getItem(BILLING_REMINDER_STORAGE_KEY);
    const parsed = raw ? JSON.parse(raw) : null;
    if (!parsed || parsed.day !== getTodayKey() || !Array.isArray(parsed.hours)) {
      return [];
    }
    return parsed.hours
      .map((h) => Number(h))
      .filter((h) => Number.isInteger(h));
  } catch (_) {
    return [];
  }
}

function saveBillingReminderSlots(hours) {
  try {
    localStorage.setItem(
      BILLING_REMINDER_STORAGE_KEY,
      JSON.stringify({ day: getTodayKey(), hours: hours })
    );
  } catch (_) {}
}

function markBillingReminderHour(hour) {
  const current = new Set(getBillingReminderSlots());
  current.add(Number(hour));
  saveBillingReminderSlots(Array.from(current.values()).sort((a, b) => a - b));
}

function hasBillingReminderHour(hour) {
  return getBillingReminderSlots().includes(Number(hour));
}

function currentPageState() {
  const visibility = document.visibilityState || 'unknown';
  const hasFocus = typeof document.hasFocus === 'function' ? document.hasFocus() : false;
  if (visibility === 'visible' && hasFocus) return 'foreground';
  if (visibility === 'hidden') return 'hidden';
  return 'background';
}

function requestBillingReminderCheck(hour) {
  if (!ws || ws.readyState !== WebSocket.OPEN) return;
  ws.send(JSON.stringify({
    type: 'billing_check',
    schedule_hour: Number(hour),
    page_state: currentPageState(),
    visibility_state: document.visibilityState || 'unknown',
    has_focus: typeof document.hasFocus === 'function' ? document.hasFocus() : false,
  }));
}

function runBillingReminderScheduleTick() {
  const now = new Date();
  const hour = now.getHours();
  if (!BILLING_REMINDER_HOURS.includes(hour)) return;
  if (hasBillingReminderHour(hour)) return;
  markBillingReminderHour(hour);
  requestBillingReminderCheck(hour);
}

function startBillingReminderScheduler() {
  if (billingReminderTimer) {
    clearInterval(billingReminderTimer);
  }
  runBillingReminderScheduleTick();
  billingReminderTimer = setInterval(runBillingReminderScheduleTick, 60 * 1000);
}

function showConnStatus(msg, isErr) {
  const el = document.getElementById('conn-status');
  el.textContent = msg || '';
  if (!msg) {
    el.className = '';
    return;
  }
  el.className = isErr ? 'error' : 'show';
}

function friendlyModelName(raw) {
  if (!raw) return '';
  const s = raw.toLowerCase();
  if (s.includes('opus'))   return 'Opus';
  if (s.includes('sonnet')) return 'Sonnet';
  if (s.includes('haiku'))  return 'Haiku';
  return raw.split('-').slice(0, 3).join('-');
}

function updateModelPill(chatModel, visionModel) {
  const pill = document.getElementById('model-pill');
  if (!pill) return;
  const chat = (chatModel || '').trim();
  const vision = (visionModel || '').trim();
  if (!chat && !vision) return;
  const chatName = friendlyModelName(chat);
  const visionName = friendlyModelName(vision);
  pill.textContent = chatName || visionName;

  // Highlight pill when using a stronger model
  const active = (chatName || '').toLowerCase();
  if (active === 'sonnet' || active === 'opus') {
    pill.style.background = 'rgba(0,153,255,0.2)';
    pill.style.borderColor = 'rgba(0,153,255,0.4)';
    pill.style.color = '#4da6ff';
  } else {
    pill.style.background = '';
    pill.style.borderColor = '';
    pill.style.color = '';
  }

  if (chat && modelSelect) {
    if (![...modelSelect.options].some((o) => o.value === chat)) {
      const opt = document.createElement('option');
      opt.value = chat;
      opt.textContent = chat;
      modelSelect.appendChild(opt);
    }
    modelSelect.value = chat;
  }

  if (vision && visionModelSelect) {
    if (![...visionModelSelect.options].some((o) => o.value === vision)) {
      const opt = document.createElement('option');
      opt.value = vision;
      opt.textContent = vision;
      visionModelSelect.appendChild(opt);
    }
    visionModelSelect.value = vision;
  }
}

function sendModelCommand(cmd) {
  if (!ws || ws.readyState !== WebSocket.OPEN) {
    showConnStatus('Not connected', true);
    return;
  }
  ws.send(JSON.stringify({ type: 'message', text: cmd }));
}

function refreshModelSettings() {
  sendModelCommand('/show-models');
}

function applyModelSettings() {
  const chat = modelSelect ? modelSelect.value.trim() : '';
  const vision = visionModelSelect ? visionModelSelect.value.trim() : '';
  if (!chat && !vision) {
    showConnStatus('Select at least one model first.', true);
    return;
  }
  if (chat) sendModelCommand('/model ' + chat);
  if (vision) sendModelCommand('/vision-model ' + vision);
  showConnStatus('Model update sent.', false);
}

// ── Message handling ───────────────────────────────────────────────────
function handleMessage(data) {
  if (data.type === 'tool_call') {
    appendToolCall(data.tool);
    return;
  }

  if (data.type === 'nova_audio') {
    if (data.audio_b64) {
      enqueueServerAudio(data.audio_b64);
    }
    return;
  }

  if (data.type === 'nova_response') {
    removeTyping();
    appendMessage('nova', data.text);
    updateModelPill(data.model, data.vision_model);
    if (data.audio_b64) {
      enqueueServerAudio(data.audio_b64);
    } else if (voiceEnabled && !data.audio_pending) {
      showConnStatus('Voice reply unavailable: ElevenLabs audio was not generated.', true);
    }
    setProcessing(false);
    avatarEl.classList.remove('speaking');
    if (data.alert) showAlert(data.alert);
    if (data.reminders) updateReminders(data.reminders);
    return;
  }

  if (data.type === 'startup') {
    window.NOVA_AUTOSEND_PENDING_ON_STARTUP = (data.autosend_pending_on_startup === true);
    updateModelPill(data.model, data.vision_model);
    if (data.alert) showAlert(data.alert);
    updateReminders(data.reminders || []);
    if (data.greeting && !greeted) {
      appendMessage('nova', data.greeting);
      if (data.audio_b64) {
        enqueueServerAudio(data.audio_b64);
      } else if (voiceEnabled && !data.audio_pending) {
        showConnStatus('Startup voice unavailable: ElevenLabs audio was not generated.', true);
      }
      greeted = true;
    }
    // Startup should stay local-only unless explicit opt-in is enabled server-side.
    if (!window._pendingChecked) {
      window._pendingChecked = true;
      if (window.NOVA_AUTOSEND_PENDING_ON_STARTUP === true) {
        fetch('/pending-message').then(r => r.json()).then(d => {
          if (d.text) {
            setTimeout(() => { appendMessage('user', d.text); sendMsg(d.text); }, 1000);
          }
        }).catch(() => {});
      }
    }
    return;
  }

  if (data.type === 'reminders_updated') {
    if (data.updated_id !== undefined && data.updated_id !== null) {
      const updatedId = Number(data.updated_id);
      if (!Number.isNaN(updatedId)) {
        reminderEditState.delete(updatedId);
      }
    }
    if (Array.isArray(data.deleted_ids)) {
      data.deleted_ids.forEach((rid) => {
        const n = Number(rid);
        if (!Number.isNaN(n)) {
          reminderSelection.delete(n);
          reminderEditState.delete(n);
        }
      });
    }
    if (data.completed_id !== undefined && data.completed_id !== null) {
      const completedId = Number(data.completed_id);
      if (!Number.isNaN(completedId)) {
        reminderSelection.delete(completedId);
        reminderEditState.delete(completedId);
      }
    }
    updateReminders(data.reminders || []);
    if (data.text) {
      showConnStatus(data.text, false);
    }
    return;
  }

  if (data.type === 'wake_trigger') {
    // Desktop wake listener triggered — enter conversation mode immediately.
    if (document.hidden || !document.hasFocus()) {
      return;
    }
    if (!wakeModeEnabled) {
      toggleWakeMode();
    } else if (!wakeConversationActive) {
      startWakeConversation();
      wakeAwaitingCommand = true;
      updateWakeListeningStatus();
    }
    return;
  }

  if (data.type === 'incoming_call_alert') {
    if (data.text) {
      showAlert(data.text);
      appendMessage('nova', data.text);
    }
    return;
  }

  if (data.type === 'reminder_due') {
    if (data.text) {
      showAlert(data.text);
      appendMessage('nova', data.text);
    }
    return;
  }

  if (data.type === 'error') {
    removeTyping();
    appendMessage('nova', '⚠️ ' + data.text);
    setProcessing(false);
    avatarEl.classList.remove('speaking');
    return;
  }
}

function appendMessage(role, text, imageDataUrl = null) {
  const wrap = document.createElement('div');
  wrap.className = `message ${role}`;

  const avatarDiv = document.createElement('div');
  avatarDiv.className = 'msg-avatar';
  avatarDiv.textContent = role === 'nova' ? '🤖' : '👤';

  const bubble = document.createElement('div');
  bubble.className = 'msg-bubble';
  if (imageDataUrl) {
    const img = document.createElement('img');
    img.src = imageDataUrl;
    img.style = 'max-width:250px;border-radius:8px;display:block;margin-bottom:6px';
    bubble.appendChild(img);
  }
  const textNode = document.createElement('div');
  textNode.innerHTML = formatText(text);
  bubble.appendChild(textNode);

  const time = document.createElement('div');
  time.className = 'msg-time';
  time.textContent = new Date().toLocaleTimeString([], {hour:'2-digit', minute:'2-digit'});

  const inner = document.createElement('div');
  inner.appendChild(bubble);
  inner.appendChild(time);

  wrap.appendChild(avatarDiv);
  wrap.appendChild(inner);
  messagesEl.appendChild(wrap);
  scrollBottom();
}

function formatText(text) {
  return text
    .replace(/[*][*](.+?)[*][*]/g, '<strong>$1</strong>')
    .replace(/[*](.+?)[*]/g, '<em>$1</em>')
    .replace(/[`](.+?)[`]/g, `<code style="font-family:var(--mono);font-size:12px;background:var(--surface2);padding:1px 5px;border-radius:4px">$1</code>`)
    .replace(/\\n/g, '<br>');
}

function appendToolCall(tool) {
  const el = document.createElement('div');
  el.className = 'tool-call';
  el.textContent = `⚙ ${tool}...`;
  messagesEl.appendChild(el);
  scrollBottom();
}

function showTyping() {
  removeTyping();
  const wrap = document.createElement('div');
  wrap.className = 'typing-indicator';
  wrap.id = 'typing';

  const av = document.createElement('div');
  av.className = 'msg-avatar';
  av.style.background = 'linear-gradient(135deg, var(--accent), var(--accent2))';
  av.textContent = '🤖';

  const dots = document.createElement('div');
  dots.className = 'typing-dots';
  dots.innerHTML = '<span></span><span></span><span></span>';

  wrap.appendChild(av);
  wrap.appendChild(dots);
  messagesEl.appendChild(wrap);
  scrollBottom();
  typingEl = wrap;
}

function removeTyping() {
  if (typingEl) { typingEl.remove(); typingEl = null; }
  // Remove tool call indicators
  document.querySelectorAll('.tool-call').forEach(e => e.remove());
}

function scrollBottom() {
  messagesEl.scrollTop = messagesEl.scrollHeight;
}

function setProcessing(val) {
  isProcessing = val;
  sendBtn.disabled = val;
  inputEl.disabled = val;
  if (val && wakeModeEnabled) {
    pauseWakeListeningForResponse();
  }
  if (!val && wakeModeEnabled) {
    updateWakeListeningStatus();
    maybeResumeWakeListening();
  }
  updateInterruptButton();
}

// ── Sending messages ───────────────────────────────────────────────────
function sendFromInput() {
  const text = inputEl.value.trim();
  if ((!text && pendingAttachments.length === 0) || isProcessing) return;

  if (pendingAttachments.length === 0 && isCloseWindowPhrase(text)) {
    appendMessage('user', text);
    appendMessage('nova', 'Closing Nova window. Say "hey nova" any time to bring me back.');
    inputEl.value = '';
    inputEl.style.height = 'auto';
    closeNovaWindowLocal();
    return;
  }

  sendMsg(text, pendingAttachments);
  inputEl.value = '';
  inputEl.style.height = 'auto';
  clearImage();
}

function sendMsg(text, attachments = []) {
  if (!ws || ws.readyState !== WebSocket.OPEN) {
    showConnStatus('Not connected', true);
    return;
  }
  markUnlocked();
  const image = attachments && attachments.length ? attachments[0] : null;
  const isPdf = image && image.mimeType === 'application/pdf';
  const isExcel = image && (
    image.mimeType === 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ||
    image.mimeType === 'application/vnd.ms-excel'
  );
  const displayText = text || (
    attachments.length > 1
      ? `📎 ${attachments.length} attached files`
      : (isPdf ? '📄 PDF document' : (isExcel ? '📊 Excel file' : '📷 Image'))
  );
  const previewDataUrl = (attachments.length === 1 && image && !isPdf && !isExcel) ? image.dataUrl : null;
  appendMessage('user', displayText, previewDataUrl);
  showTyping();
  setProcessing(true);
  avatarEl.classList.add('speaking');
  const payload = {type: 'message', text: text || 'Please analyze this image.'};
  if (attachments.length) {
    payload.attachments = attachments.map((a) => ({
      image: a.dataUrl,
      image_type: a.mimeType,
      attachment_name: a.fileName || '',
    }));
    // Backward compatibility for older backends that only accept one attachment
    payload.image = attachments[0].dataUrl;
    payload.image_type = attachments[0].mimeType;
    payload.attachment_name = attachments[0].fileName || '';
  }
  ws.send(JSON.stringify(payload));
}

// ── Sidebar ────────────────────────────────────────────────────────────
function showAlert(text) {
  const bar = document.getElementById('alert-bar');
  document.getElementById('alert-text').textContent = text;
  bar.classList.add('show');
  setTimeout(() => bar.classList.remove('show'), 8000);
}

function updateReminders(reminders) {
  const el = document.getElementById('reminders-list');
  remindersById = new Map((reminders || []).map((r) => [Number(r.id), r]));
  const currentIds = new Set((reminders || []).map((r) => Number(r.id)).filter((id) => !Number.isNaN(id)));
  reminderSelection = new Set([...reminderSelection].filter((id) => currentIds.has(id)));
  reminderEditState = new Set([...reminderEditState].filter((id) => currentIds.has(id)));

  if (!reminders || reminders.length === 0) {
    reminderSelection.clear();
    reminderEditState.clear();
    el.innerHTML = '<div style="color:var(--text-dim);font-size:12px">No active reminders</div>';
    return;
  }

  el.innerHTML = reminders.map((r) => {
    const id = Number(r.id);
    const safeTag = escHtml(String(r.tag || 'general'));
    if (reminderEditState.has(id)) {
      return renderReminderEditRow(r);
    }
    return `
      <div class="reminder-item" data-reminder-id="${id}">
        <input class="reminder-check" type="checkbox" ${reminderSelection.has(id) ? 'checked' : ''} onchange="toggleReminderSelection(${id}, this.checked)">
        <span class="reminder-tag tag-${safeTag}">${safeTag}</span>
        <span class="reminder-text">
          ${escHtml(String(r.content || ''))}
          ${renderReminderMeta(r)}
        </span>
        <div class="reminder-row-actions">
          <button type="button" onclick="startEditReminder(${id})">Edit</button>
          <button type="button" onclick="completeReminder(${id})">Done</button>
          <button type="button" class="danger" onclick="deleteReminder(${id})">Delete</button>
        </div>
      </div>
    `;
  }).join('');
}

function escHtml(value) {
  return String(value || '')
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;')
    .replace(/'/g, '&#39;');
}

function toLocalInputValue(isoValue) {
  const raw = String(isoValue || '').trim();
  if (!raw) return '';
  const d = new Date(raw);
  if (Number.isNaN(d.getTime())) return '';
  const pad = (n) => String(n).padStart(2, '0');
  return `${d.getFullYear()}-${pad(d.getMonth() + 1)}-${pad(d.getDate())}T${pad(d.getHours())}:${pad(d.getMinutes())}`;
}

function toIsoOrNull(localValue) {
  const v = String(localValue || '').trim();
  if (!v) return null;
  const d = new Date(v);
  if (Number.isNaN(d.getTime())) return v;
  return d.toISOString();
}

function renderReminderMeta(r) {
  const parts = [];
  if (r.due_at) parts.push(`Due: ${escHtml(String(r.due_at))}`);
  if (r.remind_every_minutes) parts.push(`Every ${escHtml(String(r.remind_every_minutes))}m`);
  if (r.follow_up_at) parts.push(`Follow-up: ${escHtml(String(r.follow_up_at))}`);
  if (r.follow_up_notes) parts.push(`Notes: ${escHtml(String(r.follow_up_notes))}`);
  if (!parts.length) return '';
  return `<span class="reminder-meta">${parts.join(' | ')}</span>`;
}

function renderReminderEditRow(r) {
  const id = Number(r.id);
  const content = escHtml(String(r.content || ''));
  const tag = String(r.tag || 'general');
  const dueAt = escHtml(toLocalInputValue(r.due_at));
  const followUpAt = escHtml(toLocalInputValue(r.follow_up_at));
  const followUpNotes = escHtml(String(r.follow_up_notes || ''));
  const cadence = Number(r.remind_every_minutes || 30);
  const tagOptions = ['general', 'ordering', 'calls', 'billing', 'follow_up', 'task']
    .map((opt) => `<option value="${opt}" ${tag === opt ? 'selected' : ''}>${opt}</option>`)
    .join('');

  return `
    <div class="reminder-item" data-reminder-id="${id}">
      <input class="reminder-check" type="checkbox" ${reminderSelection.has(id) ? 'checked' : ''} onchange="toggleReminderSelection(${id}, this.checked)">
      <div class="reminder-edit">
        <textarea id="rem-content-${id}" placeholder="Reminder content">${content}</textarea>
        <div class="reminder-edit-row">
          <select id="rem-tag-${id}">${tagOptions}</select>
          <input id="rem-cadence-${id}" type="number" min="1" max="1440" value="${Number.isFinite(cadence) ? cadence : 30}">
        </div>
        <div class="reminder-edit-row">
          <input id="rem-due-${id}" type="datetime-local" value="${dueAt}">
          <input id="rem-followup-at-${id}" type="datetime-local" value="${followUpAt}">
        </div>
        <textarea id="rem-followup-notes-${id}" placeholder="Follow-up notes">${followUpNotes}</textarea>
      </div>
      <div class="reminder-row-actions">
        <button type="button" onclick="saveEditReminder(${id})">Save</button>
        <button type="button" onclick="cancelEditReminder(${id})">Cancel</button>
      </div>
    </div>
  `;
}

function startEditReminder(id) {
  const n = Number(id);
  if (Number.isNaN(n) || !remindersById.has(n)) return;
  reminderEditState.add(n);
  updateReminders(Array.from(remindersById.values()));
}

function cancelEditReminder(id) {
  const n = Number(id);
  if (Number.isNaN(n)) return;
  reminderEditState.delete(n);
  updateReminders(Array.from(remindersById.values()));
}

function saveEditReminder(id) {
  if (!ws || ws.readyState !== WebSocket.OPEN) {
    showConnStatus('Not connected', true);
    return;
  }
  const n = Number(id);
  if (Number.isNaN(n)) return;
  const contentEl = document.getElementById(`rem-content-${n}`);
  const tagEl = document.getElementById(`rem-tag-${n}`);
  const dueEl = document.getElementById(`rem-due-${n}`);
  const cadenceEl = document.getElementById(`rem-cadence-${n}`);
  const followUpAtEl = document.getElementById(`rem-followup-at-${n}`);
  const followUpNotesEl = document.getElementById(`rem-followup-notes-${n}`);
  const content = String(contentEl?.value || '').trim();
  if (!content) {
    showConnStatus('Reminder content is required.', true);
    return;
  }
  ws.send(JSON.stringify({
    type: 'update_reminder',
    reminder_id: n,
    content,
    tag: String(tagEl?.value || 'general').trim() || 'general',
    due_at: toIsoOrNull(dueEl?.value),
    remind_every_minutes: Number(cadenceEl?.value || 30),
    follow_up_at: toIsoOrNull(followUpAtEl?.value),
    follow_up_notes: String(followUpNotesEl?.value || '').trim() || null,
  }));
}

function completeReminder(id) {
  if (!ws || ws.readyState !== WebSocket.OPEN) {
    showConnStatus('Not connected', true);
    return;
  }
  const n = Number(id);
  if (Number.isNaN(n)) return;
  ws.send(JSON.stringify({ type: 'complete_reminder', reminder_id: n }));
}

function deleteReminder(id) {
  if (!ws || ws.readyState !== WebSocket.OPEN) {
    showConnStatus('Not connected', true);
    return;
  }
  const n = Number(id);
  if (Number.isNaN(n)) return;
  ws.send(JSON.stringify({ type: 'delete_reminder', reminder_id: n }));
}

function toggleReminderSelection(id, checked) {
  const n = Number(id);
  if (Number.isNaN(n)) return;
  if (checked) reminderSelection.add(n);
  else reminderSelection.delete(n);
}

function toggleAllReminders(selectAll) {
  const checks = document.querySelectorAll('.reminder-check');
  checks.forEach((cb) => {
    cb.checked = !!selectAll;
    const row = cb.closest('.reminder-item');
    if (!row) return;
    const onchangeAttr = cb.getAttribute('onchange') || '';
    const m = onchangeAttr.match(/toggleReminderSelection\((\d+),/);
    if (!m) return;
    const id = Number(m[1]);
    if (Number.isNaN(id)) return;
    if (selectAll) reminderSelection.add(id);
    else reminderSelection.delete(id);
  });
}

function deleteSelectedReminders() {
  if (!ws || ws.readyState !== WebSocket.OPEN) {
    showConnStatus('Not connected', true);
    return;
  }
  const ids = [...reminderSelection];
  if (!ids.length) {
    showConnStatus('Select reminder(s) first.', true);
    return;
  }
  ws.send(JSON.stringify({ type: 'delete_reminders', reminder_ids: ids }));
}

// ── Image handling ─────────────────────────────────────────────────────
let pendingAttachments = [];  // [{ dataUrl, mimeType, fileName }]

function handleFileSelect(event) {
  const files = Array.from(event.target.files || []);
  if (!files.length) return;
  files.forEach(loadImageFile);
  event.target.value = '';
}

function loadImageFile(file) {
  const lowerName = (file.name || '').toLowerCase();
  const mimeType = file.type ||
    (lowerName.endsWith('.pdf') ? 'application/pdf' :
     (lowerName.endsWith('.xlsx') ? 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' :
      (lowerName.endsWith('.xls') ? 'application/vnd.ms-excel' : 'image/png')));
  const reader = new FileReader();
  reader.onload = (e) => {
    pendingAttachments.push({ dataUrl: e.target.result, mimeType, fileName: file.name || '' });
    renderAttachmentPreview();
  };
  reader.readAsDataURL(file);
}

function renderAttachmentPreview() {
  if (!pendingAttachments.length) {
    clearImage();
    return;
  }

  if (pendingAttachments.length === 1) {
    const only = pendingAttachments[0];
    showImagePreview(only.dataUrl, only.mimeType, only.fileName);
    return;
  }

  const wrap = document.getElementById('img-preview-wrap');
  const imgEl = document.getElementById('img-preview');
  imgEl.style.display = 'none';
  const names = pendingAttachments.map((a) => a.fileName || 'attachment').join(', ');
  let pdfLabel = document.getElementById('pdf-label');
  if (!pdfLabel) {
    pdfLabel = document.createElement('div');
    pdfLabel.id = 'pdf-label';
    pdfLabel.style = 'font-size:12px;color:var(--accent);padding:6px 0;display:flex;align-items:center;gap:6px;flex-wrap:wrap';
    wrap.insertBefore(pdfLabel, wrap.firstChild);
  }
  pdfLabel.textContent = `📎 ${pendingAttachments.length} files: ${names}`;
  wrap.style.display = 'block';
}

function showImagePreview(dataUrl, mimeType, fileName) {
  const wrap = document.getElementById('img-preview-wrap');
  const imgEl = document.getElementById('img-preview');
  const isPdf = mimeType === 'application/pdf';
  const isExcel = mimeType === 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' || mimeType === 'application/vnd.ms-excel';
  if (isPdf || isExcel) {
    imgEl.style.display = 'none';
    let pdfLabel = document.getElementById('pdf-label');
    if (!pdfLabel) {
      pdfLabel = document.createElement('div');
      pdfLabel.id = 'pdf-label';
      pdfLabel.style = 'font-size:12px;color:var(--accent);padding:6px 0;display:flex;align-items:center;gap:6px';
      wrap.insertBefore(pdfLabel, wrap.firstChild);
    }
    pdfLabel.textContent = (isExcel ? '📊 ' : '📄 ') + (fileName || (isExcel ? 'spreadsheet.xlsx' : 'document.pdf'));
  } else {
    imgEl.style.display = '';
    imgEl.src = dataUrl;
    const pdfLabel = document.getElementById('pdf-label');
    if (pdfLabel) pdfLabel.remove();
  }
  wrap.style.display = 'block';
}

function clearImage() {
  pendingAttachments = [];
  document.getElementById('img-preview-wrap').style.display = 'none';
  document.getElementById('img-preview').src = '';
  const pdfLabel = document.getElementById('pdf-label');
  if (pdfLabel) pdfLabel.remove();
  document.getElementById('img-preview').style.display = '';
}

// Ctrl+V paste from clipboard
document.addEventListener('paste', (e) => {
  const items = e.clipboardData?.items;
  if (!items) return;
  let added = false;
  for (const item of items) {
    if (item.type.startsWith('image/')) {
      e.preventDefault();
      const file = item.getAsFile();
      if (file) {
        loadImageFile(file);
        added = true;
      }
    }
  }
  if (added) renderAttachmentPreview();
});

// ── Input auto-resize + Enter to send ─────────────────────────────────
inputEl.addEventListener('input', () => {
  inputEl.style.height = 'auto';
  inputEl.style.height = Math.min(inputEl.scrollHeight, 120) + 'px';
});

inputEl.addEventListener('keydown', (e) => {
  if (e.key === 'Enter' && !e.shiftKey) {
    e.preventDefault();
    markUnlocked();
    sendFromInput();
  }
});

document.addEventListener('pointerdown', markUnlocked, { once: true });
document.addEventListener('keydown', markUnlocked, { once: true });

document.addEventListener('visibilitychange', () => {
  if (!wakeModeEnabled || !wakeRecognition) return;
  if (document.hidden) {
    releaseWakeOwner();
    try { wakeRecognition.stop(); } catch (_) {}
    wakeRecognitionActive = false;
    return;
  }
  if (!wakePausedForResponse) {
    requestWakeStart(250);
  }
});

window.addEventListener('focus', () => {
  if (!wakeModeEnabled || !wakeRecognition || wakePausedForResponse) return;
  claimWakeOwner(true);
  requestWakeStart(150);
});

window.addEventListener('blur', () => {
  releaseWakeOwner();
  if (!wakeModeEnabled || !wakeRecognition) return;
  try { wakeRecognition.stop(); } catch (_) {}
  wakeRecognitionActive = false;
});

window.addEventListener('beforeunload', () => {
  releaseWakeOwner();
});

setInterval(updateInterruptButton, 250);

document.addEventListener('keydown', (e) => {
  if (e.key === 'Escape') {
    interruptAndRedirect();
  }
});

// ── Init ───────────────────────────────────────────────────────────────
const _savedState = loadNovaUIState();
setVoicePillState();
setWakePillState();
initSpeechInput();
connect();
// Auto-enable wake mode if it was on when the user last closed the page
if (_savedState && _savedState.wakeModeEnabled && SpeechRecognitionCtor) {
  // Delay slightly so WebSocket and recognition objects are ready
  setTimeout(() => { if (!wakeModeEnabled) toggleWakeMode(); }, 800);
}
</script>
</body>
</html>
"""

# ── WebSocket handler ─────────────────────────────────────────────────
_pending_message: dict = {}  # {"text": str} or {}

@app.post("/wake_trigger")
async def wake_trigger():
    """Called by the desktop wake listener to activate Nova in any open browser."""
    notified = 0
    dead = []
    for ws in list(_active_websockets):
        try:
            await ws.send_json({"type": "wake_trigger"})
            notified += 1
        except Exception:
            dead.append(ws)
    for ws in dead:
        try:
            _active_websockets.remove(ws)
        except ValueError:
            pass
    return {"clients_notified": notified, "ok": True}


def _wake_listener_running_count() -> int:
    try:
        result = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-Command",
                "(Get-CimInstance Win32_Process | "
                "Where-Object { $_.Name -match 'powershell' -and $_.CommandLine -like '*nova_wake_listener.ps1*' } | "
                "Measure-Object).Count",
            ],
            capture_output=True,
            text=True,
            timeout=6,
        )
        if result.returncode != 0:
            return 0
        return int((result.stdout or "0").strip() or "0")
    except Exception:
        return 0


@app.post("/active-listen/ensure")
async def ensure_active_listen():
    """Best-effort re-arm of Nova background hosts and wake listener."""
    errors: list[str] = []

    try:
        from dmelogic.paths import get_project_root
        from dmelogic.services.nova_background import ensure_nova_background_services

        ensure_nova_background_services(get_project_root(), enabled=True)
    except Exception as e:
        errors.append(f"background_hosts: {e}")

    try:
        from dmelogic.services.nova_wake_listener import ensure_nova_wake_listener

        ensure_nova_wake_listener(enabled=True)
    except Exception as e:
        errors.append(f"wake_listener: {e}")

    count = _wake_listener_running_count()
    ok = count > 0 and not errors
    return {
        "ok": ok,
        "wake_listener_running": count > 0,
        "wake_listener_count": count,
        "errors": errors,
        "error": "; ".join(errors) if errors else None,
    }

@app.post("/pending-message")
async def set_pending_message(request: Request):
    body = await request.json()
    _pending_message["text"] = body.get("text", "")
    return {"ok": True}

@app.get("/pending-message")
async def get_pending_message():
    msg = _pending_message.pop("text", None)
    return {"text": msg}

@app.get("/", response_class=HTMLResponse)
async def root():
    # No-store so browsers never reuse stale softphone JS across restarts
    # (a cached old build caused 'answered but no audio' on port 8401).
    return HTMLResponse(content=HTML, headers={
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    })

@app.get("/favicon.ico")
async def favicon_ico():
  if NOVA_ICON_ICO.exists():
    return FileResponse(str(NOVA_ICON_ICO), media_type="image/x-icon")
  if NOVA_ICON_PNG.exists():
    return FileResponse(str(NOVA_ICON_PNG), media_type="image/png")
  return HTMLResponse(status_code=404, content="")

@app.get("/nova-icon.png")
async def nova_icon_png():
  if NOVA_ICON_PNG.exists():
    return FileResponse(str(NOVA_ICON_PNG), media_type="image/png")
  if NOVA_ICON_ICO.exists():
    return FileResponse(str(NOVA_ICON_ICO), media_type="image/x-icon")
  return HTMLResponse(status_code=404, content="")

@app.get("/vendor/ringcentral-web-phone.min.js")
async def vendor_webphone_js():
  if WEBPHONE_JS.exists():
    return FileResponse(str(WEBPHONE_JS), media_type="application/javascript")
  return HTMLResponse(status_code=404, content="")

@app.get("/phone/config")
async def phone_config():
  return nova_phone.get_config()

@app.post("/phone/sip-provision")
async def phone_sip_provision():
  loop = asyncio.get_event_loop()
  result = await loop.run_in_executor(EXECUTOR, nova_phone.sip_provision)
  if result.get("ok"):
    return result
  return JSONResponse(status_code=int(result.get("status", 500)), content=result)

@app.post("/phone/state")
async def phone_state(request: Request):
  try:
    body = await request.json()
  except Exception:
    body = {}
  state = nova_phone.set_answer_state(bool(body.get("enabled")), str(body.get("owner", "")))
  log.info(f"Phone answer-calls state: {state}")
  return state

@app.post("/phone/debug-log")
async def phone_debug_log(request: Request):
  """Client posts step-by-step call events here; appended to nova_call_debug.log
  so the answer flow can be traced end-to-end without browser DevTools."""
  try:
    body = await request.json()
  except Exception:
    body = {}
  step = str(body.get("step") or "").strip()
  detail = body.get("detail")
  try:
    line = f"{datetime.now().strftime('%H:%M:%S.%f')[:-3]}  CLIENT  {step}"
    if detail not in (None, ""):
      line += f"  | {detail if isinstance(detail, str) else json.dumps(detail)[:800]}"
    with open(os.path.join(_HERE, "nova_call_debug.log"), "a", encoding="utf-8") as fh:
      fh.write(line + "\n")
  except Exception as e:
    log.error(f"phone_debug_log write failed: {e}")
  return {"ok": True}

@app.get("/phone/match-caller")
async def phone_match_caller(number: str = ""):
  import dmelogic.nova_ringcentral as rc_tools
  loop = asyncio.get_event_loop()
  try:
    result = await loop.run_in_executor(
      EXECUTOR, lambda: rc_tools.match_caller_to_patient(number))
    return result if isinstance(result, dict) else {"matched": False}
  except Exception as e:
    log.error(f"phone_match_caller failed: {e}")
    return {"matched": False, "error": str(e)[:200]}

@app.post("/phone/greeting")
async def phone_greeting(request: Request):
  """Synthesize the call greeting; returns {text, audio_b64}."""
  try:
    body = await request.json()
  except Exception:
    body = {}
  text = str(body.get("text") or "").strip() or nova_phone.DEFAULT_GREETING
  loop = asyncio.get_event_loop()
  # Bilingual greeting → multilingual model so Spanish isn't English-accented.
  audio_b64 = await loop.run_in_executor(
    EXECUTOR, lambda: _synthesize_elevenlabs_b64(text, force=True, language="es"))
  return {"text": text, "audio_b64": audio_b64}


_TRANSFER_KEYWORD_RE = re.compile(
  r"\b(representative|human|operator|pharmacist|real person|speak to (a|some)one|talk to (a|some)one|representante|humano|operador|farmaceutico|persona real|hablar con alguien)\b",
  re.IGNORECASE)
_REFILL_REQUEST_RE = re.compile(
  r"\b(refill|re[- ]?fill|renew|renewal|order\s+(a\s+)?refill|need\s+(a\s+)?refill|resurtido|relleno|renovar|renovacion|ordenar\s+resurtido|necesito\s+resurtido)\b",
  re.IGNORECASE)
_VERIFY_FAIL_RE = re.compile(r"\b(not able to verify|can't verify|cannot verify|no puedo verificar|no se puede verificar)\b", re.IGNORECASE)
_FOLLOWUP_OFFER_RE = re.compile(
  r"\b(message|call\s?back|callback|transfer|team member|representative|pharmacist|mensaje|devolver\s+la\s+llamada|transferir|miembro\s+del\s+equipo|representante|farmaceutico)\b",
  re.IGNORECASE)
_LANGUAGE_SPANISH_RE = re.compile(
  r"\b(spanish|espanol|espanol por favor|en espanol|habla espanol|prefiero espanol|quiero espanol)\b",
  re.IGNORECASE)
_LANGUAGE_ENGLISH_RE = re.compile(
  r"\b(english|in english|en ingles|prefiero ingles|quiero ingles)\b",
  re.IGNORECASE)
_SPANISH_HINT_RE = re.compile(
  r"\b(hola|gracias|por favor|buenas|necesito|quiero|mensaje|equipo|farmacia|fecha de nacimiento|llamada|resurtido|relleno)\b",
  re.IGNORECASE)
_VERIFY_FAIL_SAFE_REPLY_EN = (
  "I'm not able to verify that information. "
  "I can take a callback message for staff, or transfer you to a team member."
)
_VERIFY_FAIL_SAFE_REPLY_ES = (
  "No puedo verificar esa informacion. "
  "Puedo tomar un mensaje para que el personal le devuelva la llamada, "
  "o transferirle con un miembro del equipo."
)
_TRANSFER_UNAVAILABLE_REPLY_EN = (
  "I'm sorry, no one is available to take your call right now. "
  "I've made a note for our team and someone will call you back shortly. "
  "Is there anything else I can help you with in the meantime?"
)
_TRANSFER_UNAVAILABLE_REPLY_ES = (
  "Lo siento, en este momento no hay nadie disponible para atender su llamada. "
  "He dejado una nota para nuestro equipo y alguien le devolvera la llamada muy pronto. "
  "Mientras tanto, hay algo mas en lo que le pueda ayudar?"
)


@app.websocket("/call-audio")
async def call_audio_endpoint(websocket: WebSocket):
  """Per-call voice loop: utterance audio in → transcript + Nova reply audio out."""
  await websocket.accept()
  agent = None
  call_id = ""
  caller_number = ""
  turn_busy = False
  caller_requested_refill = False
  verification_failed = False
  followup_offered = False
  transfer_attempted = False
  transfer_unavailable = False
  language_mode = "en"
  patient_match = None
  loop = asyncio.get_event_loop()

  def _dbg(step: str, detail: str = ""):
    try:
      line = f"{datetime.now().strftime('%H:%M:%S.%f')[:-3]}  SERVER  {step}"
      if detail:
        line += f"  | {detail}"
      with open(os.path.join(_HERE, "nova_call_debug.log"), "a", encoding="utf-8") as fh:
        fh.write(line + "\n")
    except Exception:
      pass

  _dbg("call_audio_ws_accepted")

  async def _tts(text: str):
    lang = language_mode
    return await loop.run_in_executor(
      EXECUTOR, lambda: _synthesize_elevenlabs_b64(text, force=True, language=lang))

  def _record_transfer_fallback_task():
    """Log a callback task when a live transfer could not be completed."""
    if agent is None:
      return None
    try:
      token = f"[Transfer unanswered {call_id}]"
      existing = agent.memory.get_reminders(tag="callback", status="active")
      for r in existing[-30:]:
        if token in str(r.get("content") or ""):
          return None
      content = (
        f"{token} Caller from {caller_number or 'unknown number'} asked for a "
        f"live person but no one answered the warm transfer. Please call them "
        f"back. (Nova took a message instead of retrying the transfer.)"
      )
      return agent.memory.add_reminder(content=content, tag="callback")
    except Exception as e:
      log.warning(f"transfer fallback task failed: {e}")
      return None

  def _record_refill_followup():
    """Deterministically create a staff follow-up when a caller requested a
    refill, so it never depends on the model remembering to call tools.

    Creates three things: a 'Must Go Out' queue entry (visible on the DMELogic
    Must Go Out tab), a repeating reminder due NOW (so the alert actively fires
    until staff mark it done), and a patient tracking note when we have a
    verified patient id."""
    if agent is None:
      return None
    try:
      # Resolve patient name / id / phone from the caller match (if any).
      p = {}
      if isinstance(patient_match, dict):
        p = patient_match.get("patient") or patient_match
      pid = p.get("id") or p.get("patient_id")
      name = str(
        p.get("name") or f"{p.get('first_name', '')} {p.get('last_name', '')}"
      ).strip()
      phone = caller_number or str(p.get("phone") or "").strip()
      display_name = name or "Unknown caller"

      # Dedup: skip if we already logged a follow-up for this call.
      token = f"[Refill call {call_id}]"
      try:
        existing = agent.memory.get_reminders(tag="refill_callback", status="active")
        for r in (existing or [])[-40:]:
          if token in str(r.get("content") or ""):
            return None
      except Exception:
        pass

      note = (
        f"Refill request via Nova phone call from {phone or 'unknown number'}. "
        "Verify the order's current delivery/pickup status on the Orders tab "
        "and follow up with the patient to confirm."
      )

      # 1) Must Go Out queue entry (staff physically see this on the tab).
      try:
        agent.dme.add_must_go_out(
          patient_name=display_name,
          patient_phone=phone,
          notes=note,
        )
      except Exception as e:
        log.warning(f"refill must-go-out entry failed: {e}")

      # 2) Repeating reminder due NOW so it actively alerts.
      rid = None
      try:
        rid = agent.memory.add_reminder(
          content=(
            f"{token} {display_name} ({phone or 'no number'}) requested a refill. "
            "Verify status and call the patient back to confirm."
          ),
          tag="refill_callback",
          due_at=datetime.now().isoformat(),
          remind_every_minutes=30,
        )
      except Exception as e:
        log.warning(f"refill reminder failed: {e}")

      # 3) Patient tracking note (only when we have a verified patient id).
      if pid:
        try:
          agent.dme.create_patient_tracking_note(
            int(pid),
            "other",
            note,
            created_by="Nova",
          )
        except Exception as e:
          log.warning(f"refill tracking note failed: {e}")

      return rid
    except Exception as e:
      log.warning(f"refill follow-up task failed: {e}")
      return None

  try:
    while True:
      data = await websocket.receive_json()
      mtype = data.get("type")

      if mtype == "call_start":
        call_id = str(data.get("call_id") or f"call-{int(datetime.now().timestamp())}")
        caller_number = str(data.get("caller_number") or "")
        _dbg("call_start_received", f"call_id={call_id} caller={caller_number}")

        def _make_agent():
          import dmelogic.nova_ringcentral as rc_tools
          from dmelogic.nova_agent import PhoneAgent
          match = None
          try:
            if caller_number:
              match = rc_tools.match_caller_to_patient(caller_number)
          except Exception:
            match = None
          return match, PhoneAgent(caller_number=caller_number, patient_match=match)

        patient_match, agent = await loop.run_in_executor(EXECUTOR, _make_agent)
        _dbg("agent_created")
        greeting_text = nova_phone.DEFAULT_GREETING
        # The greeting is bilingual, so synthesize it with the multilingual
        # model so the Spanish half doesn't come out English-accented.
        audio_b64 = await loop.run_in_executor(
          EXECUTOR, lambda: _synthesize_elevenlabs_b64(greeting_text, force=True, language="es"))
        _dbg("greeting_synth_done", f"audio_bytes_b64={len(audio_b64 or '')}")
        await websocket.send_json(
          {"type": "greeting", "text": greeting_text, "audio_b64": audio_b64})
        _dbg("greeting_sent_to_client")
        await _broadcast_json({
          "type": "call_transcript", "call_id": call_id, "role": "system",
          "text": f"📞 Nova answered a call from {caller_number or 'an unknown number'}.",
        })

      elif mtype == "utterance":
        if agent is None or turn_busy:
          continue  # no call context yet, or Nova is mid-turn — drop overlap
        turn_busy = True
        try:
          turn = int(data.get("turn") or 0)
          try:
            audio_bytes = base64.b64decode(data.get("audio_b64") or "")
          except Exception:
            audio_bytes = b""

          transcript = await loop.run_in_executor(
            EXECUTOR, lambda: nova_phone.transcribe_utterance(audio_bytes, language=language_mode))
          _dbg("utterance_transcribed", f"turn={turn} bytes={len(audio_bytes)} text={(transcript or '')[:80]!r}")

          if not transcript:
            text = ("Perdon, no le escuche bien. Puede repetirlo, por favor?"
                    if language_mode == "es"
                    else "Sorry, I didn't catch that. Could you repeat it for me?")
            await websocket.send_json({
              "type": "reply", "turn": turn, "user_text": "",
              "text": text, "audio_b64": await _tts(text), "action": None})
            continue

          await _broadcast_json({
            "type": "call_transcript", "call_id": call_id,
            "role": "caller", "text": transcript})

          if _REFILL_REQUEST_RE.search(transcript):
            caller_requested_refill = True

          if _LANGUAGE_ENGLISH_RE.search(transcript):
            language_mode = "en"
          elif _LANGUAGE_SPANISH_RE.search(transcript) or _SPANISH_HINT_RE.search(transcript):
            language_mode = "es"

          reply = await loop.run_in_executor(EXECUTOR, agent.chat, transcript)
          clean = (reply or "").strip()
          action = None
          if "<<TRANSFER>>" in clean:
            action = "transfer"
            clean = clean.replace("<<TRANSFER>>", "").strip()
          if "<<HANGUP>>" in clean:
            action = action or "hangup"
            clean = clean.replace("<<HANGUP>>", "").strip()
          # Safety net: transfer on explicit keywords even if the model
          # forgot the marker (only when a transfer number is configured).
          if not action and nova_phone.TRANSFER_NUMBER and _TRANSFER_KEYWORD_RE.search(transcript):
            action = "transfer"
            if "transfer" not in clean.lower() and "transferir" not in clean.lower():
              transfer_line = " Let me transfer you to a team member now."
              if language_mode == "es":
                transfer_line = " Le voy a transferir con un miembro del equipo ahora."
              clean = (clean + transfer_line).strip()

          # Loop guard: only ONE live transfer attempt per call. If a transfer
          # was already attempted (or already failed, or no number configured),
          # do NOT dial again — fall back to taking a message so the caller is
          # never bounced or double-dialed. Checking transfer_attempted (not
          # just transfer_unavailable) closes the race where the model emits a
          # second transfer on the next turn before the first failure is
          # reported back over the websocket.
          if action == "transfer" and (
              transfer_attempted or transfer_unavailable or not nova_phone.TRANSFER_NUMBER):
            action = None
            clean = (_TRANSFER_UNAVAILABLE_REPLY_ES if language_mode == "es"
                     else _TRANSFER_UNAVAILABLE_REPLY_EN)
            followup_offered = True
            await loop.run_in_executor(EXECUTOR, _record_transfer_fallback_task)
          elif action == "transfer":
            transfer_attempted = True

          if _VERIFY_FAIL_RE.search(clean):
            verification_failed = True
            clean = _VERIFY_FAIL_SAFE_REPLY_ES if language_mode == "es" else _VERIFY_FAIL_SAFE_REPLY_EN
            followup_offered = True
          if action == "transfer" or _FOLLOWUP_OFFER_RE.search(clean):
            followup_offered = True

          if not clean:
            clean = "Un momento, por favor." if language_mode == "es" else "One moment please."

          _dbg("nova_reply", f"turn={turn} action={action or 'none'} text={clean[:200]!r}")
          await _broadcast_json({
            "type": "call_transcript", "call_id": call_id,
            "role": "nova", "text": clean})
          await websocket.send_json({
            "type": "reply", "turn": turn, "user_text": transcript,
            "text": clean, "audio_b64": await _tts(clean), "action": action})
        finally:
          turn_busy = False

      elif mtype == "transfer_failed":
        # The browser softphone reports that a warm transfer could not be
        # completed (no answer / rejected). Nova stays on the line, tells the
        # caller, logs a callback task, and blocks any further transfer retry.
        transfer_unavailable = True
        followup_offered = True
        _dbg("transfer_failed", f"reason={str(data.get('reason') or 'no_answer')} -> message taken")
        await loop.run_in_executor(EXECUTOR, _record_transfer_fallback_task)
        fb = (_TRANSFER_UNAVAILABLE_REPLY_ES if language_mode == "es"
              else _TRANSFER_UNAVAILABLE_REPLY_EN)
        turn_busy = False
        await _broadcast_json({
          "type": "call_transcript", "call_id": call_id, "role": "nova", "text": fb})
        await websocket.send_json({
          "type": "reply", "turn": int(data.get("turn") or 0),
          "user_text": "", "text": fb, "audio_b64": await _tts(fb), "action": None})

      elif mtype == "call_end":
        reason = str(data.get("reason") or "ended")
        if agent is not None:
          auto_followup_id = None
          reason_l = reason.lower()

          # Deterministic refill follow-up: whenever the caller asked for a
          # refill, guarantee a Must Go Out entry + repeating reminder + note,
          # regardless of whether the model called the tools mid-call.
          refill_followup_id = None
          if caller_requested_refill:
            refill_followup_id = await loop.run_in_executor(
              EXECUTOR, _record_refill_followup)
            if refill_followup_id:
              try:
                fresh = await loop.run_in_executor(
                  EXECUTOR, lambda: agent.memory.get_reminders(status="active"))
                await _broadcast_json({
                  "type": "reminders_updated",
                  "reminders": fresh or [],
                  "text": "New refill callback added to reminders.",
                })
              except Exception:
                pass

          needs_auto_followup = (
            caller_requested_refill
            and verification_failed
            and ("hung up" in reason_l or "disconnect" in reason_l or not followup_offered)
          )

          if needs_auto_followup:
            def _auto_record_unverified_followup():
              try:
                token = f"[Auto call {call_id}]"
                existing = agent.memory.get_reminders(tag="follow_up", status="active")
                for r in existing[-30:]:
                  content = str(r.get("content") or "")
                  if token in content:
                    return None

                content = (
                  f"{token} Unverified caller from {caller_number or 'unknown number'} "
                  f"requested a refill, verification failed, and the call ended ({reason}). "
                  "Call back to complete identity verification and refill intake."
                )
                return agent.memory.add_reminder(content, tag="follow_up")
              except Exception as e:
                log.error(f"auto follow-up reminder failed: {e}")
                return None

            auto_followup_id = await loop.run_in_executor(EXECUTOR, _auto_record_unverified_followup)

          def _summarize():
            try:
              facts = []
              if caller_requested_refill:
                facts.append("Caller requested a refill.")
              if verification_failed:
                facts.append("Identity verification failed.")
              if followup_offered:
                facts.append("Message/transfer options were offered.")
              if auto_followup_id:
                facts.append(f"Automatic follow-up reminder #{auto_followup_id} was recorded.")
              extra = " ".join(facts).strip()

              prompt = (
                "The call has ended. For the pharmacy staff chat log, summarize in one or "
                "two plain sentences who called and what they needed, including any "
                "follow-up you recorded. Do not use call control markers."
              )
              if extra:
                prompt = f"{prompt} Additional call facts: {extra}"

              out = agent.chat(
                prompt)
              return (out or "").replace("<<TRANSFER>>", "").replace("<<HANGUP>>", "").strip()
            except Exception:
              return ""
          summary = await loop.run_in_executor(EXECUTOR, _summarize)
          _dbg("call_end", f"reason={reason} auto_followup={auto_followup_id or 'none'} summary={ (summary or '')[:200]!r}")
          if summary:
            await _broadcast_json({
              "type": "call_summary", "call_id": call_id,
              "text": f"📞 Call summary ({reason}): {summary}"})
          agent = None

  except WebSocketDisconnect:
    _dbg("call_audio_ws_disconnected", f"call_id={call_id or 'n/a'}")
    log.info(f"/call-audio disconnected (call {call_id or 'n/a'})")
  except Exception as e:
    _dbg("call_audio_ws_error", str(e)[:200])
    log.error(f"/call-audio error: {e}")



@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    _active_websockets.append(websocket)
    log.info("WebSocket client connected")

    # Normalize and heal any stale/unsupported model IDs before first startup payload.
    import dmelogic.nova_agent as na
    _set_anthropic_models(na, persist=True)

    # Send greeting immediately so UI doesn't hang
    hour = datetime.now().hour
    greeting = f"Good {'morning' if hour < 12 else 'afternoon' if hour < 17 else 'evening'}. Nova online."
    greeting_audio_pending = bool(NOVA_VOICE_ENABLED and ELEVENLABS_API_KEY)
    await websocket.send_json({
        "type": "startup",
        "alert": None,
        "reminders": [],
        "greeting": greeting,
      "audio_b64": None,
      "audio_pending": greeting_audio_pending,
        "model": os.getenv("CLAUDE_MODEL", ""),
        "vision_model": os.getenv("CLAUDE_VISION_MODEL", ""),
        "autosend_pending_on_startup": NOVA_AUTOSEND_PENDING_ON_STARTUP,
    })
    if greeting_audio_pending:
      asyncio.create_task(_send_audio_later(websocket, greeting, context="startup"))

    # Load Nova in background thread so WebSocket stays responsive
    try:
        nova = await asyncio.get_event_loop().run_in_executor(None, get_nova)
    except Exception as e:
        await websocket.send_json({"type": "error", "text": f"Failed to load Nova: {e}"})
        await websocket.close()
        return

    # Send updated startup data after Nova loads — with session continuity
    try:
        alert = nova._proactive_startup()
        reminders = nova.memory.get_reminders(status="active")

        # Build continuity greeting from last session
        continuity_greeting = None
        try:
            recent = nova.memory.get_recent_insights(limit=1)
            if recent:
                last = recent[0]
                parts = []
                summary = last.get("summary", "")
                if summary:
                    parts.append(f"Last session: {summary}")
                unresolved = last.get("unresolved", [])
                if unresolved:
                    items = "; ".join(unresolved[:3])
                    parts.append(f"Still open: {items}")
                if parts:
                    continuity_greeting = " ".join(parts)
        except Exception:
            pass

        await websocket.send_json({
            "type": "startup",
            "alert": alert,
            "reminders": reminders or [],
            "greeting": continuity_greeting,
            "model": os.getenv("CLAUDE_MODEL", ""),
            "vision_model": os.getenv("CLAUDE_VISION_MODEL", ""),
          "autosend_pending_on_startup": NOVA_AUTOSEND_PENDING_ON_STARTUP,
        })
    except Exception as e:
        log.warning(f"Startup data error: {e}")

    # Message loop
    tool_calls_made = []
    original_dispatch = na.dispatch_tool
    current_attachments = []
    auto_escalated_to_sonnet = False
    last_user_activity_monotonic = time.monotonic()

    def _resolve_uploaded_attachment_sources(client, user_message: str, attachments_payload: list) -> list:
      """Resolve uploaded filenames to existing OCR source paths (no file-copy side effects)."""
      if not attachments_payload:
        return []

      def _ext_for_mime(mime: str) -> str:
        mt = str(mime or "").lower().strip()
        mapping = {
          "application/pdf": ".pdf",
          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
          "application/vnd.ms-excel": ".xls",
          "image/jpeg": ".jpg",
          "image/jpg": ".jpg",
          "image/png": ".png",
          "image/webp": ".webp",
          "image/tiff": ".tif",
          "image/bmp": ".bmp",
        }
        return mapping.get(mt, ".bin")

      ocr_root = None
      try:
        from dmelogic.paths import ocr_folder
        ocr_root = Path(ocr_folder())
      except Exception:
        ocr_root = None

      if not ocr_root:
        try:
          configured = os.getenv("NOVA_RX_SOURCE_ROOT", "").strip()
          if configured:
            ocr_root = Path(configured)
        except Exception:
          ocr_root = None

      if not ocr_root:
        ocr_root = Path.home() / "Documents" / "FaxManagerData" / "Faxes OCR'd"

      def _extract_order_id(text: str) -> int | None:
        value = str(text or "")
        m = re.search(r"\bORD\s*[-#:]?\s*(\d{1,6})\b", value, re.IGNORECASE)
        if m:
          try:
            return int(m.group(1))
          except Exception:
            return None
        m = re.search(r"\border\s*[-#:]?\s*(\d{1,6})\b", value, re.IGNORECASE)
        if m:
          try:
            return int(m.group(1))
          except Exception:
            return None
        return None

      def _last_initial_from_name(name: str) -> str:
        raw = str(name or "").strip()
        if not raw:
          return ""
        if "," in raw:
          last = raw.split(",", 1)[0].strip()
        else:
          parts = [p for p in raw.split() if p]
          last = parts[-1] if parts else ""
        return last[:1].upper() if last else ""

      def _resolve_patient_last_initial(order_id: int | None) -> str:
        if not order_id or client is None:
          return ""
        try:
          order = client.get_order(int(order_id))
        except Exception:
          order = None
        if isinstance(order, dict):
          direct = str(order.get("patient_last_name") or order.get("last_name") or "").strip()
          if direct:
            return direct[:1].upper()
          from_name = _last_initial_from_name(order.get("patient_name") or "")
          if from_name:
            return from_name
          patient_id = order.get("patient_id")
          if patient_id:
            try:
              patient = client.get_patient(int(patient_id))
            except Exception:
              patient = None
            if isinstance(patient, dict):
              p_last = str(patient.get("last_name") or "").strip()
              if p_last:
                return p_last[:1].upper()
              return _last_initial_from_name(patient.get("name") or "")
        return ""

      order_id = _extract_order_id(user_message)
      patient_initial = _resolve_patient_last_initial(order_id)

      def _find_existing_source(file_name: str, subfolder_initial: str) -> str:
        base_name = Path(str(file_name or "").strip()).name
        if not base_name:
          return ""

        candidates = []
        if subfolder_initial:
          candidates.append(ocr_root / subfolder_initial / base_name)
        candidates.append(ocr_root / base_name)

        seen = set()
        for candidate in candidates:
          key = str(candidate).lower()
          if key in seen:
            continue
          seen.add(key)
          if candidate.exists() and candidate.is_file():
            return str(candidate)

        if subfolder_initial:
          try:
            letter_dir = ocr_root / subfolder_initial
            if letter_dir.exists():
              for match in letter_dir.rglob(base_name):
                if match.is_file():
                  return str(match)
          except Exception:
            pass

        return ""

      enriched = []
      for idx, att in enumerate(attachments_payload, start=1):
        item = dict(att or {})
        try:
          incoming_name = str(item.get("attachment_name") or item.get("fileName") or "").strip()
          safe_name = Path(incoming_name).name if incoming_name else ""
          if not safe_name:
            mime_type = item.get("image_type") or item.get("mimeType") or "application/octet-stream"
            safe_name = f"attachment_{idx}{_ext_for_mime(mime_type)}"

          resolved_source_path = _find_existing_source(safe_name, patient_initial)
          item["source_root"] = str(ocr_root)
          item["source_file_name"] = safe_name
          if patient_initial:
            item["source_subfolder_initial"] = patient_initial
          if resolved_source_path:
            item["resolved_source_path"] = resolved_source_path
        except Exception as e:
          log.warning(f"Upload source path resolution failed: {e}")
        enriched.append(item)

      return enriched

    def _persist_attachments_for_order(client, order_id: int, attachments_payload: list) -> dict:
      """Link source files to order/patient records without creating duplicate copies."""
      if not attachments_payload:
        return {"saved_paths": [], "saved_count": 0}

      def _ext_for_mime(mime: str) -> str:
        mt = str(mime or "").lower().strip()
        mapping = {
          "application/pdf": ".pdf",
          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
          "application/vnd.ms-excel": ".xls",
          "image/jpeg": ".jpg",
          "image/jpg": ".jpg",
          "image/png": ".png",
          "image/webp": ".webp",
          "image/tiff": ".tif",
          "image/bmp": ".bmp",
        }
        return mapping.get(mt, ".bin")

      def _split_refs(value: str) -> list[str]:
        refs = []
        for part in str(value or "").replace(";", "\n").splitlines():
          p = part.strip()
          if p:
            refs.append(p)
        return refs

      folder_path = getattr(client, "folder_path", None)
      if not folder_path:
        try:
          from dmelogic.paths import db_dir
          folder_path = str(db_dir())
        except Exception:
          folder_path = str(Path.home() / "Documents" / "DmeSolutionsV1" / "Data")

      saved_paths = []
      saved_names = []
      for idx, att in enumerate(attachments_payload, start=1):
        try:
          resolved = str(att.get("resolved_source_path") or "").strip()
          if not resolved:
            source_root = Path(str(att.get("source_root") or "").strip())
            file_name = Path(str(att.get("source_file_name") or "").strip()).name
            subfolder_initial = str(att.get("source_subfolder_initial") or "").strip()
            if source_root and file_name:
              candidate = source_root / subfolder_initial / file_name if subfolder_initial else source_root / file_name
              if candidate.exists() and candidate.is_file():
                resolved = str(candidate)

          if not resolved:
            continue

          resolved_path = Path(resolved)
          if not resolved_path.exists() or not resolved_path.is_file():
            continue

          saved_paths.append(str(resolved_path))
          saved_names.append(resolved_path.name)
        except Exception as e:
          log.warning(f"Attachment link prep failed for order {order_id}: {e}")

      if not saved_paths:
        return {"saved_paths": [], "saved_count": 0}

      patient_id = None
      try:
        from dmelogic.db.base import get_connection
        conn = get_connection("orders.db", folder_path=folder_path)
        cur = conn.cursor()

        try:
          cur.execute("SELECT attached_rx_files, patient_id FROM orders WHERE id = ?", (int(order_id),))
          row = cur.fetchone()
        except Exception:
          cur.execute("ALTER TABLE orders ADD COLUMN attached_rx_files TEXT")
          conn.commit()
          cur.execute("SELECT attached_rx_files, patient_id FROM orders WHERE id = ?", (int(order_id),))
          row = cur.fetchone()

        existing_refs = _split_refs(row[0] if row else "")
        patient_id = row[1] if row else None

        merged = list(existing_refs)
        seen = {r.lower() for r in merged}
        for p in saved_paths:
          if p.lower() not in seen:
            merged.append(p)
            seen.add(p.lower())

        cur.execute(
          "UPDATE orders SET attached_rx_files = ? WHERE id = ?",
          (";".join(merged), int(order_id)),
        )
        conn.commit()
        conn.close()
      except Exception as e:
        log.warning(f"Failed to update attached_rx_files for order {order_id}: {e}")

      if patient_id:
        try:
          from dmelogic.db.base import resolve_db_path
          import sqlite3

          patient_db = resolve_db_path("patients.db", folder_path=folder_path)
          conn = sqlite3.connect(patient_db)
          cur = conn.cursor()
          cur.execute(
            """
            CREATE TABLE IF NOT EXISTS patient_documents (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              patient_id INTEGER NOT NULL,
              description TEXT,
              original_name TEXT,
              stored_path TEXT,
              created_at TEXT DEFAULT (datetime('now'))
            )
            """
          )

          order_num = f"ORD-{int(order_id):03d}"
          for full_path, original_name in zip(saved_paths, saved_names):
            cur.execute(
              "SELECT id FROM patient_documents WHERE patient_id = ? AND stored_path = ?",
              (int(patient_id), full_path),
            )
            if cur.fetchone():
              continue
            cur.execute(
              "INSERT INTO patient_documents (patient_id, description, original_name, stored_path) VALUES (?, ?, ?, ?)",
              (int(patient_id), f"From {order_num}", original_name, full_path),
            )

          conn.commit()
          conn.close()
        except Exception as e:
          log.warning(f"Failed to link patient documents for order {order_id}: {e}")

      return {"saved_paths": saved_paths, "saved_count": len(saved_paths)}

    def patched_dispatch(name, inp, client, memory):
        tool_calls_made.append(name)
        result = original_dispatch(name, inp, client, memory)

        if name == "create_order" and current_attachments:
            try:
                parsed = json.loads(result) if isinstance(result, str) else result
                if isinstance(parsed, dict) and parsed.get("success") and parsed.get("order_id"):
                    link_result = _persist_attachments_for_order(
                        client,
                        int(parsed.get("order_id")),
                        current_attachments,
                    )
                    if link_result.get("saved_count", 0) > 0:
                        tool_calls_made.append(f"link_order_attachments({link_result['saved_count']})")
            except Exception as e:
                log.warning(f"post-create attachment linking failed: {e}")

        return result

    na.dispatch_tool = patched_dispatch

    try:
        while True:
            data = await websocket.receive_json()

            now_monotonic = time.monotonic()
            idle_seconds = now_monotonic - last_user_activity_monotonic
            if idle_seconds >= NOVA_MODEL_IDLE_RESET_SECONDS:
                current_models = _set_anthropic_models(na, persist=False)
                if (
                    current_models.get("model") != NOVA_ECONOMY_CHAT_MODEL
                    or current_models.get("vision_model") != NOVA_ECONOMY_VISION_MODEL
                ):
                    _reset_models_to_haiku(na, persist=False)
                    auto_escalated_to_sonnet = False
                    log.info(
                        f"Idle model reset applied after {int(idle_seconds)}s: "
                        f"chat={NOVA_ECONOMY_CHAT_MODEL}, vision={NOVA_ECONOMY_VISION_MODEL}"
                    )

            if data.get("type") == "delete_reminders":
                raw_ids = data.get("reminder_ids") or []
                ids = []
                for rid in raw_ids:
                    try:
                        ids.append(int(rid))
                    except Exception:
                        continue

                deleted = 0
                for rid in ids:
                    try:
                        nova.memory.delete_reminder(rid)
                        deleted += 1
                    except Exception:
                        pass

                reminders = nova.memory.get_reminders(status="active")
                await websocket.send_json({
                    "type": "reminders_updated",
                    "reminders": reminders or [],
                    "deleted_ids": ids,
                    "text": f"Deleted {deleted} reminder{'s' if deleted != 1 else ''}.",
                })
                continue

            if data.get("type") == "delete_reminder":
                rid_raw = data.get("reminder_id")
                rid = None
                try:
                    rid = int(rid_raw)
                except Exception:
                    rid = None

                deleted = 0
                if rid is not None:
                    try:
                        nova.memory.delete_reminder(rid)
                        deleted = 1
                    except Exception:
                        deleted = 0

                reminders = nova.memory.get_reminders(status="active")
                await websocket.send_json({
                    "type": "reminders_updated",
                    "reminders": reminders or [],
                    "deleted_ids": [rid] if rid is not None else [],
                    "text": "Deleted 1 reminder." if deleted == 1 else "No reminder deleted.",
                })
                continue

            if data.get("type") == "complete_reminder":
                rid_raw = data.get("reminder_id")
                rid = None
                try:
                    rid = int(rid_raw)
                except Exception:
                    rid = None

                completed = 0
                if rid is not None:
                    try:
                        nova.memory.complete_reminder(rid)
                        completed = 1
                    except Exception:
                        completed = 0

                reminders = nova.memory.get_reminders(status="active")
                await websocket.send_json({
                    "type": "reminders_updated",
                    "reminders": reminders or [],
                    "completed_id": rid,
                    "text": "Marked reminder complete." if completed == 1 else "Could not complete reminder.",
                })
                continue

            if data.get("type") == "update_reminder":
                rid_raw = data.get("reminder_id")
                content = str(data.get("content") or "").strip()
                tag = str(data.get("tag") or "general").strip().lower() or "general"
                due_at_raw = data.get("due_at")
                follow_up_notes_raw = data.get("follow_up_notes")
                follow_up_at_raw = data.get("follow_up_at")
                cadence_raw = data.get("remind_every_minutes", 30)

                rid = None
                try:
                    rid = int(rid_raw)
                except Exception:
                    rid = None

                due_at = str(due_at_raw or "").strip() or None
                follow_up_notes = str(follow_up_notes_raw or "").strip() or None
                follow_up_at = str(follow_up_at_raw or "").strip() or None
                try:
                    cadence = int(cadence_raw)
                except Exception:
                    cadence = 30
                cadence = max(1, min(cadence, 24 * 60))

                ok = False
                if rid is not None and content:
                    try:
                        ok = bool(
                            nova.memory.update_reminder(
                                rid,
                                content,
                                tag=tag,
                                due_at=due_at,
                                follow_up_notes=follow_up_notes,
                                follow_up_at=follow_up_at,
                                remind_every_minutes=cadence,
                            )
                        )
                    except Exception:
                        ok = False

                reminders = nova.memory.get_reminders(status="active")
                await websocket.send_json({
                    "type": "reminders_updated",
                    "reminders": reminders or [],
                    "updated_id": rid,
                    "text": "Reminder updated." if ok else "Could not update reminder.",
                })
                continue

            if data.get("type") == "billing_check":
                try:
                    rows = nova.dme.get_unbilled_orders(limit=500)
                    if isinstance(rows, dict):
                        if rows.get("error"):
                            raise RuntimeError(str(rows.get("error")))
                        rows = rows.get("orders") or []
                    rows = rows if isinstance(rows, list) else []
                    count = len(rows)

                    if count > 0:
                        page_state = str(data.get("page_state") or "unknown")
                        state_hint = ""
                        if page_state in {"background", "hidden"}:
                            state_hint = " You are away from the Nova window."

                        msg = (
                            f"Billing reminder: {count} order{'s are' if count != 1 else ' is'} pending billing."
                            f"{state_hint} Please check the Orders tab."
                        )
                        reminders = nova.memory.get_reminders(status="active")
                        await websocket.send_json({
                            "type": "nova_response",
                            "text": msg,
                            "reminders": reminders,
                            "audio_b64": _synthesize_elevenlabs_b64(msg),
                            "model": os.getenv("CLAUDE_MODEL", ""),
                            "vision_model": os.getenv("CLAUDE_VISION_MODEL", ""),
                        })
                except Exception as e:
                    log.warning(f"Scheduled billing check failed: {e}")
                continue

            if data.get("type") != "message":
                continue

            user_text = data.get("text", "").strip()
            attachments = data.get("attachments") or []
            if not isinstance(attachments, list):
              attachments = []

            # Backward compatibility with single attachment payload shape.
            if not attachments and data.get("image"):
              attachments = [{
                "image": data.get("image"),
                "image_type": data.get("image_type", "image/png"),
                "attachment_name": data.get("attachment_name", ""),
              }]

            if not user_text and not attachments:
                continue

            last_user_activity_monotonic = now_monotonic

            dme_client = getattr(nova, "dme", None)
            attachments = _resolve_uploaded_attachment_sources(dme_client, user_text, attachments)
            current_attachments = list(attachments)

            if auto_escalated_to_sonnet and not current_attachments:
              _reset_models_to_haiku(na, persist=False)
              auto_escalated_to_sonnet = False

            # Runtime model controls
            lower_text = user_text.lower()
            if lower_text.startswith("/show-models"):
              current = _set_anthropic_models(na, persist=False)
              msg = (
                "Current Anthropic models:\n"
                f"chat={current['model']}\n"
                f"vision={current['vision_model']}\n"
                f"fallbacks={current['fallbacks'] or '(none)'}"
              )
              await websocket.send_json({
                "type": "nova_response",
                "text": msg,
                "reminders": nova.memory.get_reminders(status="active"),
                "audio_b64": _synthesize_elevenlabs_b64(msg),
                "model": current["model"],
                "vision_model": current["vision_model"],
              })
              continue

            if lower_text.startswith("/model "):
              model = user_text.split(" ", 1)[1].strip()
              current = _set_anthropic_models(na, model=model, persist=True)
              auto_escalated_to_sonnet = False
              msg = f"Chat model updated to {current['model']}."
              await websocket.send_json({
                "type": "nova_response",
                "text": msg,
                "reminders": nova.memory.get_reminders(status="active"),
                "audio_b64": _synthesize_elevenlabs_b64(msg),
                "model": current["model"],
                "vision_model": current["vision_model"],
              })
              continue

            if lower_text.startswith("/vision-model "):
              vision = user_text.split(" ", 1)[1].strip()
              current = _set_anthropic_models(na, vision_model=vision, persist=True)
              auto_escalated_to_sonnet = False
              msg = f"Vision model updated to {current['vision_model']}."
              await websocket.send_json({
                "type": "nova_response",
                "text": msg,
                "reminders": nova.memory.get_reminders(status="active"),
                "audio_b64": _synthesize_elevenlabs_b64(msg),
                "model": current["model"],
                "vision_model": current["vision_model"],
              })
              continue

            if lower_text.startswith("/model-fallbacks "):
              fallbacks = user_text.split(" ", 1)[1].strip()
              current = _set_anthropic_models(na, fallbacks=fallbacks, persist=True)
              msg = "Model fallback list updated."
              await websocket.send_json({
                "type": "nova_response",
                "text": msg,
                "reminders": nova.memory.get_reminders(status="active"),
                "audio_b64": _synthesize_elevenlabs_b64(msg),
                "model": current["model"],
                "vision_model": current["vision_model"],
              })
              continue

            tool_calls_made.clear()

            escalated_for_this_turn = False
            if current_attachments and NOVA_AUTO_SONNET_FOR_ATTACHMENTS:
              current_models = _set_anthropic_models(na, persist=False)
              if not (
                _is_sonnet_or_opus(current_models.get("model", ""))
                and _is_sonnet_or_opus(current_models.get("vision_model", ""))
              ):
                _set_anthropic_models(
                  na,
                  model=NOVA_AUTO_SONNET_CHAT_MODEL,
                  vision_model=NOVA_AUTO_SONNET_VISION_MODEL,
                  persist=False,
                )
                escalated_for_this_turn = True
                auto_escalated_to_sonnet = True

            # Build message — with or without attachments
            def build_chat_message():
                if not attachments:
                    return nova.chat(user_text)

                def _normalize_attachment(att, idx):
                    image_data = att.get("image") or att.get("dataUrl")
                    if not image_data:
                        return None
                    image_type = att.get("image_type") or att.get("mimeType") or "application/octet-stream"
                    attachment_name = att.get("attachment_name") or att.get("fileName") or f"attachment_{idx}"
                    source_root = att.get("source_root") or ""
                    source_initial = att.get("source_subfolder_initial") or ""
                    resolved_source_path = att.get("resolved_source_path") or ""
                    raw_b64 = image_data.split(',', 1)[1] if ',' in image_data else image_data
                    return {
                        "raw_b64": raw_b64,
                        "image_type": image_type,
                        "attachment_name": attachment_name,
                      "source_root": source_root,
                      "source_subfolder_initial": source_initial,
                      "resolved_source_path": resolved_source_path,
                    }

                def _extract_excel_preview(raw_b64, attachment_name):
                    extracted_text = ""
                    parse_error = ""
                    try:
                        import tempfile, base64 as _b64
                        excel_bytes = _b64.b64decode(raw_b64)
                        suffix = ".xlsx"
                        lowered_name = str(attachment_name or "").lower()
                        if lowered_name.endswith(".xls"):
                            suffix = ".xls"
                        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                            tmp.write(excel_bytes)
                            tmp_path = tmp.name
                        try:
                            try:
                                import pandas as _pd
                                sheets = _pd.read_excel(tmp_path, sheet_name=None)
                                parts = []
                                for sheet_name, df in sheets.items():
                                    df2 = df.fillna("").astype(str)
                                    row_count, col_count = df2.shape
                                    preview = df2.head(80).to_csv(index=False)
                                    parts.append(f"Sheet: {sheet_name} ({row_count} rows x {col_count} cols)\\n{preview}")
                                extracted_text = "\\n\\n".join(parts)
                            except Exception:
                                from openpyxl import load_workbook
                                wb = load_workbook(tmp_path, data_only=True)
                                parts = []
                                for ws in wb.worksheets:
                                    rows = list(ws.iter_rows(values_only=True))
                                    preview_rows = rows[:81]
                                    if not preview_rows:
                                        parts.append(f"Sheet: {ws.title} (0 rows)")
                                        continue
                                    header = ["" if v is None else str(v) for v in (preview_rows[0] or [])]
                                    body = preview_rows[1:] if len(preview_rows) > 1 else []
                                    lines = []
                                    if any(h.strip() for h in header):
                                        lines.append(",".join(h.replace("\\n", " ") for h in header))
                                    for r in body[:80]:
                                        vals = ["" if v is None else str(v) for v in (r or [])]
                                        lines.append(",".join(v.replace("\\n", " ") for v in vals))
                                    parts.append(f"Sheet: {ws.title} ({max(0, len(rows)-1)} data rows)\\n" + "\\n".join(lines))
                                extracted_text = "\\n\\n".join(parts)
                        finally:
                            import os as _os
                            _os.unlink(tmp_path)
                    except Exception as e:
                        parse_error = str(e)
                    return extracted_text, parse_error

                def _extract_pdf_text(raw_b64, attachment_name):
                    ocr_text = ""
                    try:
                        import tempfile, base64 as _b64
                        pdf_bytes = _b64.b64decode(raw_b64)
                        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                            tmp.write(pdf_bytes)
                            tmp_path = tmp.name
                        try:
                            from dmelogic.ocr_tools import extract_text_from_pdf
                            ocr_text = extract_text_from_pdf(tmp_path)
                        finally:
                            import os as _os
                            _os.unlink(tmp_path)
                    except Exception as e:
                        log.warning(f"PDF text extraction skipped for {attachment_name}: {e}")
                    return ocr_text

                def _prepare_image(raw_b64, attachment_name):
                    enhanced_b64 = raw_b64
                    tesseract_text = ""
                    try:
                        import io as _io, base64 as _b64
                        from PIL import Image as _PIL, ImageFilter as _Filt, ImageEnhance as _Enh
                        img_bytes = _b64.b64decode(raw_b64)
                        img = _PIL.open(_io.BytesIO(img_bytes)).convert("RGB")
                        w, h = img.size
                        if w < 1200:
                            scale = 1200 / w
                            img = img.resize((int(w * scale), int(h * scale)), _PIL.LANCZOS)
                        img = img.filter(_Filt.SHARPEN)
                        img = _Enh.Contrast(img).enhance(1.5)
                        img = _Enh.Sharpness(img).enhance(2.0)

                        buf = _io.BytesIO()
                        img.save(buf, format="PNG")
                        enhanced_b64 = _b64.b64encode(buf.getvalue()).decode()

                        try:
                            import pytesseract
                            ocr_raw = pytesseract.image_to_string(img, config="--psm 6 --oem 3")
                            if ocr_raw.strip():
                                tesseract_text = ocr_raw.strip()
                        except Exception:
                            pass
                    except Exception as e:
                        log.warning(f"Image preprocessing skipped for {attachment_name}: {e}")
                    return enhanced_b64, tesseract_text

                normalized = []
                for idx, att in enumerate(attachments, start=1):
                    item = _normalize_attachment(att, idx)
                    if item:
                        normalized.append(item)

                if not normalized:
                    return nova.chat(user_text or "Please analyze the provided files.")

                if len(normalized) == 1:
                  one = normalized[0]
                  raw_b64 = one["raw_b64"]
                  image_type = one["image_type"]
                  attachment_name = one["attachment_name"]
                  source_root = one.get("source_root") or ""
                  source_initial = one.get("source_subfolder_initial") or ""
                  resolved_source_path = one.get("resolved_source_path") or ""
                  is_pdf = image_type == "application/pdf"
                  is_excel = image_type in {
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                  }

                  metadata_block = (
                    f"[Uploaded file name: {attachment_name}]\n"
                    + (f"[OCR source root: {source_root}]\n" if source_root else "")
                    + (f"[Expected patient subfolder: {source_initial}]\n" if source_initial else "")
                    + (f"[Resolved source path: {resolved_source_path}]\n" if resolved_source_path else "")
                  )

                  if is_excel:
                    extracted_text, parse_error = _extract_excel_preview(raw_b64, attachment_name)
                    prompt = user_text or "Please analyze this spreadsheet and summarize key insights."
                    if extracted_text.strip():
                      prompt = (
                        f"{prompt}\\n\\n"
                        f"{metadata_block}"
                        f"[Spreadsheet file: {attachment_name or 'uploaded.xlsx'}]\\n"
                        f"[Extracted sheet data preview (trimmed)]:\\n{extracted_text.strip()}"
                      )
                    elif parse_error:
                      prompt = (
                        f"{prompt}\\n\\n"
                        f"{metadata_block}"
                        f"[Spreadsheet file: {attachment_name or 'uploaded.xlsx'}]\\n"
                        f"[Note: spreadsheet parsing failed: {parse_error}]"
                      )
                    else:
                      prompt = f"{prompt}\n\n{metadata_block}".strip()
                    return nova.chat(prompt)

                  if is_pdf:
                    ocr_text = _extract_pdf_text(raw_b64, attachment_name)
                    prompt = user_text or "Please read and extract all information from this document."
                    prompt = f"{prompt}\n\n{metadata_block}".strip()
                    if ocr_text.strip():
                      prompt = (
                        f"{prompt}\\n\\n"
                        f"[Pre-extracted OCR text for cross-reference - verify against the document]:\\n"
                        f"{ocr_text.strip()}"
                      )
                    return nova.chat_with_image(prompt, raw_b64, "application/pdf")

                  if str(image_type).startswith("image/"):
                    enhanced_b64, tesseract_text = _prepare_image(raw_b64, attachment_name)
                    prompt = user_text or "Please read and extract all text from this document."
                    prompt = f"{prompt}\n\n{metadata_block}".strip()
                    if tesseract_text:
                      prompt = (
                        f"{prompt}\\n\\n"
                        f"[Pre-extracted OCR text for cross-reference - verify against the image]:\\n"
                        f"{tesseract_text}"
                      )
                    return nova.chat_with_image(prompt, enhanced_b64, "image/png")

                  prompt = user_text or "Please analyze this attachment."
                  prompt = (
                    f"{prompt}\\n\\n"
                    f"{metadata_block}"
                    f"[Attachment: {attachment_name} ({image_type})]\\n"
                    "[Unsupported file type for direct parsing. Ask for PDF/image/text conversion.]"
                  )
                  return nova.chat(prompt)

                sections = []
                for idx, item in enumerate(normalized, start=1):
                    raw_b64 = item["raw_b64"]
                    image_type = item["image_type"]
                    attachment_name = item["attachment_name"]
                    source_root = item.get("source_root") or ""
                    source_initial = item.get("source_subfolder_initial") or ""
                    resolved_source_path = item.get("resolved_source_path") or ""
                    is_pdf = image_type == "application/pdf"
                    is_excel = image_type in {
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "application/vnd.ms-excel",
                    }

                    attachment_header = f"[Attachment {idx}: {attachment_name} ({image_type})]"
                    if source_root:
                      attachment_header += f"\n[Attachment {idx} OCR root: {source_root}]"
                    if source_initial:
                      attachment_header += f"\n[Attachment {idx} expected subfolder: {source_initial}]"
                    if resolved_source_path:
                      attachment_header += f"\n[Attachment {idx} resolved source path: {resolved_source_path}]"

                    if is_excel:
                        extracted_text, parse_error = _extract_excel_preview(raw_b64, attachment_name)
                        if extracted_text.strip():
                            sections.append(
                          f"{attachment_header}\n"
                                f"[Extracted spreadsheet data preview]:\\n{extracted_text.strip()}"
                            )
                        else:
                            sections.append(
                          f"{attachment_header}\n"
                                f"[Spreadsheet parsing failed: {parse_error or 'unknown error'}]"
                            )
                        continue

                    if is_pdf:
                        ocr_text = _extract_pdf_text(raw_b64, attachment_name)
                        sections.append(
                        f"{attachment_header}\n"
                            + (ocr_text.strip() if ocr_text.strip() else "[No extractable text found]")
                        )
                        continue

                    if str(image_type).startswith("image/"):
                        _, tesseract_text = _prepare_image(raw_b64, attachment_name)
                        sections.append(
                        f"{attachment_header}\n"
                            + (tesseract_text if tesseract_text else "[No extractable text found]")
                        )
                        continue

                    sections.append(
                      f"{attachment_header}\n"
                        "[Unsupported file type for direct parsing. Mention this to user and request conversion to PDF/image/text.]"
                    )

                prompt = user_text or "Please analyze all attached files together and summarize key findings."
                if sections:
                    prompt = f"{prompt}\\n\\n" + "\\n\\n".join(sections)
                return nova.chat(prompt)

            # Run Nova in shared executor (reused across requests)
            loop = asyncio.get_event_loop()
            response = None
            error = None

            future = loop.run_in_executor(EXECUTOR, build_chat_message)
            last_sent = 0
            while not future.done():
                await asyncio.sleep(0.15)
                while last_sent < len(tool_calls_made):
                    await websocket.send_json({
                        "type": "tool_call",
                        "tool": tool_calls_made[last_sent]
                    })
                    last_sent += 1

            try:
                response = await future
            except Exception as e:
                error = str(e)

            # Send remaining tool calls
            while last_sent < len(tool_calls_made):
                await websocket.send_json({
                    "type": "tool_call",
                    "tool": tool_calls_made[last_sent]
                })
                last_sent += 1

            if error:
                await websocket.send_json({"type": "error", "text": error})
            else:
                reminders = nova.memory.get_reminders(status="active")
                response_audio_pending = bool(NOVA_VOICE_ENABLED and ELEVENLABS_API_KEY)
                actual_model = getattr(nova, 'last_model_used', '') or os.getenv("CLAUDE_MODEL", "")
                await websocket.send_json({
                    "type": "nova_response",
                    "text": response,
                    "reminders": reminders,
                  "audio_b64": None,
                  "audio_pending": response_audio_pending,
                    "model": actual_model,
                    "vision_model": os.getenv("CLAUDE_VISION_MODEL", ""),
                })
                if response_audio_pending:
                  asyncio.create_task(_send_audio_later(websocket, response, context="response"))

            if escalated_for_this_turn:
                _reset_models_to_haiku(na, persist=False)
                auto_escalated_to_sonnet = False
                log.info("Attachment turn complete; models reset to Haiku economy baseline")

            current_attachments = []

    except WebSocketDisconnect:
        log.info("WebSocket client disconnected")
    except Exception as e:
        log.error(f"WebSocket error: {e}")
        try:
            await websocket.send_json({"type": "error", "text": str(e)})
        except Exception:
            pass
    finally:
        # Always remove from active registry on any exit path
        try:
            _active_websockets.remove(websocket)
        except ValueError:
            pass
        na.dispatch_tool = original_dispatch

        # ── Session cleanup on disconnect (LLM summary is opt-in) ─────
        try:
          if NOVA_AUTO_SUMMARIZE_ON_DISCONNECT:
            log.info(f"Auto-summarizing session {nova.session_id}…")
            loop = asyncio.get_event_loop()
            insight = await loop.run_in_executor(None, nova.auto_summarize_session)
            if insight:
              log.info(f"Session summary saved: {insight.get('summary', '')[:100]}")
          nova.end_session()
        except Exception as e:
          log.warning(f"Session cleanup failed: {e}")
          try:
            nova.end_session()
          except Exception:
            pass

# ── Entry point ───────────────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s")
    log.info("Nova UI starting at http://localhost:8401")
    uvicorn.run("nova_ui_server:app", host="127.0.0.1", port=8401, reload=False)