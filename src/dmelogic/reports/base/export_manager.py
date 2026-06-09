"""
Export Manager - Universal export functionality for all reports

Handles CSV, Excel, and PDF exports with consistent formatting.
"""

from __future__ import annotations
import csv
import json
from datetime import datetime
from typing import Optional, List, Dict, Any
from pathlib import Path

from PyQt6.QtWidgets import QFileDialog, QMessageBox, QWidget


class ExportManager:
    """
    Universal export handler for report data.

    Supports:
    - CSV export (simple, universal)
    - Excel export (formatted with colors)
    - PDF export (professional layout)
    """

    def __init__(self, parent: Optional[QWidget] = None):
        """
        Initialize export manager.

        Args:
            parent: Parent widget for file dialogs
        """
        self.parent = parent

    def _excel_safe_value(self, value: Any) -> Any:
        """Convert complex Python objects to Excel-safe scalar values."""
        if isinstance(value, (dict, list, tuple, set)):
            try:
                return json.dumps(value, ensure_ascii=False)
            except Exception:
                return str(value)
        return value

    def _effective_columns(self, report_data) -> List[Any]:
        """Return report columns, forcing key order fields into exports when present."""
        columns = list(report_data.columns or [])
        title = str(getattr(report_data, "title", "") or "").lower()
        if "order report" not in title:
            return columns

        existing = {getattr(c, "name", "") for c in columns}
        row_data = [getattr(r, "data", {}) for r in (report_data.rows or [])]

        class _Column:
            def __init__(self, name: str, display_name: str, data_type: str = "text", alignment: str = "left"):
                self.name = name
                self.display_name = display_name
                self.data_type = data_type
                self.alignment = alignment

            def format_value(self, value: Any) -> str:
                if value is None:
                    return ""
                if self.data_type == "currency":
                    try:
                        return f"${float(value):,.2f}"
                    except Exception:
                        return str(value)
                if self.data_type == "number":
                    try:
                        return f"{float(value):,.1f}"
                    except Exception:
                        return str(value)
                return str(value)

        if "refill_order_number" not in existing:
            has_refill_info = any((d.get("refill_order_number") or d.get("refill_number")) for d in row_data if isinstance(d, dict))
            if has_refill_info:
                columns.insert(4 if len(columns) >= 4 else len(columns), _Column("refill_order_number", "Refill #", "text", "center"))

        if "tracking_number" not in existing:
            has_tracking = any(d.get("tracking_number") for d in row_data if isinstance(d, dict))
            if has_tracking:
                columns.append(_Column("tracking_number", "Tracking #", "text", "left"))

        return columns

    def _value_for_column(self, row, col_name: str) -> Any:
        """Get value with Order Report fallbacks for derived refill identifiers."""
        value = row.get(col_name)
        if col_name != "refill_order_number":
            return value
        if value:
            return value
        try:
            refill = int(row.get("refill_number") or 0)
        except Exception:
            refill = 0
        if refill <= 0:
            return ""
        try:
            order_id = int(row.get("order_number") or 0)
        except Exception:
            order_id = 0
        if order_id <= 0:
            return f"R{refill}"
        return f"ORD-{order_id:03d}-R{refill}"

    # ========================================================================
    # CSV Export
    # ========================================================================

    def export_csv(
        self,
        report_data,
        default_filename: Optional[str] = None
    ) -> bool:
        """
        Export report to CSV file.

        Args:
            report_data: ReportData object
            default_filename: Suggested filename

        Returns:
            True if successful, False otherwise
        """
        if not default_filename:
            safe_title = report_data.title.replace(' ', '_').replace('/', '_')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"{safe_title}_{timestamp}.csv"

        filename, _ = QFileDialog.getSaveFileName(
            self.parent,
            "Export to CSV",
            default_filename,
            "CSV Files (*.csv)"
        )

        if not filename:
            return False

        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                columns = self._effective_columns(report_data)

                # Header row
                headers = [col.display_name for col in columns]
                writer.writerow(headers)

                # Data rows
                for row in report_data.rows:
                    csv_row = []
                    for col in columns:
                        value = self._value_for_column(row, col.name)
                        formatted = col.format_value(value)
                        csv_row.append(formatted)
                    writer.writerow(csv_row)

                # Summary totals
                if report_data.summary:
                    writer.writerow([])  # Blank separator
                    for key, value in report_data.summary.items():
                        if key in ('total_rows', 'row_count', 'generated_at'):
                            continue
                        if isinstance(value, dict):
                            continue
                        display_key = key.replace('_', ' ').title()
                        if isinstance(value, float):
                            if 'margin' in key.lower() or 'percent' in key.lower():
                                writer.writerow([display_key, f"{value:.1f}%"])
                            else:
                                writer.writerow([display_key, f"${value:,.2f}"])
                        else:
                            writer.writerow([display_key, value])

            QMessageBox.information(
                self.parent,
                "Export Successful",
                f"Report exported to:\n{filename}"
            )
            return True

        except Exception as e:
            QMessageBox.critical(
                self.parent,
                "Export Failed",
                f"Failed to export CSV:\n{str(e)}"
            )
            return False

    # ========================================================================
    # Excel Export
    # ========================================================================

    def export_excel(
        self,
        report_data,
        default_filename: Optional[str] = None,
        include_summary: bool = True
    ) -> bool:
        """
        Export report to Excel with formatting.

        Args:
            report_data: ReportData object
            default_filename: Suggested filename
            include_summary: Include summary sheet

        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(
                self.parent,
                "Module Not Found",
                "openpyxl module not installed.\n\n"
                "Install it with:\npip install openpyxl --break-system-packages"
            )
            return False

        if not default_filename:
            safe_title = report_data.title.replace(' ', '_').replace('/', '_')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"{safe_title}_{timestamp}.xlsx"

        filename, _ = QFileDialog.getSaveFileName(
            self.parent,
            "Export to Excel",
            default_filename,
            "Excel Files (*.xlsx)"
        )

        if not filename:
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report_data.title[:31]  # Excel limit
            columns = self._effective_columns(report_data)

            # Header styling
            header_fill = PatternFill(
                start_color="1976D2",
                end_color="1976D2",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Header row
            headers = [col.display_name for col in columns]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Data rows
            for row in report_data.rows:
                excel_row = []
                for col in columns:
                    value = self._value_for_column(row, col.name)

                    # Store raw value for Excel
                    if col.data_type in ('currency', 'number', 'percent'):
                        try:
                            excel_row.append(float(value))
                        except (ValueError, TypeError):
                            excel_row.append(self._excel_safe_value(value))
                    else:
                        excel_row.append(self._excel_safe_value(value))

                ws.append(excel_row)

                # Apply row styling if present
                if row.style:
                    row_num = ws.max_row
                    bg_color = row.style.get('background_color')
                    if bg_color:
                        fill = PatternFill(
                            start_color=bg_color.replace('#', ''),
                            end_color=bg_color.replace('#', ''),
                            fill_type="solid"
                        )
                        for cell in ws[row_num]:
                            cell.fill = fill

            # Format columns
            for idx, col in enumerate(columns, start=1):
                column_letter = openpyxl.utils.get_column_letter(idx)

                # Set alignment
                alignment = Alignment(
                    horizontal=col.alignment,
                    vertical='center'
                )
                for cell in ws[column_letter][1:]:  # Skip header
                    cell.alignment = alignment

                # Set number format
                if col.data_type == 'currency':
                    for cell in ws[column_letter][1:]:
                        cell.number_format = '$#,##0.00'
                elif col.data_type == 'percent':
                    for cell in ws[column_letter][1:]:
                        cell.number_format = '0.0%'
                elif col.data_type == 'number':
                    for cell in ws[column_letter][1:]:
                        cell.number_format = '#,##0.0'
                elif col.data_type == 'date':
                    for cell in ws[column_letter][1:]:
                        cell.number_format = 'mm/dd/yyyy'

                # Auto-size column
                max_length = len(col.display_name)
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet if requested
            if include_summary and report_data.summary:
                summary_ws = wb.create_sheet("Summary")
                summary_ws.append(["Report Summary"])
                summary_ws['A1'].font = Font(bold=True, size=14)
                summary_ws.append([])

                for key, value in report_data.summary.items():
                    summary_ws.append([key.replace('_', ' ').title(), self._excel_safe_value(value)])

            # Freeze header row
            ws.freeze_panes = 'A2'

            wb.save(filename)

            QMessageBox.information(
                self.parent,
                "Export Successful",
                f"Report exported to:\n{filename}"
            )
            return True

        except Exception as e:
            QMessageBox.critical(
                self.parent,
                "Export Failed",
                f"Failed to export Excel:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
            return False

    # ========================================================================
    # PDF Export
    # ========================================================================

    def export_pdf(
        self,
        report_data,
        default_filename: Optional[str] = None,
        include_summary: bool = True
    ) -> bool:
        """
        Export report to PDF with professional formatting.

        Args:
            report_data: ReportData object
            default_filename: Suggested filename
            include_summary: Include summary section

        Returns:
            True if successful, False otherwise
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            QMessageBox.warning(
                self.parent,
                "Module Not Found",
                "reportlab module not installed.\n\n"
                "Install it with:\npip install reportlab --break-system-packages"
            )
            return False

        if not default_filename:
            safe_title = report_data.title.replace(' ', '_').replace('/', '_')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"{safe_title}_{timestamp}.pdf"

        filename, _ = QFileDialog.getSaveFileName(
            self.parent,
            "Export to PDF",
            default_filename,
            "PDF Files (*.pdf)"
        )

        if not filename:
            return False

        try:
            # Determine page orientation
            num_columns = len(report_data.columns)
            pagesize = landscape(letter) if num_columns > 6 else letter

            doc = SimpleDocTemplate(
                filename,
                pagesize=pagesize,
                leftMargin=0.5 * inch,
                rightMargin=0.5 * inch,
                topMargin=0.75 * inch,
                bottomMargin=0.5 * inch
            )

            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1976d2'),
                spaceAfter=12
            )
            title = Paragraph(report_data.title, title_style)
            elements.append(title)

            # Summary section
            if include_summary and report_data.summary:
                summary_lines = []
                generated_at = report_data.summary.get('generated_at', 'N/A')
                if report_data.metadata and report_data.metadata.get('generated_at'):
                    gen_time = report_data.metadata['generated_at']
                    if hasattr(gen_time, 'strftime'):
                        generated_at = gen_time.strftime('%m/%d/%Y %I:%M %p')
                summary_lines.append(f"Generated: {generated_at}")
                if 'total_rows' in report_data.summary:
                    summary_lines.append(f"Total Rows: {report_data.summary['total_rows']}")

                summary_para = Paragraph(" | ".join(summary_lines), styles['Normal'])
                elements.append(summary_para)
                elements.append(Spacer(1, 0.2 * inch))

            # Build table data
            table_data = []
            columns = self._effective_columns(report_data)

            # Headers
            headers = [col.display_name for col in columns]
            table_data.append(headers)

            # Data rows
            for row in report_data.rows:
                pdf_row = []
                for col in columns:
                    value = self._value_for_column(row, col.name)
                    formatted = col.format_value(value)
                    pdf_row.append(formatted)
                table_data.append(pdf_row)

            # Create table
            table = Table(table_data, repeatRows=1)

            # Table styling
            table_style = TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),

                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
            ])

            # Apply column-specific alignment
            for idx, col in enumerate(columns):
                if col.alignment == 'right':
                    table_style.add('ALIGN', (idx, 1), (idx, -1), 'RIGHT')
                elif col.alignment == 'center':
                    table_style.add('ALIGN', (idx, 1), (idx, -1), 'CENTER')

            table.setStyle(table_style)
            elements.append(table)

            # Summary totals section after table
            if include_summary and report_data.summary:
                elements.append(Spacer(1, 0.25 * inch))

                summary_style = ParagraphStyle(
                    'SummaryStyle',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#333333'),
                    spaceAfter=4,
                    fontName='Helvetica-Bold'
                )

                totals_parts = []
                for key, value in report_data.summary.items():
                    if key in ('total_rows', 'row_count', 'generated_at'):
                        continue
                    if isinstance(value, dict):
                        continue
                    display_key = key.replace('_', ' ').title()
                    if isinstance(value, float):
                        if 'margin' in key.lower() or 'percent' in key.lower():
                            totals_parts.append(f"{display_key}: {value:.1f}%")
                        else:
                            totals_parts.append(f"{display_key}: ${value:,.2f}")
                    else:
                        totals_parts.append(f"{display_key}: {value}")

                if totals_parts:
                    # Build a small summary table for clean formatting
                    summary_table_data = []
                    summary_table_data.append(['REPORT TOTALS', '', ''])
                    for i in range(0, len(totals_parts), 3):
                        row_items = totals_parts[i:i + 3]
                        while len(row_items) < 3:
                            row_items.append('')
                        summary_table_data.append(row_items)

                    summary_table = Table(summary_table_data)
                    summary_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('SPAN', (0, 0), (-1, 0)),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('TOPPADDING', (0, 1), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e8f4f8')),
                        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1976d2')),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                    ]))
                    elements.append(summary_table)

            # Build PDF
            doc.build(elements)

            QMessageBox.information(
                self.parent,
                "Export Successful",
                f"Report exported to:\n{filename}"
            )
            return True

        except Exception as e:
            QMessageBox.critical(
                self.parent,
                "Export Failed",
                f"Failed to export PDF:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
            return False

    # ========================================================================
    # Convenience methods
    # ========================================================================

    def export_all_formats(
        self,
        report_data,
        base_filename: Optional[str] = None
    ) -> Dict[str, bool]:
        """
        Export report to all formats.

        Args:
            report_data: ReportData object
            base_filename: Base filename (without extension)

        Returns:
            Dictionary mapping format to success status
        """
        results = {}

        if not base_filename:
            safe_title = report_data.title.replace(' ', '_').replace('/', '_')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = f"{safe_title}_{timestamp}"

        results['csv'] = self.export_csv(
            report_data,
            f"{base_filename}.csv"
        )

        results['excel'] = self.export_excel(
            report_data,
            f"{base_filename}.xlsx"
        )

        results['pdf'] = self.export_pdf(
            report_data,
            f"{base_filename}.pdf"
        )

        return results
